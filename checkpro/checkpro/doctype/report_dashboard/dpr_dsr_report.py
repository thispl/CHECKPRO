import frappe
import pandas as pd
from frappe.utils import get_site_path, today
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.styles import Font, Alignment, Border, Side

@frappe.whitelist()
def download_excel(date):
    file_name = "Production Task.xlsx"
    file_path = get_site_path("private", "files", secure_filename(file_name))
   
    # Get dataframes
    df1 = production(date)
    df2 = in_sprint_not_taken(date)
    df3 = not_in_sprint(date)

    # Write to Excel with multiple sheets
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df1.to_excel(writer, sheet_name="Production", index=False)
        df2.to_excel(writer, sheet_name="In Sprint Not Taken", index=False)
        df3.to_excel(writer, sheet_name="Not in Sprint", index=False)
        writer._save()
    
    # Create Summary sheet
    summary(date, file_path)

    # Borders for all cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    workbook = load_workbook(file_path)

    # Define sheet-specific column widths
    sheet_column_widths = {
        "Production": {
            "A":3,"B": 9, "C": 8, "D": 6, "E": 30, "F": 9, "G": 62, "H": 9, "I": 4, "J":4,"K":5, "L":5, "M":13
        },
        "In Sprint Not Taken": {
            "A":3,"B": 9, "C": 8, "D": 6, "E": 30, "F": 9, "G": 62, "H": 9,"I": 4, "J":4,"K":5, "L":10,"M":11
        },
        "Not in Sprint": {
             "A":3,"B":9, "C": 30, "D": 62, "E": 9, "F": 4, "G": 4, "H": 5, "I":11,"J":12
        },
    }

    fill_color1 = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")  # light yellow
    fill_color2 = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")  # white

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]

        if sheet_name != "Summary":
            # Header formatting
            header_fill = PatternFill(start_color="FF0070C0", end_color="FF0070C0", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            center_align = Alignment(horizontal="center", vertical="center")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = center_align

            # Apply borders and alternating row colors
            for i, row in enumerate(ws.iter_rows(min_row=2), start=2):  # start from row 2 (first data row)
                row_fill = fill_color1 if i % 2 == 0 else fill_color2
                for cell in row:
                    cell.border = thin_border
                    cell.fill = row_fill
                    cell.alignment = center_align

        else:
            # Summary sheet: bold font for header
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Apply sheet-specific column widths
        col_widths = sheet_column_widths.get(sheet_name, {})
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        subject_columns = {
        "Production": "G",           
        "In Sprint Not Taken": "G",  
        "Not in Sprint": "D"         
        }

        for sheet_name, col_letter in subject_columns.items():
            ws = workbook[sheet_name]
            wrap_alignment = Alignment(wrap_text=True, vertical="top")
            for row in ws.iter_rows(min_row=2):  
                cell = row[ord(col_letter.upper()) - 65]  
                cell.alignment = wrap_alignment

        if "Summary" in workbook.sheetnames:
            summary_sheet = workbook["Summary"]
            workbook._sheets.remove(summary_sheet)
            workbook._sheets.insert(0, summary_sheet)
            workbook.active = 0  # make Summary the active sheet

    workbook.save(file_path)
    workbook.close()

    with open(file_path, "rb") as f:
        file_content = f.read()

    return {
        "filename": file_name,
        "content": file_content
    }



def production(date):
    columns = [
        "S.NO", "Sprint","Team","CB","Project","Task","Subject",
        "Priority","KT","ET","AT","TRT","Status"
    ]

    data = []

    daily_monitors = frappe.get_all(
        "Daily Monitor",
        filters={"custom_dm_production_date": date},
        fields=["name","sprint","dev_team"],
    )

    for dm in daily_monitors:
        doc = frappe.get_doc("Daily Monitor", dm.name)
        for row in doc.task_details:
            kt = frappe.db.get_value("Task", {"name": row.id}, "kt_confirmed")
            data.append([
                dm.sprint,
                dm.dev_team,
                row.cb,
                row.project_name,
                row.id,
                row.subject,
                row.priority,
                kt,
                row.et,
                round(row.at, 2),
                row.today_rt,
                row.current_status
            ])

    data = sorted(data, key=lambda x: (x[1] or "" , x[0] or "", x[2] or "", x[3] or "", x[6] or ""))

    data_with_sn = []
    for idx, row in enumerate(data, start=1):
        data_with_sn.append([idx] + row)  
    df = pd.DataFrame(data_with_sn, columns=columns)
    return df


def in_sprint_not_taken(date):
    columns = [
        "S.NO", "Sprint","Team","CB","Project","Task","Subject",
        "Priority","KT","ET","RT","Status","Creation"
    ]

    data = []

    tasks = frappe.get_all(
        "Task",
        filters={
            "custom_sprint": ["is", "set"],
            "custom_production_date": ["is", "not set"],
            "status": ["in", ["Open","Working"]]
        },
        fields=[
            "name",
            "custom_sprint",
            "custom_dev_team",
            "cb",
            "project",
            "subject",
            "priority",
            "kt_confirmed",
            "expected_time",
            "rt",
            "status",
            "creation"
        ],
        order_by="custom_sprint asc, custom_dev_team asc"
    )
    # tasks = sorted(tasks, key=lambda x: (x.get("custom_dev_team"), x.get("custom_sprint"), x.get("cb"), x.get("project"), x.get("priority")))
    tasks = sorted(
            tasks,
            key=lambda x: (
                x.get("custom_dev_team") or "",
                x.get("custom_sprint") or "",
                x.get("cb") or "",
                x.get("project") or "",
                x.get("priority") or ""
            )
        )


    for idx, i in enumerate(tasks, start=1):
        creation_date = (i.get("creation").strftime("%d-%m-%Y") if i.get("creation") else None)
        data.append([
            idx,  
            i.get("custom_sprint"),
            i.get("custom_dev_team"),
            i.get("cb"),
            i.get("project"),
            i.get("name"),
            i.get("subject"),
            i.get("priority"),
            i.get("kt_confirmed"),
            i.get("expected_time"),
            i.get("rt"),
            i.get("status"),
            creation_date
        ])

    df = pd.DataFrame(data, columns=columns)
    return df



def not_in_sprint(date):
    columns = [
        "S.NO","Task","Project Name","Subject",
        "Priority","KT","ET","RT","Status","Creation"
    ]

    data = []

    tasks = frappe.db.get_all(
        "Task",
        filters={
            "custom_sprint": ["is", "not set"],
            "status": ["not in", ["Completed","Cancelled","Client Review","Pending Review","Code Review","Hold"]],
            "service":"IT-SW"
        },
        fields=[
            "name",
            "custom_sprint",
            "custom_dev_team",
            "cb",
            "project",
            "subject",
            "priority",
            "kt_confirmed",
            "expected_time",
            "rt",
            "status",
            "creation"
        ]
    )
    tasks = sorted(tasks, key=lambda x: (x.get("project") or "", x.get("priority") or ""))

    for idx, i in enumerate(tasks, start=1):
        creation_date = (i.get("creation").strftime("%d-%m-%Y") if i.get("creation") else None)
        data.append([
            idx,  
            i.get("name"),
            i.get("project"),
            i.get("subject"),
            i.get("priority"),
            i.get("kt_confirmed"),
            i.get("expected_time"),
            i.get("rt"),
            i.get("status"),
            creation_date
        ])

    df = pd.DataFrame(data, columns=columns)
    return df



# def summary(date, file_path):
#     from collections import defaultdict
#     from openpyxl import load_workbook
#     from openpyxl.styles import Font, Alignment

#     production_data = production(date).values.tolist()
#     in_sprint_data = in_sprint_not_taken(date).values.tolist()
#     not_in_sprint_data = not_in_sprint(date).values.tolist()

#     wb = load_workbook(file_path)
#     ws = wb.create_sheet("Summary")

#     bold_font = Font(bold=True)
#     thin_border_1 = Border(
#         top=Side(style='thin'),
#         bottom=Side(style='thin')
#     )
#     align_center = Alignment(horizontal="center", vertical="center")
#     align_left = Alignment(horizontal="left", vertical="center")

#     # --- Header ---
#     ws.merge_cells("A1:B1")
#     ws.merge_cells("D1:E1")
#     ws.merge_cells("G1:H1")
#     ws["A1"]= "Production"
#     ws["D1"]= "In Sprint Not Taken"
#     ws["G1"]= "Not in Sprint (Inventory)"
#     for cell in ["A1","B1","D1","E1","G1","H1"]:
#         ws[cell].font = bold_font
#         ws[cell].alignment = align_center

#     ws["A2"], ws["B2"] = "Row Labels", "Sum of TRT"
#     ws["D2"], ws["E2"] = "Row Labels", "Sum of ET"
#     ws["G2"], ws["H2"] = "Row Labels", "ET (hrs)"
#     for cell in ["A2","B2","D2","E2","G2","H2"]:
#         ws[cell].font = bold_font
#         ws[cell].alignment = align_center

#     prod_dict = defaultdict(lambda: defaultdict(float)) 
#     for row in production_data:
#         team, cb, trt = row[2], row[3], row[11] or 0
#         prod_dict[team][cb] += trt

#     sprint_dict = defaultdict(lambda: defaultdict(float))  
#     for row in in_sprint_data:
#         team, sprint, et = row[2], row[1], row[9] or 0
#         sprint_dict[team][sprint] += et

#     inv_dict = defaultdict(float)  
#     for row in not_in_sprint_data:
#         project, et = row[2], row[6] or 0
#         inv_dict[project] += et

#     thin_border = Border(
#     top=Side(style='thin'),
#     )
#     row_num = 3
#     for team, cbs in prod_dict.items():
#         team_total = sum(cbs.values())

#         ws.cell(row=row_num, column=1, value=team).font = bold_font
#         ws.cell(row=row_num, column=1).border = thin_border
#         ws.cell(row=row_num, column=2, value=team_total).font = bold_font
#         ws.cell(row=row_num, column=2).alignment = align_center
#         ws.cell(row=row_num, column=2).border = thin_border

#         row_num += 1
#         for cb, trt in cbs.items():
#             ws.cell(row=row_num, column=1, value=cb).alignment = align_left
#             ws.cell(row=row_num, column=2, value=trt).alignment = align_center
#             row_num += 1

#     # --- Production Grand Total immediately after Production section ---
#     ws.cell(row=row_num, column=1, value="Grand Total").font = bold_font
#     ws.cell(row=row_num, column=1).border = thin_border
#     ws.cell(row=row_num, column=2, value=sum([trt for cbs in prod_dict.values() for trt in cbs.values()])).alignment = align_center
#     ws.cell(row=row_num, column=2).border = thin_border
#     row_num += 2  # leave a blank row before next section

#     # --- Write In Sprint Not Taken with borders for team name & total only ---
#     row_num = 3
#     for team, sprints in sprint_dict.items():
#         team_total = sum(sprints.values())

#         ws.cell(row=row_num, column=4, value=team).font = bold_font
#         ws.cell(row=row_num, column=4).border = thin_border
#         ws.cell(row=row_num, column=5, value=team_total).font = bold_font
#         ws.cell(row=row_num, column=5).alignment = align_center
#         ws.cell(row=row_num, column=5).border = thin_border

#         row_num += 1
#         for sprint, rt in sprints.items():
#             ws.cell(row=row_num, column=4, value=sprint).alignment = align_left
#             ws.cell(row=row_num, column=5, value=rt).alignment = align_center
#             row_num += 1

#     # --- In Sprint Not Taken Grand Total immediately after section ---
#     ws.cell(row=row_num, column=4, value="Grand Total").font = bold_font
#     ws.cell(row=row_num, column=4).border = thin_border
#     ws.cell(row=row_num, column=5, value=sum([rt for sprints in sprint_dict.values() for rt in sprints.values()])).alignment = align_center
#     ws.cell(row=row_num, column=5).border = thin_border
#     row_num += 2  # leave a blank row before next section

#     # --- Write Not in Sprint ---
#     row_num = 3
#     for project, rt in inv_dict.items():
#         ws.cell(row=row_num, column=7, value=project).alignment = align_left
#         ws.cell(row=row_num, column=8, value=rt).alignment = align_center
#         row_num += 1

#     # --- Not in Sprint Grand Total immediately after section ---
#     ws.cell(row=row_num, column=7, value="Grand Total").font = bold_font
#     ws.cell(row=row_num, column=7).border = thin_border
#     ws.cell(row=row_num, column=8, value=sum(inv_dict.values())).alignment = align_center
#     ws.cell(row=row_num, column=8).border = thin_border


#     # --- Column widths ---
#     for col in ["A","B","D","E","G","H"]:
#         ws.column_dimensions[col].width = 25
#     for col in ["C","F"]:  # Empty spacing columns
#         ws.column_dimensions[col].width = 3

#     wb.save(file_path)


def summary(date, file_path):
    from collections import defaultdict
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    # --- Load data ---
    production_data = production(date).values.tolist()
    in_sprint_data = in_sprint_not_taken(date).values.tolist()
    not_in_sprint_data = not_in_sprint(date).values.tolist()

    wb = load_workbook(file_path)
    ws = wb.create_sheet("Summary")

    # --- Styles ---
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    # Header colors
    header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    # --- Header Titles ---
    ws.merge_cells("A1:B1")
    ws.merge_cells("D1:E1")
    ws.merge_cells("G1:H1")

    ws["A1"] = "Production"
    ws["D1"] = "In Sprint Not Taken"
    ws["G1"] = "Not in Sprint (Inventory)"

    for cell in ["A1", "B1", "D1", "E1", "G1", "H1"]:
        ws[cell].font = bold_font
        ws[cell].alignment = align_center
        ws[cell].fill = header_fill
        ws[cell].border = thin_border

    # --- Subheaders ---
    ws["A2"], ws["B2"] = "TEAM & CB", "TRT(H)"
    ws["D2"], ws["E2"] = "TEAM & SPRINT", "ET(H)"
    ws["G2"], ws["H2"] = "Project", "ET(H)"

    for cell in ["A2", "B2", "D2", "E2", "G2", "H2"]:
        ws[cell].font = bold_font
        ws[cell].alignment = align_center
        ws[cell].border = thin_border
        ws[cell].fill = header_fill

    # === Production Summary ===
    prod_dict = defaultdict(lambda: defaultdict(float))
    for row in production_data:
        team, cb, trt = row[2], row[3], row[11] or 0
        prod_dict[team][cb] += trt

    row_num = 3
    for team, cbs in prod_dict.items():
        team_total = sum(cbs.values())
        ws.cell(row=row_num, column=1, value=team).font = bold_font
        ws.cell(row=row_num, column=1).alignment = align_left
        ws.cell(row=row_num, column=1).border = thin_border
        ws.cell(row=row_num, column=2, value=round(team_total, 2)).alignment = align_center
        ws.cell(row=row_num, column=2).font = bold_font
        ws.cell(row=row_num, column=2).border = thin_border
        row_num += 1

        for cb, trt in cbs.items():
            ws.cell(row=row_num, column=1, value=cb).alignment = align_left
            ws.cell(row=row_num, column=1).border = thin_border
            ws.cell(row=row_num, column=2, value=round(trt, 2)).alignment = align_center
            ws.cell(row=row_num, column=2).border = thin_border
            row_num += 1

    # --- Production Grand Total ---
    ws.cell(row=row_num, column=1, value="Grand Total").font = bold_font
    ws.cell(row=row_num, column=1).alignment = align_left
    ws.cell(row=row_num, column=1).border = thin_border
    ws.cell(row=row_num, column=1).fill = header_fill
    ws.cell(row=row_num, column=2, value=round(sum([trt for cbs in prod_dict.values() for trt in cbs.values()]), 2)).alignment = align_center
    ws.cell(row=row_num, column=2).border = thin_border
    ws.cell(row=row_num, column=2).font = bold_font
    ws.cell(row=row_num, column=2).fill = header_fill
    row_num += 2

    # === In Sprint Not Taken Summary ===
    sprint_dict = defaultdict(lambda: defaultdict(float))
    for row in in_sprint_data:
        team, sprint, et = row[2], row[1], row[9] or 0
        sprint_dict[team][sprint] += et

    row_num2 = 3
    for team, sprints in sprint_dict.items():
        team_total = sum(sprints.values())
        ws.cell(row=row_num2, column=4, value=team).font = bold_font
        ws.cell(row=row_num2, column=4).alignment = align_left
        ws.cell(row=row_num2, column=4).border = thin_border
        ws.cell(row=row_num2, column=5, value=round(team_total, 2)).alignment = align_center
        ws.cell(row=row_num2, column=5).font = bold_font
        ws.cell(row=row_num2, column=5).border = thin_border
        row_num2 += 1

        for sprint, et in sprints.items():
            ws.cell(row=row_num2, column=4, value=sprint).alignment = align_left
            ws.cell(row=row_num2, column=4).border = thin_border
            ws.cell(row=row_num2, column=5, value=round(et, 2)).alignment = align_center
            ws.cell(row=row_num2, column=5).border = thin_border
            row_num2 += 1

    ws.cell(row=row_num2, column=4, value="Grand Total").font = bold_font
    ws.cell(row=row_num2, column=4).alignment = align_left
    ws.cell(row=row_num2, column=4).border = thin_border
    ws.cell(row=row_num2, column=4).fill = header_fill
    ws.cell(row=row_num2, column=5, value=round(sum([et for sprints in sprint_dict.values() for et in sprints.values()]), 2)).alignment = align_center
    ws.cell(row=row_num2, column=5).border = thin_border
    ws.cell(row=row_num2, column=5).font = bold_font
    ws.cell(row=row_num2, column=5).fill = header_fill
    row_num2 += 2

    # === Not in Sprint (Inventory) Summary ===
    inv_dict = defaultdict(float)
    for row in not_in_sprint_data:
        project, et = row[2], row[6] or 0
        inv_dict[project] += et

    row_num3 = 3
    for project, et in inv_dict.items():
        ws.cell(row=row_num3, column=7, value=project).alignment = align_left
        ws.cell(row=row_num3, column=7).border = thin_border
        ws.cell(row=row_num3, column=8, value=round(et, 2)).alignment = align_center
        ws.cell(row=row_num3, column=8).border = thin_border
        row_num3 += 1

    ws.cell(row=row_num3, column=7, value="Grand Total").font = bold_font
    ws.cell(row=row_num3, column=7).alignment = align_left
    ws.cell(row=row_num3, column=7).border = thin_border
    ws.cell(row=row_num3, column=7).fill = header_fill
    ws.cell(row=row_num3, column=8, value=round(sum(inv_dict.values()), 2)).alignment = align_center
    ws.cell(row=row_num3, column=8).border = thin_border
    ws.cell(row=row_num3, column=8).font = bold_font
    ws.cell(row=row_num3, column=8).fill = header_fill

    # --- Column Widths ---
    for col, width in {"A": 13, "B": 8, "D": 13, "E": 8, "G": 40, "H": 8}.items():
        ws.column_dimensions[col].width = width
    for col in ["C", "F"]:
        ws.column_dimensions[col].width = 3  # spacing columns

    wb.save(file_path)




