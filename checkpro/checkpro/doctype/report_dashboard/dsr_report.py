import frappe
import pandas as pd
from frappe.utils import get_site_path, today
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.styles import Font, Alignment, Border, Side

# @frappe.whitelist()
# def download_excel(date):
# 	file_name = "Production Task.xlsx"
# 	file_path = get_site_path("private", "files", secure_filename(file_name))
   
# 	# Get dataframes
# 	df1 = production(date)
# 	df2 = in_sprint_not_taken(date)
# 	df3 = not_in_sprint(date)

# 	# Write to Excel with multiple sheets
# 	with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
# 		df1.to_excel(writer, sheet_name="DSR", index=False)
# 		df2.to_excel(writer, sheet_name="In Sprint Not Taken", index=False)
# 		df3.to_excel(writer, sheet_name="Not in Sprint", index=False)
# 		writer._save()
	
# 	# Create Summary sheet
# 	summary(date,file_path)

# 	# Borders for all cells
# 	thin_border = Border(
# 		left=Side(style='thin'),
# 		right=Side(style='thin'),
# 		top=Side(style='thin'),
# 		bottom=Side(style='thin')
# 	)

# 	workbook = load_workbook(file_path)

# 	# Define sheet-specific column widths
# 	sheet_column_widths = {
# 		"DSR": {
# 			"A":3,"B": 9, "C": 8, "D": 6, "E": 30, "F": 9, "G": 62, "H": 9, "I": 4, "J":4,"K":5, "L":5, "M":5, "N":13, "O":29, "P":35
# 		},
# 		"In Sprint Not Taken": {
# 			"A":3,"B": 9, "C": 8, "D": 6, "E": 30, "F": 9, "G": 62, "H": 9,"I": 4, "J":4,"K":5, "L":13,"M":13
# 		},
# 		"Not in Sprint": {
# 			 "A":3,"B":9, "C": 30, "D": 62, "E": 9, "F": 4, "G": 4, "H": 5, "I":12,"J":12
# 		},
# 	}

# 	fill_color1 = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")  # light yellow
# 	fill_color2 = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")  # white

# 	for sheet_name in workbook.sheetnames:
# 		ws = workbook[sheet_name]

# 		if sheet_name != "Summary":
# 			# Header formatting
# 			header_fill = PatternFill(start_color="FF0070C0", end_color="FF0070C0", fill_type="solid")
# 			header_font = Font(color="FFFFFF", bold=True)
# 			for cell in ws[1]:
# 				cell.fill = header_fill
# 				cell.font = header_font
# 				cell.border = thin_border

# 			# Apply borders and alternating row colors
# 			for i, row in enumerate(ws.iter_rows(min_row=2), start=2):  # start from row 2 (first data row)
# 				row_fill = fill_color1 if i % 2 == 0 else fill_color2
# 				for cell in row:
# 					cell.border = thin_border
# 					cell.fill = row_fill

# 		else:
# 			# Summary sheet: bold font for header
# 			for cell in ws[1]:
# 				cell.font = Font(bold=True)
# 				cell.alignment = Alignment(horizontal="center", vertical="center")

# 		# Apply sheet-specific column widths
# 		col_widths = sheet_column_widths.get(sheet_name, {})
# 		for col, width in col_widths.items():
# 			ws.column_dimensions[col].width = width

# 		subject_columns = {
# 			"DSR": ["G", "P"],        
# 			"In Sprint Not Taken": ["G"],  
# 			"Not in Sprint": ["D"]         
# 		}

# 		for sheet_name, col_letters in subject_columns.items():
# 			ws = workbook[sheet_name]
# 			wrap_alignment = Alignment(wrap_text=True, vertical="top")
			
# 			# Ensure col_letters is always a list
# 			if not isinstance(col_letters, list):
# 				col_letters = [col_letters]
			
# 			for col_letter in col_letters:
# 				for row in ws.iter_rows(min_row=2):
# 					cell = row[ord(col_letter.upper()) - 65]  # works now
# 					cell.alignment = wrap_alignment


# 	workbook.save(file_path)
# 	workbook.close()

# 	with open(file_path, "rb") as f:
# 		file_content = f.read()

# 	return {
# 		"filename": file_name,
# 		"content": file_content
# 	}


@frappe.whitelist()
def download_excel(date):
	file_name = "Production Task.xlsx"
	file_path = get_site_path("private", "files", secure_filename(file_name))
   
	# Get dataframes
	df1 = combined_production_report(date)
	# df4 = production_1(date)
	df2 = in_sprint_not_taken(date)
	df3 = not_in_sprint(date)

	# Write to Excel with multiple sheets
	with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
		df1.to_excel(writer, sheet_name="DSR", index=False)
		# df4.to_excel(writer, sheet_name="CDR Tasks", index=False)
		df2.to_excel(writer, sheet_name="In Sprint Not Taken", index=False)
		df3.to_excel(writer, sheet_name="Not in Sprint", index=False)
		writer._save()
	
	# Create Summary sheet
	summary(date, file_path)

	# Borders for all cells
	thin_border = Border(
		left=Side(style='thin'),
		right=Side(style='thin'),
		top=Side(style='thin'),
		bottom=Side(style='thin')
	)

	workbook = load_workbook(file_path)

	# Define sheet-specific column widths
	sheet_column_widths = {
		"DSR": {
			"A":3,"B":9,"C":8,"D":6,"E":30,"F":9,"G":62,"H":9,"I":4,"J":4,"K":5,"L":5,"M":5,"N":13,"O":31,"P":35
		},
		# "CDR Tasks": {
		# 	"A":3,"B":9,"C":8,"D":6,"E":30,"F":9,"G":62,"H":9,"I":4,"J":4,"K":5,"L":5,"M":5,"N":13,"O":29,"P":35
		# },
		"In Sprint Not Taken": {
			"A":3,"B":9,"C":8,"D":6,"E":30,"F":9,"G":62,"H":9,"I":4,"J":4,"K":5,"L":13,"M":13
		},
		"Not in Sprint": {
			"A":3,"B":9,"C":30,"D":62,"E":9,"F":4,"G":4,"H":5,"I":12,"J":12
		},
	}

	fill_color1 = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")  # light color
	fill_color2 = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")  # white

	for sheet_name in workbook.sheetnames:
		ws = workbook[sheet_name]

		if sheet_name != "Summary":
			# Header formatting
			header_fill = PatternFill(start_color="FF0070C0", end_color="FF0070C0", fill_type="solid")
			header_font = Font(color="FFFFFF", bold=True)
			center_align = Alignment(horizontal="center", vertical="center")
			for cell in ws[1]:
				cell.fill = header_fill
				cell.font = header_font
				cell.border = thin_border
				cell.alignment=center_align

			for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
				first_cell = str(row[0].value).strip().lower() if row[0].value else ""
				if first_cell == "s.no":
					for cell in row:
						cell.fill = header_fill
						cell.font = header_font
						cell.border = thin_border
						cell.alignment = center_align

				

			# Apply borders and alternating row colors
			for i, row in enumerate(ws.iter_rows(min_row=2), start=2):  # start from row 2 (data rows)
				first_cell = str(row[0].value).strip().lower() if row[0].value else ""
				if first_cell == "s.no":
					continue  

				if all(not (cell.value and str(cell.value).strip()) for cell in row):
					continue

				row_fill = fill_color1 if i % 2 == 0 else fill_color2
				for cell in row:
					cell.border = thin_border
					cell.fill = row_fill
					cell.alignment = center_align


		else:
			# Summary sheet: bold header and centered
			for cell in ws[1]:
				cell.font = Font(bold=True)
				cell.alignment = Alignment(horizontal="center", vertical="center")

		# Apply sheet-specific column widths
		col_widths = sheet_column_widths.get(sheet_name, {})
		for col, width in col_widths.items():
			ws.column_dimensions[col].width = width

	# Wrap text for subject columns
	subject_columns = {
		"DSR": ["G", "P"], 
		# "CDR Tasks": ["G", "P"],        
		"In Sprint Not Taken": ["G"],  
		"Not in Sprint": ["D"]         
	}

	for sheet_name, col_letters in subject_columns.items():
		ws = workbook[sheet_name]
		wrap_alignment = Alignment(wrap_text=True, vertical="top")
		
		if not isinstance(col_letters, list):
			col_letters = [col_letters]
		
		for col_letter in col_letters:
			for row in ws.iter_rows(min_row=2):
				cell = row[ord(col_letter.upper()) - 65]
				cell.alignment = wrap_alignment


	if "Summary" in workbook.sheetnames:
		summary_sheet = workbook["Summary"]
		workbook._sheets.remove(summary_sheet)
		workbook._sheets.insert(0, summary_sheet)
		workbook.active = 0  

	workbook.save(file_path)
	workbook.close()

	with open(file_path, "rb") as f:
		file_content = f.read()

	return {
		"filename": file_name,
		"content": file_content
	}


import pandas as pd

def combined_production_report(date):
	# Get the two dataframes
	df1 = production(date)
	df2 = production_1(date)

	blank_row = pd.DataFrame([[""] * len(df1.columns)], columns=df1.columns)

	header_row = pd.DataFrame([df1.columns], columns=df1.columns)
	combined_df = pd.concat([df1, blank_row, header_row, df2], ignore_index=True)

	return combined_df




def production(date):
	columns = [
		"S.NO", "Sprint","Team","CB","Project","Task","Subject",
		"Priority","KT","ET","AT","TRT","ATP","Status","Allocated to", "Remarks"
	]

	data = []

	daily_monitors = frappe.get_all(
		"Daily Monitor",
		filters={"custom_dm_production_date": date},
		fields=["name","sprint","dev_team"],
	)

	for dm in daily_monitors:
		doc = frappe.get_doc("Daily Monitor", dm.name)
		for row in doc.task_details:
			kt = frappe.db.get_value("Task", {"name": row.id}, "kt_confirmed")
			allocated_to = frappe.db.get_value("Task",{'name':row.id},'custom_allocated_to')
			if row.cb == row.cb_against_task:
				data.append([
					dm.sprint,
					dm.dev_team,
					row.cb,
					row.project_name,
					row.id,
					row.subject,
					row.priority,
					kt,
					row.et,
					round(row.at, 2),
					row.today_rt,
					float(row.at_taken or 0),
					row.current_status,
					allocated_to,
					row.remark
				])

	data = sorted(data, key=lambda x: (x[1] or "", x[0] or "", x[2] or "", x[3] or "", x[6] or ""))

	data_with_sn = []
	for idx, row in enumerate(data, start=1):
		data_with_sn.append([idx] + row)  
	df = pd.DataFrame(data_with_sn, columns=columns)
	return df


def production_1(date):
	columns = [
		"S.NO", "Sprint","Team","CB","Project","Task","Subject",
		"Priority","KT","ET","AT","TRT","ATP","Status","Allocated to", "Remarks"
	]

	data = []

	daily_monitors = frappe.get_all(
		"Daily Monitor",
		filters={"custom_dm_production_date": date},
		fields=["name","sprint","dev_team"],
	)

	for dm in daily_monitors:
		doc = frappe.get_doc("Daily Monitor", dm.name)
		for row in doc.task_details:
			kt = frappe.db.get_value("Task", {"name": row.id}, "kt_confirmed")
			allocated_to = frappe.db.get_value("Task",{'name':row.id},'custom_allocated_to')
			if row.cb != row.cb_against_task and float(row.at_taken or 0) > 0:
				data.append([
					dm.sprint,
					dm.dev_team,
					row.cb,
					row.project_name,
					row.id,
					row.subject,
					row.priority,
					kt,
					row.et,
					round(row.at, 2),
					row.today_rt,
					float(row.at_taken or 0),
					row.current_status,
					allocated_to,
					row.remark
				])

	data = sorted(data, key=lambda x: (x[1] or "", x[0] or "", x[2] or "", x[3] or "", x[6] or ""))

	data_with_sn = []
	for idx, row in enumerate(data, start=1):
		data_with_sn.append([idx] + row)  
	df = pd.DataFrame(data_with_sn, columns=columns)
	return df


def in_sprint_not_taken(date):
	columns = [
		"S.NO", "Sprint","Team","CB","Project","Task","Subject",
		"Priority","KT","ET","RT","Status","Creation"
	]

	data = []

	tasks = frappe.get_all(
		"Task",
		filters={
			"custom_sprint": ["is", "set"],
			"custom_production_date": ["is", "not set"],
			"status": ["in", ["Open","Working"]]
		},
		fields=[
			"name",
			"custom_sprint",
			"custom_dev_team",
			"cb",
			"project",
			"subject",
			"priority",
			"kt_confirmed",
			"expected_time",
			"rt",
			"status",
			"creation"
		],
		order_by="custom_sprint asc, custom_dev_team asc"
	)
	tasks = sorted(
			tasks,
			key=lambda x: (
				x.get("custom_dev_team") or "",
				x.get("custom_sprint") or "",
				x.get("cb") or "",
				x.get("project") or "",
				x.get("priority") or ""
			)
		)
	for idx, i in enumerate(tasks, start=1):
		creation_date = (i.get("creation").strftime("%d-%m-%Y") if i.get("creation") else None)
		data.append([
			idx,  
			i.get("custom_sprint"),
			i.get("custom_dev_team"),
			i.get("cb"),
			i.get("project"),
			i.get("name"),
			i.get("subject"),
			i.get("priority"),
			i.get("kt_confirmed"),
			i.get("expected_time"),
			i.get("rt"),
			i.get("status"),
			creation_date
		])

	df = pd.DataFrame(data, columns=columns)
	return df



def not_in_sprint(date):
	columns = [
		"S.NO","Task","Project Name","Subject",
		"Priority","KT","ET","RT","Status","Creation"
	]

	data = []

	tasks = frappe.db.get_all(
		"Task",
		filters={
			"custom_sprint": ["is", "not set"],
			"status": ["not in", ["Completed","Cancelled","Client Review","Pending Review","Code Review","Hold"]],
			"service":"IT-SW"
		},
		fields=[
			"name",
			"custom_sprint",
			"custom_dev_team",
			"cb",
			"project",
			"subject",
			"priority",
			"kt_confirmed",
			"expected_time",
			"rt",
			"status",
			"creation"
		]
	)
	tasks = sorted(tasks, key=lambda x: (x.get("project") or "", x.get("priority") or ""))

	for idx, i in enumerate(tasks, start=1):
		creation_date = (i.get("creation").strftime("%d-%m-%Y") if i.get("creation") else None)
		data.append([
			idx,  
			i.get("name"),
			i.get("project"),
			i.get("subject"),
			i.get("priority"),
			i.get("kt_confirmed"),
			i.get("expected_time"),
			i.get("rt"),
			i.get("status"),
			creation_date
		])

	df = pd.DataFrame(data, columns=columns)
	return df




def summary(date, file_path):
	import frappe
	from openpyxl import load_workbook
	from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
	from openpyxl.utils import get_column_letter
	from collections import defaultdict


	today_date = date
	wb = load_workbook(file_path)
	ws = wb.create_sheet("Summary")

	# --- Styles ---
	bold_font = Font(bold=True)
	center_align = Alignment(horizontal="center", vertical="center")
	align_left = Alignment(horizontal="left", vertical="center")
	thin_border = Border(
		left=Side(style='thin'), right=Side(style='thin'),
		top=Side(style='thin'), bottom=Side(style='thin')
	)

	# --- Headers ---
	headers = ["Sprint", "Dev Team", "CB",
			   "Sprint (Hrs) Plan", "Sprint (Hrs) Comp", "Sprint (Hrs) Working","Sprint (Hrs) NT",
			   "Others (Hrs) Plan", "Others (Hrs) Completed", "Others (Hrs) Working", "Others (Hrs) NT",
			   "Total (Hrs) Planned", "Total (Hrs) Completed", "Total (Hrs) Working","Total (Hrs) Not Taken",
			   "Revisions", "NC",
			   "Sprint Count Plan", "Sprint Count Comp", "Sprint Count Working", "Sprint Count NT",
			   "Others Count Plan", "Others Count Comp", "Others Count Working", "Others Count NT","Revision(count)","NC"]
	ws.append(headers)
	for cell in ws[1]:
		cell.font = bold_font
		cell.alignment = center_align
		cell.border = thin_border

	# --- Fetch Data from Daily Monitor ---
	daily_data = frappe.get_all(
		"Daily Monitor",
		filters={"custom_dm_production_date": date},
		fields=["sprint", "dev_team", "name"]
	)

	# --- Prepare set of task-names already present in Daily Monitor (to exclude from Others) ---
	dm_task_names = set()
	for row in daily_data:
		dm_childs = frappe.get_all(
			"Allocated Tasks",
			filters={"parent": row["name"]},
			fields=["id"]
		)
		for d in dm_childs:
			# 'id' in Allocated Tasks corresponds to the linked Task name in your earlier code
			dm_task_names.add(d["id"])

	# --- Aggregate all CB totals first (from Daily Monitor / Allocated Tasks) ---
	cb_totals = {}
	cb_info = {}  # Store sprint, dev_team, and Daily Monitor name per CB

	for row in daily_data:
		tasks = frappe.get_all(
			"Allocated Tasks",
			filters={"parent": row["name"]},
			fields=["cb", "today_rt", "current_status", "at_taken", "revisions", "id"]
		)
		for t in tasks:
			cb = t["cb"]
			today_rt = t["today_rt"] or 0

			if cb not in cb_totals:
				cb_totals[cb] = {
					"total": 0,
					"not_open_working_code": 0,
					"open_working_code": 0,
					"at_taken_zero": 0,
					"revisions_code": 0,
					"revisions_count":0,
					"nc_rt": 0,
					"sprint_count_plan": 0,
					"sprint_count_completed": 0,
					"sprint_count_working": 0,
					"sprint_count_nt": 0
				}
				cb_info[cb] = {
					"sprint": row["sprint"],
					"dev_team": row["dev_team"],
					"daily_monitor_name": row["name"]
				}

			# ---- Aggregate sprint values from the allocated task row ----
			emp_email = frappe.db.get_value("Employee", {"short_code": cb,'department':'IT. Development - THIS'}, "user_id")

			sprint_plan = cb_totals[cb]["total"]
			sprint_completed = cb_totals[cb]["not_open_working_code"]
			sprint_working = cb_totals[cb]["open_working_code"]

			if emp_email:
				rt = today_rt
				sprint_plan += rt
				cb_totals[cb]["sprint_count_plan"] += 1

				status = t.get("current_status")
				# keep same logic you used earlier for sprint completed/working
				if status not in ("Open", "Working", "Code Review"):
					sprint_completed += rt
					cb_totals[cb]["sprint_count_completed"] += 1
				else:
					sprint_working += rt
					cb_totals[cb]["sprint_count_working"] += 1

				# if no timesheet exists → this is NT
				timesheet_exists = frappe.db.exists(
					"Timesheet Detail", {"task": t["id"], 'start_date': today_date}
				)
				if not timesheet_exists:
					cb_totals[cb]["sprint_count_nt"] += 1

				# Save totals
				cb_totals[cb]["total"] = sprint_plan
				cb_totals[cb]["not_open_working_code"] = sprint_completed
				cb_totals[cb]["open_working_code"] = sprint_working

			# Revisions (sum from Task)
			task_doc = frappe.get_all(
				"Task",
				filters={"name": t["id"], "cb": cb},
				fields=["revisions","rt"]
			)
			if task_doc:
				cb_totals[cb]["revisions_code"] += task_doc[0]["rt"] or 0

			task_doc = frappe.get_all(
				"Task",
				filters={"name": t["id"], "cb": cb},
				fields=["revisions","rt"]
			)
			if task_doc:
				cb_totals[cb]["revisions_count"] += task_doc[0]["revisions"] or 0

			# NC RT
			emp_id = frappe.db.get_value("Employee", {"short_code": cb}, "name")
			if emp_id and frappe.db.exists(
				"Energy Point And Non Conformity",
				{"task": t["id"], "docstatus": ["!=", 2], "emp": emp_id}
			):
				cb_totals[cb]["nc_rt"] += today_rt

			nc_count = frappe.db.sql("""
				SELECT SUM(e.nc_score)
				FROM `tabEnergy Point And Non Conformity` e
				JOIN `tabTask` t
				ON e.task = t.name			
				WHERE t.cb = %s
				AND (t.custom_production_date = %s OR t.custom_pr_date = %s)
				AND e.emp = %s
				AND e.docstatus != 2
			""", (cb, today_date, today_date, emp_id))[0][0] or 0

	# --- Write all CBs with Others (Hrs) Plan ---
	row_num = 2
	for cb, totals in cb_totals.items():
		# Others counts per CB
		others_count_plan = 0
		others_count_completed = 0
		others_count_working = 0
		others_count_nt = 0
		others_hrs_plan = 0
		completed_hrs = 0
		working_hrs = 0
		not_taken_hrs = 0

		# Check if CB is TL
		employee_docs = frappe.get_all("Employee", filters={"short_code": cb, 'department': 'IT. Development - THIS'}, fields=["name", "custom_is_tl","user_id"])
		is_tl = employee_docs[0]["custom_is_tl"] if employee_docs else False

		# Build list of relevant CBs (the CB itself + team members if TL)
		relevant_cbs = {cb}
		if is_tl:
			tl_name = employee_docs[0]["name"]
			team_members = frappe.get_all(
				"Employee",
				filters={"custom_tl": tl_name},
				fields=["short_code"]
			)
			for member in team_members:
				relevant_cbs.add(member["short_code"])

		# Collect tasks for Others from relevant_cbs where (custom_pr_date OR custom_production_date) == today
		tasks_to_sum = []
		added_task_names = set()

		for person_cb in relevant_cbs:
			# tasks with custom_pr_date == today
			tasks_pr = frappe.get_all(
				"Task",
				filters={"cb": person_cb, "custom_pr_date": today_date},
				fields=["rt", "status", "name"]
			) or []

			# tasks with custom_production_date == today
			tasks_prod = frappe.get_all(
				"Task",
				filters={"cb": person_cb, "custom_production_date": today_date},
				fields=["rt", "status", "name"]
			) or []

			for tsk in tasks_pr + tasks_prod:
				# exclude tasks already included in Daily Monitor
				if tsk["name"] in dm_task_names:
					continue
				# dedupe
				if tsk["name"] in added_task_names:
					continue
				added_task_names.add(tsk["name"])
				tasks_to_sum.append(tsk)

		# 🔹 Aggregate Others Hrs and Counts
		for t in tasks_to_sum:
			rt_val = t.get("rt", 0) or 0

			# Plan
			others_hrs_plan += rt_val
			others_count_plan += 1

			# Your rule: status NOT in (Open/Working/Code Review) -> Others Working
			#              status IN (Open/Working/Code Review) -> Others Completed
			if t.get("status") in ("Open", "Working", "Code Review"):
				working_hrs += rt_val
				others_count_working += 1
			else:
				completed_hrs += rt_val
				others_count_completed += 1

			# Timesheet check -> Not Taken
			timesheet_exists = frappe.db.exists(
				"Timesheet Detail", {"task": t["name"], "start_date": today_date}
			)
			if not timesheet_exists:
				not_taken_hrs += rt_val
				others_count_nt += 1

		# --- Totals ---
		total_plan = totals["total"] + others_hrs_plan
		total_completed = totals["not_open_working_code"] + completed_hrs
		total_working = totals["open_working_code"] + working_hrs
		total_nt = totals["at_taken_zero"] + not_taken_hrs

		# --- Append row ---
		ws.append([
			cb_info[cb]["sprint"],
			cb_info[cb]["dev_team"],
			cb,
			totals["total"],
			totals["not_open_working_code"],
			totals["open_working_code"],
			totals["at_taken_zero"],
			others_hrs_plan,
			completed_hrs,
			working_hrs,
			not_taken_hrs,
			total_plan,
			total_completed,
			total_working,
			total_nt,
			totals["revisions_count"],
			totals["nc_rt"],
			totals["sprint_count_plan"],
			totals["sprint_count_completed"],
			totals["sprint_count_working"],
			totals["sprint_count_nt"],
			others_count_plan,
			others_count_completed,
			others_count_working,
			others_count_nt,
			totals["revisions_count"],
			nc_count
		])

		# Style row
		for cell in ws[row_num]:
			# cell.border = thin_border
			cell.alignment = center_align
		row_num += 1

	# --- Column widths ---
	for idx in range(1, len(headers) + 1):
		ws.column_dimensions[get_column_letter(idx)].width = 18

	# --------------------- SECOND SUMMARY (Production / In Sprint / Inventory) --------------------- #
	try:
		production_data = production(date).values.tolist()
		in_sprint_data = in_sprint_not_taken(date).values.tolist()
		not_in_sprint_data = not_in_sprint(date).values.tolist()
	except Exception:
		production_data, in_sprint_data, not_in_sprint_data = [], [], []

	if production_data or in_sprint_data or not_in_sprint_data:
		start_row = row_num + 2  # leave a blank row after CB summary

		# Headers
		ws.merge_cells(f"A{start_row}:B{start_row}")
		ws.merge_cells(f"D{start_row}:E{start_row}")
		ws.merge_cells(f"G{start_row}:H{start_row}")
		ws[f"A{start_row}"], ws[f"D{start_row}"], ws[f"G{start_row}"] = "", "In Sprint Not Taken", "Not in Sprint (Inventory)"
		for cell in [f"A{start_row}", f"B{start_row}", f"D{start_row}", f"E{start_row}", f"G{start_row}", f"H{start_row}"]:
			ws[cell].font = bold_font
			ws[cell].alignment = center_align

		# ws[f"A{start_row+1}"], ws[f"B{start_row+1}"] = "Row Labels", "Sum of TRT"
		ws[f"D{start_row+1}"], ws[f"E{start_row+1}"] = "Row Labels", "Sum of RT"
		ws[f"G{start_row+1}"], ws[f"H{start_row+1}"] = "Row Labels", "RT (hrs)"
		for cell in [f"A{start_row+1}", f"B{start_row+1}", f"D{start_row+1}", f"E{start_row+1}", f"G{start_row+1}", f"H{start_row+1}"]:
			ws[cell].font = bold_font
			ws[cell].alignment = center_align

		# --- Build dictionaries ---
		prod_dict = defaultdict(lambda: defaultdict(float))
		for row in production_data:
			team, cb, trt = row[2], row[3], row[11] or 0
			prod_dict[team][cb] += trt

		sprint_dict = defaultdict(lambda: defaultdict(float))
		for row in in_sprint_data:
			team, sprint, rt = row[2], row[1], row[10] or 0
			sprint_dict[team][sprint] += rt

		inv_dict = defaultdict(float)
		for row in not_in_sprint_data:
			project, rt = row[2], row[7] or 0
			inv_dict[project] += rt

		# row_num = start_row + 2
		# # --- Production ---
		# for team, cbs in prod_dict.items():
		# 	team_total = sum(cbs.values())
		# 	ws.cell(row=row_num, column=1, value=team).font = bold_font
		# 	ws.cell(row=row_num, column=2, value=team_total).font = bold_font
		# 	ws.cell(row=row_num, column=2).alignment = center_align
		# 	row_num += 1
		# 	for cb, trt in cbs.items():
		# 		ws.cell(row=row_num, column=1, value=cb).alignment = align_left
		# 		ws.cell(row=row_num, column=2, value=trt).alignment = center_align
		# 		row_num += 1
		# ws.cell(row=row_num, column=1, value="Grand Total").font = bold_font
		# ws.cell(row=row_num, column=2, value=sum([trt for cbs in prod_dict.values() for trt in cbs.values()])).alignment = center_align


		row_num = start_row + 2
		# --- In Sprint Not Taken ---
		for team, sprints in sprint_dict.items():
			team_total = sum(sprints.values())
			ws.cell(row=row_num, column=4, value=team).font = bold_font
			ws.cell(row=row_num, column=5, value=team_total).font = bold_font
			ws.cell(row=row_num, column=5).alignment = center_align
			row_num += 1
			for sprint, rt in sprints.items():
				ws.cell(row=row_num, column=4, value=sprint).alignment = align_left
				ws.cell(row=row_num, column=5, value=rt).alignment = center_align
				row_num += 1
		ws.cell(row=row_num, column=4, value="Grand Total").font = bold_font
		ws.cell(row=row_num, column=5, value=sum([rt for sprints in sprint_dict.values() for rt in sprints.values()])).alignment = center_align


		row_num = start_row + 2
		# --- Not in Sprint ---
		for project, rt in inv_dict.items():
			ws.cell(row=row_num, column=7, value=project).alignment = align_left
			ws.cell(row=row_num, column=8, value=rt).alignment = center_align
			row_num += 1
		ws.cell(row=row_num, column=7, value="Grand Total").font = bold_font
		ws.cell(row=row_num, column=8, value=sum(inv_dict.values())).alignment = center_align

		for col in ["A","B","D","E","G","H"]:
			ws.column_dimensions[col].width = 25

	wb.save(file_path)

# from frappe.utils.csvutils import UnicodeWriter, read_csv_content
# from frappe.utils.csvutils import read_csv_content
# @frappe.whitelist()
# def batch_status_update_exisiting():
# 	filename='d2425582a710043SAMS Status Update.csv'
# 	from frappe.utils.file_manager import get_file
# 	filepath = get_file(filename)
# 	pps = read_csv_content(filepath[1])
# 	ind=0
# 	for pp in pps:
# 		if pp[0] != "ID":
# 			frappe.db.set_value("SAMS",{"name":pp[0]},"sa_status",pp[1])
# 			ind+=1
# 			print(pp[0])
# 			print(pp[1])
# 	print(ind)





from frappe.utils import formatdate

import frappe

@frappe.whitelist()
def get_tasks_by_date_and_employee():
	employee = "TC00031"
	date = "2025-12-04"
	custom_dev_team = "CHARLIE"
	has_sub_tl = frappe.db.exists(
		"Employee",
		{
			"custom_dev_team": custom_dev_team,
			"custom_is_sub_tl": 1
		}
	)
	emp = frappe.db.get_value("Employee", employee, ["user_id"])
	tasks = frappe.get_all(
		"Task",
		filters={
			"custom_production_date": date,
			"custom_allocated_to": emp
		},
		fields=["name", "status", "rt", "project", "subject", "priority",'custom_dev_team','cb']
	)

	priority_order = {"Urgent": 1, "High": 2, "Medium": 3, "Low": 4}
	tasks.sort(key=lambda x: priority_order.get(x.get("priority") or "Low", 5))
	cdr_list = []
	is_tl = frappe.db.get_value('Employee', {'name': employee}, ['custom_is_tl'])
	is_sub_tl = frappe.db.get_value('Employee', {'name': employee}, ['custom_is_sub_tl'])
	allocated_persons = []
	if not has_sub_tl:
		if is_tl == 1:
			print("if1")
			user = frappe.db.get_value('Employee', {'name': employee}, ['user_id'])
			team = frappe.db.get_value('Dev Team', {'code_reviewer': user}, ['name'])
			if team:
				team_tl = frappe.db.get_value('Employee', {
					'status': 'Active',
					'custom_is_tl': 1,
					'custom_dev_team': team
				}, ['user_id'])
				if team_tl:
					allocated_persons.append(team_tl)
			cdr_employees = frappe.db.get_all(
				'Employee',
				{'custom_is_tl': 0, 'custom_tl': employee},
				['user_id']
			)
			for cdr in cdr_employees:
				allocated_persons.append(cdr.user_id)
			cdr_tasks = frappe.db.get_all(
				'Task',
				filters={
					'custom_allocated_to': ('in', allocated_persons),
					'custom_pr_date': date
				},
				fields=['name', 'subject', 'project', 'status', 'priority','custom_dev_team','cb']
			)

			print(cdr_tasks)


			cdr_tasks.sort(key=lambda x: priority_order.get(x.get("priority") or "Low", 5))
			for tsk in cdr_tasks:
				cdr_list.append({
					'task': tsk['name'],
					'subject': tsk.get('subject', ''),
					'project': tsk.get('project', ''),
					'hours': 0,
					'task_status': tsk.get('status', ''),
					'priority':tsk.get('priority', ''),
					'description': ''
				})
	else:
		if is_tl == 1 and is_sub_tl ==0:
			print("if2")
			user = frappe.db.get_value('Employee', {'name': employee}, ['user_id'])
			team = frappe.db.get_value('Dev Team', {'code_reviewer': user}, ['name'])
			if team:
				team_tl = frappe.db.get_value('Employee', {
					'status': 'Active',
					'custom_is_tl': 1,
					'custom_dev_team': team
				}, ['user_id'])
				if team_tl:
					allocated_persons.append(team_tl)
			cdr_employees = frappe.db.get_all(
				'Employee',
				{'custom_is_tl': 0, 'custom_tl': employee},
				['user_id']
			)
			for cdr in cdr_employees:
				allocated_persons.append(cdr.user_id)
			cdr_tasks = frappe.db.get_all(
				'Task',
				filters={
					'custom_allocated_to': ('in', allocated_persons),
					'custom_pr_date': date
				},
				fields=['name', 'subject', 'project', 'status', 'priority','custom_dev_team','cb']
			)

			print(cdr_tasks)


			cdr_tasks.sort(key=lambda x: priority_order.get(x.get("priority") or "Low", 5))
			for tsk in cdr_tasks:
				cdr_list.append({
					'task': tsk['name'],
					'subject': tsk.get('subject', ''),
					'project': tsk.get('project', ''),
					'hours': 0,
					'task_status': tsk.get('status', ''),
					'priority':tsk.get('priority', ''),
					'description': ''
				})
		if is_tl==1 and is_sub_tl==1:
			print("if3")
			user = frappe.db.get_value('Employee', {'name': employee}, ['user_id'])
			cdr_employees = frappe.db.get_all(
				'Employee',
				{'custom_is_tl': 0, 'custom_tl': employee},
				['user_id']
			)
			for cdr in cdr_employees:
				allocated_persons.append(cdr.user_id)
			cdr_tasks = frappe.db.get_all(
				'Task',
				filters={
					'custom_allocated_to': ('in', allocated_persons),
					'custom_pr_date': date
				},
				fields=['name', 'subject', 'project', 'status', 'priority','custom_dev_team','cb']
			)

			print(cdr_tasks)


			cdr_tasks.sort(key=lambda x: priority_order.get(x.get("priority") or "Low", 5))
			for tsk in cdr_tasks:
				cdr_list.append({
					'task': tsk['name'],
					'subject': tsk.get('subject', ''),
					'project': tsk.get('project', ''),
					'hours': 0,
					'task_status': tsk.get('status', ''),
					'priority':tsk.get('priority', ''),
					'description': ''
				})




	if not tasks and not cdr_list:
		return "<div>No tasks found for the selected date and employee.</div>"

	html = """
	<style>
		.task-table, .task-table th, .task-table td {
			border: 1px solid black;
			border-collapse: collapse;
		}
	</style>
	<table class="table table-bordered">
		<thead>
			<tr style='background-color:#0f1568;color:white;text-align:center'>
				<th>Task ID</th>
				<th>Project</th>
				<th>Subject</th>
				<th>Status</th>
				<th>Priority</th>
				<th>RT</th>
				<th>Today RT</th>
			</tr>
		</thead>
		<tbody>
	"""
	for task in tasks:
		today_rt = get_today_rt_from_child(task["name"], date,task["custom_dev_team"],task["cb"])

		html += f"""
			<tr>
				<td>{task.name}</td>
				<td>{task.project}</td>
				<td>{task.subject}</td>
				<td>{task.status}</td>
				<td>{task.priority}</td>
				<td style='text-align:right'>{task.rt or ''}</td>
				<td style='text-align:right'>{today_rt or task.rt}</td>
			</tr>
		"""

	for cdr in cdr_list:
		html += f"""
			<tr style="background-color:#d3e8f2;">
				<td>{cdr['task']}</td>
				<td>{cdr['project']}</td>
				<td>{cdr['subject']}</td>
				<td>{cdr['task_status']}</td>
				<td>{cdr['priority']}</td>
				<td style='text-align:right'>0.5</td>
				<td style='text-align:right'>0.5</td>
			</tr>
		"""


	html += "</tbody></table>"
	return html

def get_today_rt_from_child(task_id, date,team,cb=None):
    daily_monitors = frappe.get_all(
        "Daily Monitor",
        filters={"date": date,"dev_team":team},
        fields=["name"]
    )
    for dm in daily_monitors:
        child = frappe.db.get_value(
            "Allocated Tasks",
            {"parent": dm.name, "id": task_id,"cb":cb},
            "today_rt"
        )
        if child:
            return child
    return 0
