from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
# import pandas as pd
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles



@frappe.whitelist()
def download():
	filename = 'Batch'
	build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	
	if wb is None:
		wb = Workbook()
	
	ws = wb.create_sheet(sheet_name, 0)
	ws.append(["Batch ID","Customer","Expected Start Date","Expected End Date","Batch Status","No Of Cases","#Pending","#Comp","#Insuff","#Billing Status"])
	align_center = Alignment(horizontal='center',vertical='center')
	for header in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=10):
		for cell in header:
			cell.font = Font(bold=True)
			cell.alignment = align_center
			cell.fill = PatternFill(fgColor='D7BDE2', fill_type = "solid")
	data1= get_data(args)
	for row in data1:
		ws.append(row)

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'


@frappe.whitelist()
def get_data(args):
    ss = []
    checks = frappe.get_all('Batch', fields=['*'])
    
    for i in checks:
        expected_start_date = i.expected_start_date
        expected_end_date = i.expected_end_date
        
        if expected_start_date and expected_end_date:
            parsed_start_date = datetime.strptime(str(expected_start_date), '%Y-%m-%d')
            formatted_start_date = parsed_start_date.strftime('%d-%m-%Y')
            parsed_end_date = datetime.strptime(str(expected_end_date), '%Y-%m-%d')
            formatted_end_date = parsed_end_date.strftime('%d-%m-%Y')
        else:
            formatted_start_date = ''
            formatted_end_date = ''
        
        sa = frappe.db.count("Case", filters={"batch": i.name, "entry_status": "Pending"})
        sb = frappe.db.count("Case", filters={"batch": i.name, "entry_status": "Completed"})
        sc = frappe.db.count("Case", filters={"batch": i.name, "entry_status": "Insufficient"})
        row = [
            i.name, i.customer, formatted_start_date, formatted_end_date,i.batch_status, i.no_of_cases, sa, sb, sc, i.billing_status
        ]
        ss.append(row)
    
    return ss
