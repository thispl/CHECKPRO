from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
# import pandas as pd
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles



@frappe.whitelist()
def download():
	filename = 'Check Status Report'
	build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	
	if wb is None:
		wb = Workbook()
	
	ws = wb.create_sheet(sheet_name, 0)
	ws.append(["Batch Number","Client Name","TAT Date","Age","Checks","Cases","Insuff","Entry By","Entry QC","Education Checks","Employement Checks","Address Checks","Court Checks","Reference Checks","Identity Checks","Family Checks","Social Checks","Final QC","Completed"])
	align_center = Alignment(horizontal='center',vertical='center')
	for header in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=19):
		for cell in header:
			cell.font = Font(bold=True)
			cell.fill = PatternFill(fgColor='D7BDE2', fill_type = "solid")
			cell.alignment = align_center

	data1= get_data(args)
	for row in data1:
		ws.append(row)
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file


def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'


@frappe.whitelist()
def get_data(args):
	ss = []
	batches = frappe.get_all('Batch', filters={
		'batch_status': ['in', ['Open', 'Overdue', 'Insufficient']]
	}, fields=['*'])
	for batch in batches:
		check_package_name = batch.get('check_package')
		
	

		if check_package_name:
			check_package = frappe.get_doc('Check Package', check_package_name)
			package_tat = check_package.get('package_tat')
	insuff = frappe.get_all('Batch', fields=['insuff'])
	
	checks = frappe.get_all('Batch', filters={
		'batch_status': ['in', ['Open', 'Overdue', 'Insufficient']]
	}, fields=['*'])
	for i in checks:
		sa = frappe.db.count("Case", filters={"batch": i.name, "entry_status": "Pending"})
		sb = frappe.db.count("Case", filters={"batch": i.name, "entry_status": "Completed"})
		sc = frappe.db.count("Case", filters={"batch": i.name, "entry_status": "Insufficient"})
		expected_end_date = getdate(i.expected_end_date)
		today = date.today()
		tat_day = (expected_end_date - today).days
		
		cases = frappe.get_all('Case', filters={"batch": i.name}, fields=["entered_by", "entered_by_qc"])
		sg = sum(1 for case in cases if case.entered_by)
		sh = sum(1 for case in cases if case.entered_by_qc)
		new = frappe.get_doc("Batch",i.name)
		
		expected_end_date = i.expected_end_date
		
		if expected_end_date:
			parsed_end_date = datetime.strptime(str(expected_end_date), '%Y-%m-%d')
			formatted_end_date = parsed_end_date.strftime('%d-%m-%Y')
		else:
			formatted_end_date = ''
		count = 0
		for j in new.check_list:
			count = j.idx
		row = [i.name, i.customer, formatted_end_date,tat_day,count,i.no_of_cases,i.insuff,sg,sh]
		query = frappe.db.sql(""" select `tabCheckwise Report`.checks,`tabCheckwise Report`.verification_status,count(`tabCheckwise Report`.verification_status) as count from `tabCheckwise Report` where parent = '%s' group by `tabCheckwise Report`.checks  """%(i.name),as_dict=1)
		education = 0
		employment = 0
		address = 0
		court = 0
		reference = 0
		identity = 0
		family = 0
		social = 0
		for k in query:
			if k.checks == "Education Checks" and k.verification_status == "YTS":
				education = k.count
			if k.checks == "Employment" and k.verification_status == "YTS":
				employment = k.count
			if k.checks == "Address Check" and k.verification_status == "YTS":
				address = k.count
			if k.checks == "Court" and k.verification_status == "YTS":
				court = k.count
			if k.checks == "Reference Check" and k.verification_status == "YTS":
				reference = k.count
			if k.checks == "Identity Aadhar" and k.verification_status == "YTS":
				identity = k.count
			if k.checks == "Family" and k.verification_status == "YTS":
				family = k.count
			if k.checks == "Social Media" and k.verification_status == "YTS":
				social = k.count
		
		row += [education,employment,address,court,reference,identity,family,social]
		education_checks_pending_count = frappe.db.count("Education Checks",
														 filters={"batch": i.name, "workflow_state": "Pending for Verification"})
		employment_checks_pending_count = frappe.db.count("Employment",
														 filters={"batch": i.name, "workflow_state": "Pending for Verification"})
		
		Address_checks_pending_count = frappe.db.count("Address Check",
														 filters={"batch": i.name, "workflow_state": "Pending for Verification"})
		court_pending_count = frappe.db.count("Court",
														 filters={"batch": i.name, "workflow_state": "Pending for Verification"})
		
		refer_pending_count = frappe.db.count("Reference Check",
														 filters={"batch": i.name, "workflow_state": "Pending for Verification"})
		
		identity_pending_count = frappe.db.count("Identity Aadhar",
														 filters={"batch": i.name, "workflow_state": "Pending for Verification"})
		
		family_pending_count = frappe.db.count("Family",
														 filters={"batch": i.name, "workflow_state": "Pending for Verification"})
		

		social_pending_count = frappe.db.count("Social Media",
														 filters={"batch": i.name, "workflow_state": "Pending for Verification"})
		
		
		
		var = (education_checks_pending_count + employment_checks_pending_count + Address_checks_pending_count + court_pending_count + refer_pending_count + identity_pending_count + family_pending_count + social_pending_count)
		
		row += [var]

		education_checks_com = frappe.db.count("Education Checks",
														 filters={"batch": i.name, "verification_status": "completed"})
		employment_checks_com = frappe.db.count("Employment",
														 filters={"batch": i.name, "verification_status": "completed"})
		
		Address_checks_com = frappe.db.count("Address Check",
														 filters={"batch": i.name, "verification_status": "completed"})
		court_com = frappe.db.count("Court",
														 filters={"batch": i.name, "verification_status": "completed"})
		
		refer_com = frappe.db.count("Reference Check",
														 filters={"batch": i.name, "verification_status": "completed"})
		
		identity_com = frappe.db.count("Identity Aadhar",
														 filters={"batch": i.name, "verification_status": "completed"})
		
		family_com = frappe.db.count("Family",
														 filters={"batch": i.name, "verification_status": "completed"})
		

		social_com = frappe.db.count("Social Media",
														 filters={"batch": i.name, "verification_status": "completed"})
		
		com = (education_checks_com + employment_checks_com + Address_checks_com + court_com + refer_com + identity_com + family_com + social_com)
		
		row += [com]

		ss.append(row)
	
	return ss
