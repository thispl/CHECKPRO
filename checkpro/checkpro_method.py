import frappe
import frappe.utils
from frappe.utils.csvutils import read_csv_content
from frappe.utils import get_first_day, get_last_day, format_datetime, get_url_to_form
from frappe.utils import cint
from frappe.utils.data import date_diff, now_datetime, nowdate, today, add_days
import datetime
from io import BytesIO
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment, Border, Side
import openpyxl
from frappe import _
from frappe import _
from frappe.model.document import Document
from frappe.utils import getdate, nowdate
from frappe import throw, msgprint
import frappe
from frappe.utils import flt, fmt_money
from datetime import timedelta
from datetime import date
from frappe import throw, _
from frappe.utils import getdate, today
today = date.today()
from frappe.model.document import Document
import datetime 
import frappe,erpnext
from frappe.utils import cint
import json
from frappe.utils import date_diff, add_months,today,add_days,add_years,nowdate,flt
from frappe.model.mapper import get_mapped_doc
from frappe.utils.file_manager import get_file
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
import datetime
from datetime import date,datetime,timedelta
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
import pandas as pd
from frappe.utils import formatdate
from frappe.utils import now
from erpnext.setup.utils import get_exchange_rate
from datetime import date
from six import BytesIO, string_types
from frappe.utils import time_diff

@frappe.whitelist()
def create_so_case(case_id):
    doc_name = json.loads(case_id)
    customer = []
    check_package = []
    for c in doc_name:
        case=frappe.get_doc("Case",c)
        customer.append(case.customer)
        check_package.append(case.check_package)
    if all(cust == customer[0] for cust in customer) and all(check_pac == check_package[0] for check_pac in check_package):
        so = frappe.new_doc("Sales Order")
        so.company = "TEAMPRO HR & IT Services Pvt. Ltd."
        so.customer = case.customer
        so.service = "BCS"
        so.order_type = "Sales"
        so.delivery_date = today()  
        so.transaction_date = today()
        batch_delivery_manager = frappe.db.get_value("Batch",case.batch,['batch_manager']) 
        so.delivery_manager = batch_delivery_manager
        so.posa_notes=case.case_report
        so.tc_name="Account Details - THIS"
        for i in doc_name:
            case=frappe.get_doc("Case",i)
            if case.case_status =="To be Billed":
                batch = frappe.db.get_value("Batch",case.batch,['customers_purchase_order'])
                item = frappe.new_doc("Item")
                item.item_code = i
                item.item_name= case.case_name
                item.item_group = "BCS Cases"
                item.item_group_code= "BCS"
                item.stock_uom = "Nos"
                item.qty = "1"
                item.gst_hsn_code = '998521'
                item.is_stock_item = "0"
                item.include_item_in_manufacturing = "0"
                dict_list = []
                dict_list.append(frappe._dict({"item_tax_template":"GST 18% - THIS","tax_category":"Tamil Nadu","valid_from": today()}))
                dict_list.append(frappe._dict({"item_tax_template":"I - GST @ 18% - THIS","tax_category":"Inter State","valid_from": today()}))
                for j in dict_list:
                    item.append("taxes", {
                        "item_tax_template":j.item_tax_template,
                        "tax_category":j.tax_category,
                        "valid_from": j.valid_from
                        })
                item.append("item_defaults", {
                            "company": "TeamPRO HR & IT Services Pvt. Ltd.",
                            "buying_cost_center":"Main - THIS",
                            "selling_cost_center":"Main - THIS",
                            "income_account":"Sales - THIS",
                            "expense_account":"Cost of Goods Sold - THIS"
                        })
                item.insert()
                item.save(ignore_permissions=True)

                
                # rate = frappe.db.get_value("Check Package", {"name":case.check_package},["total_sp"])
                package_price=frappe.db.get_value("Check Package",{"name":case.check_package},["pricing_model"])
                if package_price=="Lumpsum":
                    rate = frappe.db.get_value("Check Package", {"name":case.check_package},["total_sp"])
                elif package_price=="Check Based":
                    frappe.log_error(message='i',title='errors')
                    check_doc=frappe.get_doc("Check Package",{"name":case.check_package})
                    list=[]
                    rate = 0
                    for k in case.checkwise_status:
                        if k.checks_status=="Report Completed" and k.check_report !="Not Applicable":
                            check_type=k.checks
                            list.append(check_type)
                    for p in check_doc.checks_list:
                        if p.check_name in list:
                            rate+=p.check_sp
                    # for k in check_doc.checks_list:
                    #     if k.check_name in list:
                    #         rate+=k.check_sp
                so.append('items', {
                    'item_code': i,
                    'item_name':case.case_name,
                    'case_batch':case.batch,
                    'qty':1,
                    'posa_notes':case.case_report,
                    'rate':rate
                    })
                case_status=case.case_status
                billing_status = case.billing_status

                frappe.db.set_value("Case",i,"case_status","SO Created")
            
            else:
                frappe.msgprint("Case Status is not To be Billed for this Case-"+" "+i)
        so.insert()
        so.save(ignore_permissions=True)
        frappe.msgprint("Sales Order Created"+" "+"-<b> "+so.name+"</b>")
        # frappe.set_value("Case",i,"billing_status","Billed")
        
    else:
        frappe.msgprint("All Cases are not belong to same Customer and same Check Package")


@frappe.whitelist()
def case_report_submitted(case_id,mode_of_submission,proof_of_submission):
    doc_name = json.loads(case_id)
    for i in doc_name:
        case=frappe.get_doc("Case",i)
        if case.case_status =="Case Completed":
            frappe.set_value("Case",i,"mode_of_submission",mode_of_submission)
            frappe.set_value("Case",i,"proof_of_submission",proof_of_submission)
            frappe.set_value("Case",i,"case_status","To be Billed")
            
        else:
            frappe.msgprint("Case Status is not Case Completed for this Case-"+" "+i)

@frappe.whitelist()
def create_so(case_id):
    frappe.log_error('Check error','error')
    case=frappe.get_doc("Case",case_id)
    batch = frappe.db.get_value("Batch",case.batch,['customers_purchase_order'])
    item = frappe.new_doc("Item")
    item.item_code = case_id
    item.item_name= case.case_name
    item.item_group = "BCS Cases"
    item.item_group_code= "BCS"
    item.stock_uom = "Nos"
    item.qty = "1"
    item.gst_hsn_code = '998521'
    item.is_stock_item = "0"
    item.include_item_in_manufacturing = "0"
    dict_list = []
    dict_list.append(frappe._dict({"item_tax_template":"GST 18% - THIS","tax_category":"Tamil Nadu","valid_from": today()}))
    dict_list.append(frappe._dict({"item_tax_template":"I - GST @ 18% - THIS","tax_category":"Inter State","valid_from": today()}))
    for j in dict_list:
        item.append("taxes", {
            "item_tax_template":j.item_tax_template,
            "tax_category":j.tax_category,
            "valid_from": j.valid_from
            })
    item.append("item_defaults", {
                "company": "TeamPRO HR & IT Services Pvt. Ltd.",
                "buying_cost_center":"Main - THIS",
                "selling_cost_center":"Main - THIS",
                "income_account":"Sales - THIS",
                "expense_account":"Cost of Goods Sold - THIS"
            })
    item.insert()
    item.save(ignore_permissions=True)
    frappe.log_error('check error','error')

    so = frappe.new_doc("Sales Order")
    so.company = "TEAMPRO HR & IT Services Pvt. Ltd."
    so.customer = case.customer
    so.service = "BCS"
    so.order_type = "Sales"
    so.delivery_date = today()  
    so.transaction_date = today() 
    so.po_no = batch
    # so.delivery_manager = batch.delivery_manager
    so.posa_notes:case.case_report
    so.tc_name="Account Details - THIS"
    rate = frappe.db.get_value("Check Package", {"name":case.check_package},["total_sp"])
    
    so.append('items', {
        'item_code': case_id,
        'item_name':case.case_name,
        'case_batch':case.batch,
        'qty':1,
        'posa_notes':case.case_report,
        'rate':rate,
        })
    case_status=case.case_status
    billing_status = case.billing_status
    
    so.insert()
    so.save(ignore_permissions=True)
    frappe.msgprint("Sales Order Created"+" "+"-<b> "+so.name+"</b>")
    frappe.set_value("Case",case_id,"billing_status","Billed")
    frappe.set_value("Case",case_id,"case_status","Case Completed")
    frappe.set_value("Case",case_id,"custom_case_update_status","Case Completed")
    
@frappe.whitelist()
def update_next_action_sm(check_id,allocated_to):
    doc_name = json.loads(check_id)
    for j in doc_name:
        frappe.set_value("Social Media",j,"allocated_to",allocated_to)
        doc = frappe.get_doc("Social Media",j)
        check_status=['Draft','Entry Completed','Entry QC Pending','Entry QC Completed','Execution Initiated','Execution Pending','Execution Completed','Final QC Pending']
        if doc.workflow_state in check_status:
            indx=check_status.index(doc.workflow_state)
            next_indx=check_status[indx+1]
            frappe.set_value("Social Media",j,"workflow_state",next_indx)

@frappe.whitelist()
def update_next_action_fam(check_id,allocated_to):
    doc_name = json.loads(check_id)
    for j in doc_name:
        frappe.set_value("Family",j,"allocated_to",allocated_to)
        doc = frappe.get_doc("Family",j)
        check_status=['Draft','Entry Completed','Entry QC Pending','Entry QC Completed','Execution Initiated','Execution Pending','Execution Completed','Final QC Pending']
        if doc.workflow_state in check_status:
            indx=check_status.index(doc.workflow_state)
            next_indx=check_status[indx+1]
            frappe.set_value("Family",j,"workflow_state",next_indx)

@frappe.whitelist()
def update_next_action_edu(check_id,allocated_to):
    doc_name = json.loads(check_id)
    for j in doc_name:
        frappe.set_value("Education Checks",j,"allocated_to",allocated_to)
        doc = frappe.get_doc("Education Checks",j)
        check_status=['Draft','Entry Completed','Entry QC Pending','Entry QC Completed','Execution Initiated','Execution Pending','Execution Completed','Final QC Pending']
        if doc.workflow_state in check_status:
            indx=check_status.index(doc.workflow_state)
            next_indx=check_status[indx+1]
            frappe.set_value("Education Checks",j,"workflow_state",next_indx)

@frappe.whitelist()
def update_next_action_emp(check_id,allocated_to):
    doc_name = json.loads(check_id)
    for j in doc_name:
        frappe.set_value("Employment",j,"allocated_to",allocated_to)
        doc = frappe.get_doc("Employment",j)
        check_status=['Draft','Entry Completed','Entry QC Pending','Entry QC Completed','Execution Initiated','Execution Pending','Execution Completed','Final QC Pending']
        if doc.workflow_state in check_status:
            indx=check_status.index(doc.workflow_state)
            next_indx=check_status[indx+1]
            frappe.set_value("Employment",j,"workflow_state",next_indx)

@frappe.whitelist()
def update_next_action_addrs(check_id,allocated_to,allocate_to_supplier,supplier=None):
    doc_name = json.loads(check_id)
    for j in doc_name:
        frappe.set_value("Address Check",j,"allocated_to",allocated_to)
        doc = frappe.get_doc("Address Check",j)
        if supplier is not None:
            check_status=['Draft','Entry Completed','Entry QC Pending','Entry QC Completed','Execution Initiated','Supplier Pending','Execution Pending','Execution Completed','Final QC Pending']
            if doc.workflow_state in check_status:
                frappe.set_value("Address Check",j,"supplier",supplier)
                frappe.set_value("Address Check",j,"custom_supplier_allocation_date",frappe.utils.nowdate())
                indx=check_status.index(doc.workflow_state)
                next_indx=check_status[indx+1]
                frappe.set_value("Address Check",j,"workflow_state",next_indx)
        else:
            if doc.workflow_state == "Supplier Pending":
                frappe.set_value("Address Check",j,"workflow_state","Execution Pending")
                frappe.set_value("Address Check",j,"execution_allocation_date",frappe.utils.nowdate())
            else:
                check_status=['Draft','Entry Completed','Entry QC Pending','Entry QC Completed','Execution Pending','Execution Completed','Final QC Pending']
                if doc.workflow_state in check_status:
                    indx=check_status.index(doc.workflow_state)
                    next_indx=check_status[indx+1]
                    frappe.set_value("Address Check",j,"workflow_state",next_indx)

@frappe.whitelist()
def update_next_action_court(check_id,allocated_to):
    doc_name = json.loads(check_id)
    for j in doc_name:
        frappe.set_value("Court",j,"allocated_to",allocated_to)
        doc = frappe.get_doc("Court",j)
        check_status=['Draft','Entry Completed','Entry QC Pending','Entry QC Completed','Execution Initiated','Execution Pending','Execution Completed','Final QC Pending']
        if doc.workflow_state in check_status:
            indx=check_status.index(doc.workflow_state)
            next_indx=check_status[indx+1]
            frappe.set_value("Court",j,"workflow_state",next_indx)

@frappe.whitelist()
def update_next_action_criminal(check_id,allocated_to):
    doc_name = json.loads(check_id)
    for j in doc_name:
        frappe.set_value("Criminal",j,"allocated_to",allocated_to)
        doc = frappe.get_doc("Criminal",j)
        check_status=['Draft','Entry Completed','Entry QC Pending','Entry QC Completed','Execution Initiated','Execution Pending','Execution Completed','Final QC Pending']
        if doc.workflow_state in check_status:
            indx=check_status.index(doc.workflow_state)
            next_indx=check_status[indx+1]
            frappe.set_value("Criminal",j,"workflow_state",next_indx)

@frappe.whitelist()
def update_next_action_ref(check_id,allocated_to):
    doc_name = json.loads(check_id)
    for j in doc_name:
        frappe.set_value("Reference Check",j,"allocated_to",allocated_to)
        doc = frappe.get_doc("Reference Check",j)
        check_status=['Draft','Entry Completed','Entry QC Pending','Entry QC Completed','Execution Initiated','Execution Pending','Execution Completed','Final QC Pending']
        if doc.workflow_state in check_status:
            indx=check_status.index(doc.workflow_state)
            next_indx=check_status[indx+1]
            frappe.set_value("Reference Check",j,"workflow_state",next_indx)

@frappe.whitelist()
def update_next_action_id(check_id,allocated_to):
    doc_name = json.loads(check_id)
    for j in doc_name:
        frappe.set_value("Identity Aadhar",j,"allocated_to",allocated_to)
        doc = frappe.get_doc("Identity Aadhar",j)
        check_status=['Draft','Entry Completed','Entry QC Pending','Entry QC Completed','Execution Initiated','Execution Pending','Execution Completed','Final QC Pending']
        if doc.workflow_state in check_status:
            indx=check_status.index(doc.workflow_state)
            next_indx=check_status[indx+1]
            frappe.set_value("Identity Aadhar",j,"workflow_state",next_indx)

@frappe.whitelist()
def total_wh_appointment(in_time, name):
    current_datetime_str = frappe.utils.now_datetime().strftime('%Y-%m-%d %H:%M:%S')
    wh = total_appointment(current_datetime_str, in_time)
    
    data_list = [{'wh': wh, 'current_datetime_str': current_datetime_str}]
    
    return data_list

def total_appointment(out_time_str, in_time_str):
    out_time = datetime.strptime(out_time_str, '%Y-%m-%d %H:%M:%S')
    in_time = datetime.strptime(in_time_str, '%Y-%m-%d %H:%M:%S')
    time_difference = out_time - in_time
    
    hours = int(time_difference.total_seconds() // 3600)
    minutes = int((time_difference.total_seconds() % 3600) // 60)
    seconds = int(time_difference.total_seconds() % 60)
    formatted_time = f"{hours:02}:{minutes:02}:{seconds:02}"
    
    return formatted_time

@frappe.whitelist()
def create_appointment_from_sfu_lead(lead):
    data = frappe.db.get_value("Lead", {"name":lead}, ["email_id", "mobile_no"])
    return data

@frappe.whitelist()
def create_appointment_from_sfu_customer(customer):
    contact_name = frappe.db.get_value("Dynamic Link", {
        "link_doctype": "Customer",
        "link_name": customer,
        "parenttype": "Contact"
    }, "parent")
    if contact_name:
        email, mobile_no = frappe.db.get_value("Contact", contact_name, ["email_id", "mobile_no"])
        return {
            "email": email,
            "mobile_no": mobile_no
        }
    else:
        return {
            "email": None,
            "mobile_no": None
        }
    
@frappe.whitelist()
def update_appointment_in_sfu(name, customer_name):
    app = frappe.get_doc("Appointment", name)
    sfu = frappe.get_doc("Sales Follow Up", {"organization_name":customer_name})
    sfu.appointment = name
    sfu.appointment_status = app.status
    sfu.sheduled_time = app.scheduled_time
    sfu.appointment_with = app.appointment_with
    sfu.name1 = app.customer_name
    sfu.party = app.party
    sfu.phone_no = app.customer_phone_number
    sfu.details = app.customer_details
    sfu.skype_id = app.customer_skype
    sfu.email = app.customer_email
    sfu.calendar_event = app.calendar_event
    sfu.appointment_remarks = app.custom_remarks
    sfu.save()
    

@frappe.whitelist()
def create_exp_claim(doc, name):
    employee = frappe.db.get_value("Employee", {"user_id": frappe.session.user}, ['name'])
    approval_status = 'Draft'
    return employee, approval_status

@frappe.whitelist()
def retrieve_expenses(app):
    km = frappe.get_value("Appointment", {'name': app}, ['custom_distance'])
    km = int(km)
    a = 0
    return km, a

@frappe.whitelist()
def validate_expense(app):
    frappe.set_value("Appointment", app, "custom_expense_claimed", 1)

@frappe.whitelist()
def send_mail_for_nc(cause, id,owner, allocated, subject, project, revision, service, spoc, domain, live,dev_spoc=None):
    if service == 'IT-SW':
        reports = frappe.db.get_value("Employee", {'user_id': allocated}, ['reports_to'])
        reports_to = frappe.db.get_value("Employee", {'name': reports}, ['user_id'])
        tl=frappe.db.get_value("Employee",{'user_id':allocated},["custom_tl"])
        tl_mail=''
        if tl:
            tl_mail=frappe.db.get_value("Employee",{'name':tl},['user_id'])
        data = f"""
        <table width='100%' style='border-collapse: collapse; border: 1px solid black; text-align: center;'>
            <tr>
                <td colspan='2' style='text-align: center; background-color: #0f1568; color: white; font-size: 17px; border: 1px solid black;'>
                    <b>Task Re-Open Note</b>
                </td>
            </tr>
            <tr style='text-align: left;'>
                <td width='25%'style='border: 1px solid black;'><b>Task ID</b></td>
                <td style='border: 1px solid black;'><a href='https://erp.teamproit.com/app/task/{id}' target='_blank'>{id}</a></td>
            </tr>

            <tr style='text-align: left;'>
                <td style='border: 1px solid black;'><b>Task Statement</b></td>
                <td style='border: 1px solid black;'>{subject}</td>
            </tr>
            <tr style='text-align: left;'>
                <td style='border: 1px solid black;'><b>Project</b></td>
                <td style='border: 1px solid black;'>{project}</td>
            </tr>
            <tr style='text-align: left;'>
                <td style='border: 1px solid black;'><b>Task Raised By</b></td>
                <td style='border: 1px solid black;'>{owner}</td>
            </tr>
            <tr style='text-align: left;'>
                <td style='border: 1px solid black;'><b>Task Developed By</b></td>
                <td style='border: 1px solid black;'>{allocated}</td>
            </tr>
            <tr style='text-align: left;'>
                <td style='border: 1px solid black;'><b>Live At</b></td>
                <td style='border: 1px solid black;'>{live}</td>
            </tr>
            <tr style='text-align: left;'>
                <td style='border: 1px solid black;'><b>Domain</b></td>
                <td style='border: 1px solid black;'>{domain}</td>
            </tr>
            <tr style='text-align: left;'>
                <td style='border: 1px solid black;'><b>Re-Open Count</b></td>
                <td style='border: 1px solid black;'>{revision}</td>
            </tr>
            <tr style='text-align: left;'>
                <td style='border: 1px solid black;'><b>Re-Open Remarks</b></td>
                <td style='border: 1px solid black;'>{cause}</td>
            </tr>
        </table>
        """
        
        frappe.sendmail(
            sender=frappe.session.user,
            # recipients='divya.p@groupteampro.com',
            recipients=spoc,
            cc=[reports_to,allocated, dev_spoc,"dineshbabu.k@groupteampro.com",tl_mail],
            subject=f'Task : {id} Re-open: Forward for Re-Open',
            message=f"""
                <b>Dear Patron,<br><br>Greeting !!!</b><br><br>
                The attached Task has been re-opened and forwarded for your kind reference<br><br>
                {data}<br><br>
                Thanks & Regards,<br>TEAM ERP<br>
                <i>This email has been automatically generated. Please do not reply</i>
            """
        )

