import frappe
import frappe.utils
from frappe.utils.csvutils import read_csv_content
from frappe.utils import get_first_day, get_last_day, format_datetime, get_url_to_form
from frappe.utils import cint
from frappe.utils.data import date_diff, now_datetime, nowdate, today, add_days
import datetime
from io import BytesIO
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment, Border, Side
import openpyxl
from frappe import _
from frappe import _
from frappe.model.document import Document
from frappe.utils import getdate, nowdate
from frappe import throw, msgprint
import frappe
from frappe.utils import flt, fmt_money
from datetime import timedelta
from datetime import date
from frappe import throw, _
from frappe.utils import getdate, today
today = date.today()
from frappe.model.document import Document
import datetime 
import frappe,erpnext
from frappe.utils import cint
import json
from frappe.utils import date_diff, add_months,today,add_days,add_years,nowdate,flt
from frappe.model.mapper import get_mapped_doc
from frappe.utils.file_manager import get_file
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
import datetime
from datetime import date,datetime,timedelta
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
import pandas as pd
from frappe.utils import formatdate
from frappe.utils import now
from erpnext.setup.utils import get_exchange_rate
from datetime import date
from six import BytesIO, string_types
from frappe.utils import time_diff


@frappe.whitelist()
def update_batch():
    batch = frappe.get_all("Batch",{'batch_status':"Completed"},['*'])
    ind=1
    for b in batch:
        ind+=1
        case = frappe.get_all("Case",{"batch":b.name},["name","case_status","end_date"])
        for c in case:
            if c.end_date:
                pass
            else:
                print(c.name)
                list = ["Education Checks","Family","Reference Check","Court","Social Media","Criminal","Employment","Identity Aadhar","Address Check"]
                for i in list:
                    doc=frappe.get_all(i,{"case_id":c.name},["name","workflow_state"])
                    for j in doc:
                        if j.workflow_state != "Report Completed":
                            frappe.db.set_value(i,j.name,"workflow_state","Report Completed")
                frappe.db.set_value("Case",c.name,"case_status","Case Completed")
                frappe.db.set_value("Case",c.name,"custom_case_update_status","Case Completed")
                print(ind)
    print(ind)
    
@frappe.whitelist()
def update_case_check():
    filename='96e1e0fa9fa3573billingstatus.csv'
    from frappe.utils.file_manager import get_file
    filepath = get_file(filename)
    pps = read_csv_content(filepath[1])
    ind=0
    for pp in pps:
        frappe.db.sql("""update `tabCase` set billing_status = 'Billed' where name = %s""",(pp[0]))
        ind+=1
    print(ind)

@frappe.whitelist()
def update_status_case():
    frappe.enqueue(
        update_case_status_report, 
        queue="long", 
        timeout=36000,
        is_async=True, 
        now=False,  
        job_name='Update Case Status',
        enqueue_after_commit=False,

    ) 
@frappe.whitelist()
def update_case_status_report():
    i=0
    case=frappe.db.get_all("Case",{"case_report":["in", ['Positive', 'Negative', 'Dilemma']],"case_status":"Generate Report"},['*'])
    for c in case:
        i+=1
        frappe.db.set_value("Case",c.name,"case_status","Case Report Completed")
        frappe.db.set_value("Case",c.name,"custom_case_update_status","Case Report Completed")
        frappe.db.set_value("Case",c.name,"case_completion_date",c.end_date)
    print(i)

@frappe.whitelist()
def cases_beyond_tat_age_10():
    cases = frappe.get_all("Case", {"batch_age": (">=", 10), "case_status": ("not in", ['Case Report Completed', 'Case Completed', 'Drop','Execution-Insuff','Entry-Insuff','To be Billed','Generate Report','SO Created'])},['*'],order_by='batch_age DESC')
    data = '<table border="1" width="100%" style="border-collapse: collapse;">'
    data += '<tr style="background-color: #009dd1;"><td width=5% >S.No</td><td width=15% >Batch</td><td width=15% >Case ID</td><td width=25% >Customer</td><td width=20% >Employee Name</td><td width=10% >TAT Age</td><td width=10% >Case Status</td></tr>'
    i=1
    for c in cases:
        data += '<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>'%(i, c.batch, c.name, c.customer, c.case_name, c.batch_age, c.case_status)
        i+=1
    data += '</table>'
    frappe.sendmail(
        recipients=['dineshbabu.k@groupteampro.com','sangeetha.s@groupteampro.com','sangeetha.a@groupteampro.com',"keerthana.b@groupteampro.com"],
        cc=[''],
        subject=_("Cases having TAT Age 10 and above"),
        message="""
            Dear Sir/Madam,<br>Kindly Find the below List of Cases that are having TAT Age 10 and above %s<br>
            Thanks & Regards,<br>TEAM ERP<br>"This email has been automatically generated. Please do not reply"
            """%(data)
    )
    print(i)


import frappe
from frappe.utils import formatdate

@frappe.whitelist()
def cases_with_insuff():
    cases = frappe.get_all(
        "Case",
        filters={"case_status": ("in", ['Execution-Insuff', 'Entry-Insuff'])},
        fields=['*'],
        order_by='insufficiency_reported ASC'
    )
    

    data = '''
    <table border="1" width="100%" style="border-collapse: collapse; text-align: center;">
        <thead style="background-color: #0f1568; color: white;">
            <tr>
                <th width="5%">S.No</th>
                <th width="10%">Insuff Reported On</th>
                <th width="15%">Batch</th>
                <th width="15%">Case ID</th>
                <th width="25%">Customer</th>
                <th width="20%">Employee Name</th>
                <th width="10%">Employee Code</th>
                <th width="10%">Case Status</th>
                <th width="50%">Insuff Check(s)</th>
                <th width="5%">Age of Insufficiency</th>
            </tr>
        </thead>
        <tbody>
    '''
    
    i = 1
    for c in cases:
        check_types = ["Education Checks", "Family", "Reference Check", "Court", "Social Media", "Criminal", "Employment", "Identity Aadhar", "Address Check"]
        checks = []
        
        for check_type in check_types:
            docs = frappe.get_all(
                check_type,
                filters={"case_id": c.name, "check_status": "Insufficient Data"},
                fields=["name"]
            )
            checks.extend([doc.name for doc in docs])
        
        checks_str = ", ".join(checks)
        insuff_reported = formatdate(c.insufficiency_reported) if c.insufficiency_reported else ''
        
        data += f'''
        <tr>
            <td>{i}</td>
            <td>{insuff_reported}</td>
            <td>{c.batch}</td>
            <td>{c.name}</td>
            <td>{c.customer}</td>
            <td>{c.case_name}</td>
            <td>{c.client_employee_code or "-"}</td>
            <td>{c.case_status}</td>
            <td>{checks_str}</td>
            <td>{c.insufficiency_days or "-"}</td>
        </tr>
        '''
        i += 1
    
    
    data += '''
        </tbody>
    </table>
    '''
    
    frappe.sendmail(
        # recipients=['divya.p@groupteampro.com'],
        # recipients="siva.m@groupteampro.com",
        recipients=['dineshbabu.k@groupteampro.com','sangeetha.s@groupteampro.com','sangeetha.a@groupteampro.com',"keerthana.b@groupteampro.com"],
        cc=[''],
        subject=_("Cases with Insuff"),
        message=f"""
            Dear Madam,<br>Kindly find the below list of cases that are in Insuff status:<br>{data}<br><br>
            Thanks & Regards,<br>TEAMPRO<br>"This email has been automatically generated. Please do not reply"<br><br>"initiate further action and intimate a direct manager through email."
        """
    )
    print(i)



@frappe.whitelist()
def cases_with_insuff_daily_report():
    customers_with_cases = frappe.get_all("Case", {
        "case_status": ("in", ['Execution-Insuff', 'Entry-Insuff'])
    }, ["customer"], distinct=True, pluck="customer")

    for customer in customers_with_cases:
        cases = frappe.get_all("Case", {
            "customer": customer,
            "case_status": ("in", ['Execution-Insuff', 'Entry-Insuff'])
        }, ['*'])

        if cases:
            cust_mail=''
            batch=''
            cs=''
            data = '<table border="1" width="100%" style="border-collapse: collapse;">'
            data += '<tr style="background-color: #009dd1;"><td width=5% >S.No</td><td width=10% >Insuff Reported On</td><td width=15% >Batch</td><td width=15% >Case ID</td><td width=10% >Customer</td><td width=10% >Employee Name</td><td width=10% >Employee Code</td><td width=10% >Check Type</td><td width=10% >ID</td><td width=15% >Insuff Reported By</td><td width=20% >Remarks</td></tr>'
            ind = 0
            check_types = ["Education Checks", "Family", "Reference Check", "Court", "Social Media", "Criminal", "Employment", "Identity Aadhar", "Address Check"]
            for c in cases:
                batch = c.batch
                cs=c.name
                cust_mail=frappe.db.get_value("Batch",{"name":c.batch},['customer_mail_ids'])
                for check_type in check_types:
                    doc = frappe.get_all(check_type, {
                        "case_id": c.name,
                        "check_status": "Insufficient Data",
                        "insufficiency_date": frappe.utils.nowdate()
                    }, ["name", "workflow_state", "custom_insufficiency_reported_by", "insufficiency_date", "case_id", "batch", 'insufficient_remarks'])

                    for j in doc:
                        ind += 1
                        data += '<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (
                            ind, j.insufficiency_date or '', j.batch, j.case_id, c.customer, c.case_name,c.client_employee_code, check_type, j.name, j.custom_insufficiency_reported_by or '',  j.insufficient_remarks)

            data += '</table>'
            if ind>0:
                formatted_date = frappe.utils.format_datetime(frappe.utils.nowdate(), "dd-MMM-yyyy")
                frappe.sendmail(
                    # recipients=[cust_mail],
                    # recipients=["giftyannie6@gmail.com"],
                    recipients=['sangeetha.s@groupteampro.com',"sangeetha.a@groupteampro.com","keerthana.b@groupteampro.com"],  
                    subject=_("Insufficiency Report - Customer: %s - Date: %s" % (customer, formatted_date)),
                    message="""
                        Dear Sir/Madam,<br>Kindly Find the below List of Cases that are Reported as Insuff on Today for Customer %s %s<br>
                        Thanks & Regards,<br>TEAM ERP<br>"This email has been automatically generated. Please do not reply"
                    """ % (customer, data)
                )

@frappe.whitelist()
def dsr_mail():
    # current_date = datetime.now().date()
    # previous_day = current_date - timedelta(days=1)
    user = frappe.get_all("User", filters={"role": "BCS User",'enabled':1},  fields=["*"])
    # user = frappe.get_all("User",{"roles":"BCS User"},["*"])
    table = '<table  text-align: center; border="1" width="100%" style="border-collapse: collapse;"><tr><td style="width: 40%; font-weight: bold;">Executive</td><td style="width: 20%; font-weight: bold;">Total Allocated To</td><td style="width: 20%; font-weight: bold;">Completed by Today</td><td style="width: 20%; font-weight: bold;">Total Pending</td></tr> '
    for u in user:
        total_tasks = 0
        pending_tasks = 0
        completed_today = 0
        case=frappe.get_all("Case",{"case_status":("in",['Draft',"Entry Completed"]),"allocated_to":u.name},["*"])
        for c in case:
            total_tasks += 1
            pending_tasks += 1
            if c.date_of_entry_completion.date() == frappe.utils.nowdate():
                completed_today += 1
        list = ["Education Checks","Family","Reference Check","Court","Social Media","Criminal","Employment","Identity Aadhar","Address Check"]
        for i in list:
            doc=frappe.get_all(i,{"allocated_to":u.name},['*'])
            for j in doc:
                if j.check_status in ["Draft","Execution Pending"]:
                    total_tasks += 1
                    pending_tasks += 1
                    # table += '<td></td><td>{}</td><td></td><td>{}</td>'.format(total_tasks, pending_tasks)
                if str(j.date_of_entry_completion) == frappe.utils.nowdate() and j.entered_by == u.name:
                    # if j.date_of_entry_completion == nowdate():
                    # if j.date_of_entry_completion == frappe.utils.add_days(frappe.utils.nowdate(),-1):
                    completed_today += 1
                    total_tasks += 1
                        # table += '<td></td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(total_tasks, completed_today, pending_tasks)
                if str(j.date_of_execution_completion) == frappe.utils.nowdate() and j.execution_by == u.name:
                    # if j.date_of_execution_completion == nowdate():
                    # if j.date_of_execution_completion == frappe.utils.add_days(frappe.utils.nowdate(),-1):
                    completed_today += 1
                    total_tasks += 1
                        # table += '<td></td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(total_tasks, completed_today, pending_tasks)
        table += '<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (u.name, total_tasks, completed_today, pending_tasks)
    table += '</table>'
    frappe.sendmail(
        recipients=['sangeetha.s@groupteampro.com',"sangeetha.a@groupteampro.com","keerthana.b@groupteampro.com"],
        subject=_("DSR-%s"%(nowdate()) ),
        message="""
            Dear Sir/Madam,<br>Kindly Find the below attached DSR - %s<br>
            Thanks & Regards,<br>TEAM ERP<br>"This email has been automatically generated. Please do not reply"
            """%(table)
    )
    return "ok"     


from datetime import datetime, timedelta
from frappe.utils import add_days
from frappe import _
from datetime import datetime, timedelta

@frappe.whitelist()
def submitted_bg_entry():
    data = '<table  text-align: center; border="1" width="100%" style="border-collapse: collapse;text-align: center;">'
    data += '<tr style="font-weight: bold;background-color: #009dd1;"><td width=15%>ID</td><td width=25%>Name</td><td width=15%>DOB</td><td width=25%>Case Type</td><td width=25%>Status</td><td width=25%>Submitted Date</td><td width=25%>Submitted Time</td></tr>'
    today = datetime.now().date()
    prev_date = today - timedelta(days=1)
    start_time = datetime.combine(prev_date, datetime.min.time()) + timedelta(hours=18) 
    end_time = datetime.combine(today, datetime.min.time()) + timedelta(hours=18)   
    saved = frappe.db.sql("""
        SELECT * 
        FROM `tabBG Entry Form` 
        WHERE modified BETWEEN %s AND %s and docstatus = 1 order by experience DESC
    """, (start_time, end_time), as_dict=True)
    ind=0
    print(saved)
    for i in saved:
        print("hi")
        ind+=1
        modified_date = i.modified.date()
        modified_time = i.modified.strftime("%H:%M:%S")
        data += '<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>Submitted</td><td>%s</td><td>%s</td></tr>' % (i.name,i.employee_name,i.date_of_birth, i.experience, modified_date,modified_time)
    data += '</table>'
    if ind >= 1:   
        frappe.sendmail(
            # recipients=['divya.p@groupteampro.com'],
            recipients=['sangeetha.s@groupteampro.com','hrops@kblservices.in',"sangeetha.a@groupteampro.com","keerthana.b@groupteampro.com"],
            subject=_("KBL New Cases"),
            message="""
                Dear Sir/Madam,<br>
                Kindly Find the below attached KBL New Cases  %s<br>
                Thanks & Regards,<br>
                TEAM ERP<br>
                "This email has been automatically generated. Please do not reply"
            """ % data
        )
    else:
        frappe.sendmail(
            # recipients=['divya.p@groupteampro.com'],
            recipients=['sangeetha.s@groupteampro.com','hrops@kblservices.in',"sangeetha.a@groupteampro.com","keerthana.b@groupteampro.com"],
            subject=_("KBL New Cases"),
            message="""
                Dear Sir/Madam,<br>
                No New Cases has been Submitted today  %s<br>
                Thanks & Regards,<br>
                TEAM ERP<br>
                "This email has been automatically generated. Please do not reply"
            """ % today
        )




@frappe.whitelist()
def insuff_consolidated_mail():
    user = frappe.get_all("User", filters={"role": "BCS User"},  fields=["*"])
    for u in user:
        ind=0
        table = '<table  text-align: center; border="1" width="100%" style="border-collapse: collapse;"><tr><td style="width: 15%; font-weight: bold;">ID</td><td style="width: 15%; font-weight: bold;">Batch</td><td style="width: 15%; font-weight: bold;">Employee Name</td><td style="width: 10%; font-weight: bold;">Employee Code</td><td style="width: 20%; font-weight: bold;">Client</td><td style="width: 10%; font-weight: bold;">Case ID</td><td style="width: 10%; font-weight: bold;">Check Type</td><td style="width: 10%; font-weight: bold;">Check ID</td><td style="width: 10%; font-weight: bold;">Check Status</td><td style="width: 10%; font-weight: bold;">Actual Age</td><td style="width: 20%; font-weight: bold;">Allocated To</td></tr> '
        list = ["Education Checks","Family","Reference Check","Court","Social Media","Criminal","Employment","Identity Aadhar","Address Check"]
        for i in list:
            doc=frappe.get_all(i,{"allocated_to":u.name},['*'],order_by='actual_tat DESC')
            for j in doc:
                if j.clear_insufficiency:
                    if j.clear_insufficiency.strftime('%Y-%m-%d') == today():

                        ind+=1
                        if i == "Address Check":
                            table += '<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (j.name, j.batch,j.name1,j.client_employee_code,j.client,j.case_id,i,j.name,j.check_status,j.actual_tat, u.name)
                        else:
                            table += '<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (j.name, j.batch,j.name1,j.client_employee_code,j.customer,j.case_id,i,j.name,j.check_status,j.actual_tat, u.name)
        table += '</table>'
        if ind>0:
            frappe.sendmail(
                # recipients=['giftyannie6@gmail.com'],
                recipients=['sangeetha.s@groupteampro.com',u.name,"sangeetha.a@groupteampro.com","keerthana.b@groupteampro.com"],
                subject=_("Insuff Cleared-%s"%(nowdate()) ),
                message="""
                    Dear Sir/Madam,<br>Kindly Find the below attached List of Insuff Cleared Checks, %s<br>
                    Thanks & Regards,<br>TEAM ERP<br>"This email has been automatically generated. Please do not reply"
                    """%(table)
            )
    return "ok"


@frappe.whitelist()
def cases_with_gr_daily_report():
    data = '<table border="1" width="100%" style="border-collapse: collapse;">'
    data += '<tr style="background-color: #009dd1;"><td width=5% >S.No</td><td width=70% >Customer</td><td width=25% >Generate Report Count</td></tr>'
    customers_with_cases = frappe.get_all("Case", {
        "case_status": 'Generate Report'
    }, ["customer"], distinct=True, pluck="customer")
    ind = 0
    for customer in customers_with_cases:
        count=0
        ind += 1
        cases = frappe.get_all("Case", {
            "customer": customer,
            "case_status": 'Generate Report'
        }, ['*'])
        if cases:
            for c in cases:
                count+=1			
        data += '<tr><td>%s</td><td>%s</td><td>%s</td></tr>' % (
                ind, customer, count)
    data += '</table>'	
    formatted_date = frappe.utils.format_datetime(frappe.utils.nowdate(), "dd-MMM-yyyy")
    frappe.sendmail(
        # recipients=["giftyannie6@gmail.com"],
        recipients=['sangeetha.s@groupteampro.com',"sangeetha.a@groupteampro.com","keerthana.b@groupteampro.com"],  
        subject=_("Customer-Wise Generate Report Count - %s" % (formatted_date)),
        message="""
            Dear Sir/Madam,<br>Kindly Find the below attached Customer-Wise Generate Report Count %s<br>
            Thanks & Regards,<br>TEAM ERP<br>"This email has been automatically generated. Please do not reply"
        """ % (data)
    )



import openpyxl
from io import BytesIO
@frappe.whitelist()
def dpr_excel_format_bcs():
    filename = "DPR_" + today()
    users = frappe.get_all("User", filters={"role": "BCS User","enabled":1,"name":"thelothamma.r@groupteampro.com"}, fields=["*"])
    for user in users:
        email = user.name
        xlsx_file = build_xlsx_response_file(filename,user.name)
        send_mail_with_dpr_attachment(email, filename, xlsx_file.getvalue())


def send_mail_with_dpr_attachment(recipient, filename, file_content):
    subject = ("DPR-%s-%s"%(nowdate(),recipient) )
    message = "Dear Sir/Madam,<br> Please find attached the Daily Progress Report.<br>Thanks & Regards,<br>TEAM ERP<br>This email has been automatically generated. Please do not reply"
    attachments = [{"fname": filename + '.xlsx', "fcontent": file_content}]
    frappe.sendmail(
        recipients=[recipient],
        # recipients="divya.p@groupteampro.com",
        cc=['sangeetha.s@gmail.com',"sangeetha.a@groupteampro.com","keerthana.b@groupteampro.com"],
        sender=None,  
        subject=subject,
        message=message,
        attachments=attachments,
    )


def build_xlsx_response_file(filename,user_name):
    xlsx_file = make_xlsx_file(filename,user_name)
    return xlsx_file

def make_xlsx_file(filename, user_name, sheet_name=None, wb=None, column_widths=None):
    from collections import defaultdict

    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
        # Remove the default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

    # Sheet 1: Main DPR
    ws1 = wb.create_sheet(title="DPR", index=0)
    ws1.append(["ID", "Batch", "Employee Name", "Employee Code", 'Client', 'Case ID', 'Case/Check Type',
                'Case/Check Status', 'Actual Age', 'Allocated To', 'Entry Allocated Date', 'Execution Allocated Date'])
    # Dict to track summary counts by Check Type only
    summary_counts = defaultdict(int)

    # Add Cases
    cases = frappe.get_all("Case", {"case_status": "Draft", "allocated_to": user_name},
                           ['*'], order_by='actual_tat DESC')
    for c in cases:
        ws1.append([
            c.name, c.batch, c.case_name, c.client_employee_code, c.customer, c.name,
            "Case", c.case_status, c.actual_tat, user_name,
            c.custom_allocation_date or '', ''
        ])
        summary_counts["Case"] += 1
    case_id_set = set()
    # Add Checks
    check_types = ["Education Checks", "Family", "Reference Check", "Court", "Social Media", "Criminal", "Employment", "Identity Aadhar", "Address Check"]
    for check in check_types:
        docs = frappe.get_all(check, {"allocated_to": user_name}, ['*'], order_by='actual_tat DESC')
        for d in docs:
            if d.check_status in ["Draft", "Entry QC Completed", "Execution Pending", "Execution Initiated"]:
                row = [
                    d.name, d.batch, d.name1, d.client_employee_code,
                    d.client if check in ["Address Check", "Court", "Employment", "Criminal", "Social Media", "Family"] else d.customer,
                    d.case_id, check, d.check_status, d.actual_tat, user_name,
                    d.custom_allocation_date or '', d.custom_date_of_execution_initiated or ''
                ]
                ws1.append(row)
                summary_counts[check] += 1
                if d.case_id:
                    case_id_set.add(d.case_id)

    # Sheet 2: Horizontal Summary
    ws2 = wb.create_sheet(title="Summary", index=1)
    ws2.append([])
    # Check types to include (as-is)
    check_types = ["Case","Education Checks", "Family", "Reference Check", "Court", "Social Media", "Criminal", "Employment", "Identity Aadhar", "Address Check"]

    # Add header row with original names
    # header_row = ["User"]+check_types
    # Mapping original keys to new column headers
    check_type_labels = {
        "Case": "Draft",
        "Education Checks": "Education Checks",
        "Criminal": "Criminal",
        "Employment": "Employment",
        "Identity Aadhar": "Identity Aadhar",
        "Address Check": "Address Check",
        "Family":"Family",
        "Reference Check":"Reference Check",
         "Court": "Court",
         "Social Media":"Social Media",

    }
    check_types = list(check_type_labels.keys())
    header_row = ["User"] + [check_type_labels[ct] for ct in check_types]

    ws2.append(header_row)

    # Add data row: user name + counts
    summary_row = [user_name] +[summary_counts.get(ct, 0) for ct in check_types]
    ws2.append(summary_row)

    # Apply formatting: bold header, blue fill
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    bold_font = Font(bold=True)

    for cell in ws2[2]:  # first row (headers)
        cell.font = bold_font
        cell.fill = header_fill
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file




@frappe.whitelist()
def cases_with_generate_report_status():
    cases = frappe.get_all("Case", {"case_status": "Generate Report"}, ['*'])

    if cases:
        data = '<table border="1" width="100%" style="border-collapse: collapse;">'
        data += '<tr style="background-color: #009dd1;"><td width=5% >S.No</td><td width=15% >Batch</td><td width=15% >Case ID</td><td width=10% >Customer</td><td width=10% >Employee Name</td><td width=10% >Employee Code</td></tr>'
        ind = 0
        for c in cases:	
            ind += 1
            data += '<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (
                ind, c.batch, c.name, c.customer, c.case_name,c.client_employee_code)

        data += '</table>'
        if ind>0:
            formatted_date = frappe.utils.format_datetime(frappe.utils.nowdate(), "dd-MMM-yyyy")
            frappe.sendmail(
                # recipients=[cust_mail],
                # recipients=["giftyannie6@gmail.com"],
                recipients=['sangeetha.s@groupteampro.com',c.allocated_to_batch_manager,"sangeetha.a@groupteampro.com","keerthana.b@groupteampro.com"],  
                subject=_("Cases in Generate Report- Date: %s" % ( formatted_date)),
                message="""
                    Dear Sir/Madam,<br>Kindly Find the below List of Cases that are in "Generate Report" Status %s<br>
                    Thanks & Regards,<br>TEAM ERP<br>"This email has been automatically generated. Please do not reply"
                """ % (data)
            )

@frappe.whitelist()
def cases_with_to_be_billed_status():
    cases = frappe.get_all("Case", {"case_status": "To be Billed"}, ['*'])

    if cases:
        data = '<table border="1" width="100%" style="border-collapse: collapse;">'
        data += '<tr style="background-color: #009dd1;"><td width=5% >S.No</td><td width=15% >Batch</td><td width=15% >Case ID</td><td width=10% >Customer</td><td width=10% >Employee Name</td><td width=10% >Employee Code</td></tr>'
        ind = 0
        for c in cases:	
            ind += 1
            data += '<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (
                ind, c.batch, c.name, c.customer, c.case_name,c.client_employee_code)

        data += '</table>'
        if ind>0:
            formatted_date = frappe.utils.format_datetime(frappe.utils.nowdate(), "dd-MMM-yyyy")
            frappe.sendmail(
                # recipients=[cust_mail],
                # recipients=["giftyannie6@gmail.com"],
                recipients=['sangeetha.s@groupteampro.com',c.allocated_to_batch_manager,"sangeetha.a@groupteampro.com"],  
                subject=_("Cases in To be Billed- Date: %s" % ( formatted_date)),
                message="""
                    Dear Sir/Madam,<br>Kindly Find the below List of Cases that are in "To be Billed" Status %s<br>
                    Thanks & Regards,<br>TEAM ERP<br>"This email has been automatically generated. Please do not reply"
                """ % (data)
            )

@frappe.whitelist()
def delete_batch_projects():
    projects=frappe.get_all("Project",{'project_type':"BCS","service":"BCS"},['*'])
    for p in projects:
        frappe.db.sql("""delete from `tabProject` where project_type = 'BCS'""",as_dict = True)
        print(p.project_type)

import frappe

import frappe

@frappe.whitelist()
def delete_batch_projects():
    projects = frappe.get_all("Project", {'project_type': "BCS", 'service': "BCS"}, ['*'])
    for p in projects:
        frappe.db.sql("""delete from `tabProject` where project_type = 'BCS'""", as_dict=True)
        print(p.project_type)

import frappe

@frappe.whitelist()
def delete_batch_projects():
    projects = frappe.get_all("Project", {'project_type': "BCS", 'service': "BCS"}, ['*'])
    for p in projects:
        frappe.db.sql("""delete from `tabProject` where project_type = 'BCS'""", as_dict=True)
        print(p.project_type)

@frappe.whitelist()
def gl_report(entry_report):
    data= ""
    data= ""
    data += '<table class="table table-bordered"><tr><th style="padding:1px;border: 1px solid black;color:black;background-color:lightblue" colspan=7><center>General Ledger</center></th></tr>'
    data += '''
    <tr>
    <td  style="padding:1px;border: 1px solid black" ><b>Posting Date</b></td>
    <td style="padding:1px;border: 1px solid black" colspan =1><b>Account</b></td>
    <td style="padding:1px;border: 1px solid black" colspan=1><b>Debit(INR)</b></td>
    <td style="padding:1px;border: 1px solid black" colspan=1><b>Credit(INR)</b></td>
    <td style="padding:1px;border: 1px solid black" colspan=1><b>Balance(INR)</b></td>
    <td style="padding:1px;border: 1px solid black" colspan=1><b>Vocher Type</b></td>
    <td style="padding:1px;border: 1px solid black" colspan=1><b>Vocher NO</b>
    </td></tr>'''
    sq = frappe.db.sql(""" select * from `tabGL Entry` where account='%s' and posting_date between '%s' and '%s' and is_opening = 'No' and is_cancelled = 0 """%(entry_report,today(),today()),as_dict=True)
    for i in sq:
        balance=i.debit-i.credit
        data += '''<tr>
            <td style="padding:1px;border: 1px solid black" colspan =1>%s</td>
            <td style="padding:1px;border: 1px solid black" colspan=1>%s</td>
            <td style="padding:1px;border: 1px solid black" colspan=1 >%s</td>
            <td style="padding:1px;border: 1px solid black" colspan=1>%s</td>
            <td style="padding:1px;border: 1px solid black" colspan=1>%s</td>
            <td style="padding:1px;border: 1px solid black" colspan=1 >%s</td>
            <td style="padding:1px;border: 1px solid black" colspan=1>%s</td>
            </tr>'''%(i.posting_date,i.account,i.debit,i.credit,balance,i.voucher_type,i.voucher_no)

    data += '</table>'
    return data


# Daily Transaction Report For New Correction
@frappe.whitelist()
def statement_of_account_test_1():
    data = """
    <style>
        .responsive-table {
            width: 100%;
            border-collapse: collapse;
        }
        .responsive-table th, .responsive-table td {
            padding: 8px;
            text-align: center;
        }
        .responsive-table th {
            background-color: #063970;
            color: white;
        }
        .responsive-table td.account {
            text-align: left;
        }
        .company-header {
            text-align: center;
            font-weight: bold;
        }
        @media (max-width: 600px) {
            .responsive-table thead {
                display: none;
            }
            .responsive-table, .responsive-table tbody, .responsive-table tr, .responsive-table td {
                display: block;
                width: 100%;
            }
            .responsive-table tr {
                margin-bottom: 15px;
            }
            .responsive-table td {
                text-align: right;
                padding-left: 50%;
                position: relative;
            }
            .responsive-table td::before {
                content: attr(data-label);
                position: absolute;
                left: 0;
                width: 50%;
                padding-left: 15px;
                text-align: left;
                font-weight: bold;
            }
        }
    </style>
    """
    company_order = [
        'TEAMPRO HR & IT Services Pvt. Ltd.',
        'TEAMPRO General Trading Pvt. Ltd.',
        'TEAMPRO General Trading',
        'TEAMPRO Food Products'
    ]

    company = frappe.db.get_all('Company', {'name': ('Not in', ['TEAMPRO Saudi Arabia'])}, ['*'])
    company_dict = {c.name: c for c in company}

    for company_name in company_order:
        j = company_dict.get(company_name)
        if not j:
            continue

        accounts = []

        if j.name == 'TEAMPRO HR & IT Services Pvt. Ltd.':
            accounts = ['50200054611436 - HDFC - THIS', '777705160983 - ICICI Bank - THIS', 'Cash - THIS']
        elif j.name == 'TEAMPRO General Trading Pvt. Ltd.':
            accounts = ['777705755022 - ICICI Bank - TGTP', 'Cash - TGTP']
        elif j.name == 'TEAMPRO General Trading':
            accounts = ['50200050787897 - HDFC Account - TGT', 'Cash - TGT']
        elif j.name == 'TEAMPRO Food Products':
            accounts = ['50200059117831 - HDFC Bank - TFP', 'Cash - TFP']

        data += f"<br><table class='responsive-table' border=1 style='margin:2px;'><tr class='company-header' style='text-align:center;font-size:10px;background-color:#063970;color:#FFFFFF;'><td width='100%'><b>{j.name}</b></td></tr></table>"
        data += "<table class='responsive-table' border=1 style='margin:2px;'><thead><tr style='font-size:10px;background-color:#063970;color:#FFFFFF;'><th width='10%'><b>Posting Date</b></th><th width='10%'><b>Voucher Type</b></th><th width='10%'><b>Voucher No</b></th><th width='30%'><b>Against Account</b></th><th width='10%'><b>Debit (INR)</b></th><th width='10%'><b>Credit (INR)</b></th><th width='10%'><b>Balance (INR)</b></th></tr></thead><tbody>"

        today_date = frappe.utils.now_datetime().date()

        for a in accounts:
            data += f'<tr style="font-size:10px"><td class="account" colspan=7><b>{a}</b></td></tr>'

            gl_entry = frappe.db.sql("""
                select voucher_type, voucher_no, posting_date, sum(debit) as debit, sum(credit) as credit, account, against
                from `tabGL Entry`
                where account=%s and posting_date=%s and is_cancelled = 0 and company=%s
                group by voucher_type, voucher_no, posting_date, account, against
                order by posting_date
            """, (a, today_date, j.name), as_dict=True)

            gle = frappe.db.sql("""
                select sum(debit) as opening_debit, sum(credit) as opening_credit
                from `tabGL Entry`
                where account=%s and posting_date < %s and is_cancelled = 0 and company=%s
            """, (a, today_date, j.name), as_dict=True)

            opening_balance = round((gle[0].opening_debit or 0) - (gle[0].opening_credit or 0), 2)
            data += f'<tr style="font-size:10px"><td colspan=6 style="text-align:right" data-label="Opening Balance"><b>Opening Balance</b></td><td style="text-align:right" data-label="Opening Balance"><b>{opening_balance}</b></td></tr>'

            balance = opening_balance
            total_debit = 0
            total_credit = 0

            for entry in gl_entry:
                posting_date = entry.posting_date.strftime("%d-%m-%Y") if entry.posting_date else "-"
                debit = round(entry.debit or 0, 2)
                credit = round(entry.credit or 0, 2)
                balance += debit - credit

                data += f'<tr style="font-size:10px"><td data-label="Posting Date">{posting_date}</td><td data-label="Voucher Type">{entry.voucher_type or "-"}</td><td data-label="Voucher No">{entry.voucher_no or "-"}</td><td data-label="Against Account">{entry.against or "-"}</td><td style="text-align:right" data-label="Debit (INR)">{debit}</td><td style="text-align:right" data-label="Credit (INR)">{credit}</td><td style="text-align:right" data-label="Balance (INR)">{round(balance, 2)}</td></tr>'

                total_debit += debit
                total_credit += credit

            total_balance = round(balance, 2)
            data += f'<tr style="font-size:10px"><td colspan=4 style="text-align:right" data-label="Total"><b>Total</b></td><td style="text-align:right" data-label="Total Debit"><b>{round(total_debit, 2)}</b></td><td style="text-align:right" data-label="Total Credit"><b>{round(total_credit, 2)}</b></td><td></td></tr>'
            data += f'<tr style="font-size:10px"><td colspan=6 style="text-align:right" data-label="Closing Balance"><b>Closing Balance</b></td><td style="text-align:right" data-label="Closing Balance"><b>{total_balance}</b></td></tr>'

        data += '</tbody></table><br><br>'


    frappe.sendmail(
        recipients=['dineshbabu.k@groupteampro.com'],
        cc=['sangeetha.a@groupteampro.com', 'sangeetha.s@groupteampro.com', 'accounts@groupteampro.com'],
        # recipients='divya.p@groupteampro.com',
        subject='Daily Transaction Report',
        message=f"""
            Dear Sir,<br>
            <p>Please find the enclosed details for your reference. Kindly check the Daily Transaction Report</p>
            {data}
            "This email has been automatically generated. PLEASE DONOT REPLY, Initiate further action and intimate your direct manager through email."
            <br><br>
            "With Best Wishes & Regards "
            <br><br>
            <span style="color:#203ed5;">
            "TEN – Auto Mail "
            </span>
            <br><br>
            <span style="color:#203ed5;">
                "Disclaimers:<br>
                This email and any files transmitted with it are confidential and intended solely for the use of the individual or entity to whom they are addressed. If you have received this email in error please notify the system manager. Please note that any views or opinions presented in this email are solely those of the author and do not necessarily represent those of the company. Finally, the recipient should check this email and any attachments for the presence of viruses. The company accepts no liability for any damage caused by any virus transmitted by this email."
            </span>
        """
    )


# #Sales Order Follow Up New Correction

from frappe.utils import date_diff
from frappe import _

@frappe.whitelist()
def sales_order_follow_up():
    sales_orders = frappe.get_list(
        "Sales Order",
        filters={"status": ["not in", ["Hold", "To Deliver", "Closed", "Cancelled", "Completed"]]},
        fields=["name", "account_manager", "service", "status", "customer", "company", "transaction_date", "base_grand_total", "per_billed", "advance_paid"],
        order_by='customer asc'  
    )
    
    additional_table = '<br><br><table border=1><tr><td style="background-color:#063970;color:white">S.No</td><td style="background-color:#063970;color:white">Service</td><td style="background-color:#063970;color:white">Grand Total</td><td style="background-color:#063970;color:white">Outstanding</td></tr>'
    tfp = '<table border=1><tr style="text-align: center"><td style="background-color:#063970;color:white">ID</td><td style="background-color:#063970;color:white">Account Manager</td><td style="background-color:#063970;color:white">Service</td><td style="background-color:#063970;color:white">Status</td><td style="background-color:#063970;color:white">Customer Name</td><td style="background-color:#063970;color:white">Company</td><td style="background-color:#063970;color:white">Date</td><td style="background-color:#063970;color:white">Grand Total</td><td style="background-color:#063970;color:white">% Amount Billed</td><td style="background-color:#063970;color:white">Advance Paid</td><td style="background-color:#063970;color:white">To Be Billed</td></tr>'
    
    total_amount = 0
    grand_total = 0
    serial_number = 1  

    for j in sales_orders:
        if j.get('service') == 'TFP':
            to_be_billed = j.get('base_grand_total') - (j.get('advance_paid') + ((j.get('per_billed') / 100) * j.get('base_grand_total')))
            total_amount += to_be_billed
            grand_total += j.get('base_grand_total')
            formatted_date = j.get("transaction_date").strftime('%d-%m-%Y')
            tfp += '<tr style="font-size:14px"><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td style="text-align:left;">{}</td><td style="text-align:right;">{}</td><td style="text-align:right;">{}</td><td style="text-align:right;">{}</td><td style="text-align:right;">{}</td><td style="text-align:right;">{}</td></tr>'.format(
               j.get('name'), j.get('account_manager'), j.get('service'), j.get('status'), j.get('customer'), j.get('company'), formatted_date, "{:,.0f}".format(j.get('base_grand_total')), "{:,.0f}".format(j.get('per_billed')), "{:,.0f}".format(j.get('advance_paid')), "{:,.0f}".format(to_be_billed))
    
    additional_table += '<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(serial_number, "TFP", "{:,.0f}".format(grand_total), "{:,.0f}".format(total_amount))
    
    tfp += '<tr><td style="text-align:center;" colspan=7>Total</td><td style="text-align:right">{}</td><td></td><td></td><td style="text-align:right;">{}</td></tr>'.format("{:,.0f}".format(grand_total), "{:,.0f}".format(total_amount))
    additional_table += '</table>'
    tfp += '</table>'

    frappe.sendmail(
        # recipients=["siva.m@groupteampro.com","accounts@groupteampro.com"],
        recipients='amirtham.g@groupteampro.com',
        cc=['sangeetha.s@groupteampro.com','dineshbabu.k@groupteampro.com','accounts@groupteampro.com'],
        subject='Sales Invoice Follow Up-Sales Order Outstanding',
        message="""
        Dear Mam,<br>
        <p>Collection Outstanding Report For Further Action.</p>
        TFP : SBMK/AM
        <br>
        {}
        <br>
        {}<br><br>
        "This email has been automatically generated. PLEASE DONOT REPLY, Initiate further action and intimate your direct manager through email."
            <br><br>
            "With Best Wishes & Regards "
            <br><br>
            <span style="color:#203ed5;">
            "TEN – Auto Mail "
            </span>
            <br><br>
            <span style="color:#203ed5;">
                "Disclaimers:<br>
                This email and any files transmitted with it are confidential and intended solely for the use of the individual or entity to whom they are addressed. If you have received this email in error please notify the system manager. Please note that any views or opinions presented in this email are solely those of the author and do not necessarily represent those of the company. Finally, the recipient should check this email and any attachments for the presence of viruses. The company accepts no liability for any damage caused by any virus transmitted by this email."
            </span>
        """.format(additional_table, tfp)
    )



    additional_table = '<br><br><table border=1><tr><td style="background-color:#063970;color:white">S.No</td><td style="background-color:#063970;color:white">Service</td><td style="background-color:#063970;color:white">Grand Total</td><td style="background-color:#063970;color:white">Outstanding</td></tr>'
    bcs = '<table border=1><tr style="text-align: center"><td style="background-color:#063970;color:white">ID</td><td style="background-color:#063970;color:white">Account Manager</td><td style="background-color:#063970;color:white">Service</td><td style="background-color:#063970;color:white">Status</td><td style="background-color:#063970;color:white">Customer Name</td><td style="background-color:#063970;color:white">Company</td><td style="background-color:#063970;color:white">Date</td><td style="background-color:#063970;color:white">Grand Total</td><td style="background-color:#063970;color:white">% Amount Billed</td><td style="background-color:#063970;color:white">Advance Paid</td><td style="background-color:#063970;color:white">To Be Billed</td></tr>'
    
    total_amount_bcs = 0
    grand_total_bcs = 0
    serial_number = 1  
    
    for i in sales_orders:
        if i.get('service') == 'BCS':
            to_be_billed = i.get('base_grand_total') - (i.get('advance_paid') + ((i.get('per_billed') / 100) * i.get('base_grand_total')))
            total_amount_bcs += to_be_billed
            grand_total_bcs += i.get('base_grand_total')
            formatted_date = i.get("transaction_date").strftime('%d-%m-%Y')
            bcs += '<tr style="font-size:14px"><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td style="text-align:left;">{}</td><td style="text-align:right;">{}</td><td style="text-align:right;">{}</td><td style="text-align:right;">{}</td><td style="text-align:right;">{}</td><td style="text-align:right;">{}</td></tr>'.format(
                i.get('name'), i.get('account_manager'), i.get('service'), i.get('status'), i.get('customer'), i.get('company'), formatted_date, "{:,.0f}".format(i.get('base_grand_total')), "{:,.0f}".format(i.get('per_billed')), "{:,.0f}".format(i.get('advance_paid')), "{:,.0f}".format(to_be_billed))
    
    additional_table += '<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(serial_number, "BCS", "{:,.0f}".format(grand_total_bcs), "{:,.0f}".format(total_amount_bcs))
    
    bcs += '<tr><td style="text-align:center;" colspan=7>Total</td><td style="text-align:right">{}</td><td></td><td></td><td style="text-align:right;">{}</td></tr>'.format("{:,.0f}".format(grand_total_bcs), "{:,.0f}".format(total_amount_bcs))
    additional_table += '</table>'
    bcs += '</table>'

    frappe.sendmail(
        # recipients=["siva.m@groupteampro.com","accounts@groupteampro.com"],
        recipients=['sangeetha.a@groupteampro.com'],
        cc=['dineshbabu.k@groupteampro.com', 'accounts@groupteampro.com', 'sangeetha.s@groupteampro.com'],
        subject='Sales Invoice Follow Up-Sales Order Outstanding',
        message="""
        Dear Mam,<br>
        <p>Collection Outstanding Report For Further Action.</p>
        BCS : SBMK
        <br>
        {}
        <br>
        {}<br><br>
        "This email has been automatically generated. PLEASE DONOT REPLY, Initiate further action and intimate your direct manager through email."
            <br><br>
            "With Best Wishes & Regards "
            <br><br>
            <span style="color:#203ed5;">
            "TEN – Auto Mail "
            </span>
            <br><br>
            <span style="color:#203ed5;">
                "Disclaimers:<br>
                This email and any files transmitted with it are confidential and intended solely for the use of the individual or entity to whom they are addressed. If you have received this email in error please notify the system manager. Please note that any views or opinions presented in this email are solely those of the author and do not necessarily represent those of the company. Finally, the recipient should check this email and any attachments for the presence of viruses. The company accepts no liability for any damage caused by any virus transmitted by this email."
            </span>
        """.format(additional_table, bcs)
    )

    additional_table = '<br><br><table border=1><tr><td style="background-color:#063970;color:white">S.No</td><td style="background-color:#063970;color:white">Service</td><td style="background-color:#063970;color:white">Grand Total</td><td style="background-color:#063970;color:white">Outstanding</td></tr>'
    rec = '<table border=1><tr style="text-align: center"><td style="background-color:#063970;color:white">ID</td><td style="background-color:#063970;color:white">Account Manager</td><td style="background-color:#063970;color:white">Service</td><td style="background-color:#063970;color:white">Status</td><td style="background-color:#063970;color:white">Customer Name</td><td style="background-color:#063970;color:white">Company</td><td style="background-color:#063970;color:white">Date</td><td style="background-color:#063970;color:white">Grand Total</td><td style="background-color:#063970;color:white">% Amount Billed</td><td style="background-color:#063970;color:white">Advance Paid</td><td style="background-color:#063970;color:white">To Be Billed</td></tr>'
    
    total_amount_rec = 0
    grand_total_rec = 0
    serial_number = 1  
    
    for k in sales_orders:
        if k.get('service') in ['REC-I', 'REC-D']:
            to_be_billed = k.get('base_grand_total') - (k.get('advance_paid') + ((k.get('per_billed') / 100) * k.get('base_grand_total')))
            total_amount_rec += to_be_billed
            grand_total_rec += k.get('base_grand_total')
            formatted_date = k.get("transaction_date").strftime('%d-%m-%Y')
            rec += '<tr style="font-size:14px"><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td style="text-align:left;">{}</td><td style="text-align:right;" nowrap>{}</td><td style="text-align:right;">{}</td><td style="text-align:right;">{}</td><td style="text-align:right;">{}</td><td style="text-align:right;">{}</td></tr>'.format(
               k.get('name'), k.get('account_manager'), k.get('service'), k.get('status'), k.get('customer'), k.get('company'), formatted_date, "{:,.0f}".format(k.get('base_grand_total')), "{:,.0f}".format(k.get('per_billed')), "{:,.0f}".format(k.get('advance_paid')), "{:,.0f}".format(to_be_billed))
    
    additional_table += '<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(serial_number, "REC", "{:,.0f}".format(grand_total_rec), "{:,.0f}".format(total_amount_rec))
    
    rec += '<tr><td style="text-align:center;" colspan=7>Total</td><td style="text-align:right">{}</td><td></td><td></td><td style="text-align:right;">{}</td></tr>'.format("{:,.0f}".format(grand_total_rec), "{:,.0f}".format(total_amount_rec))
    additional_table += '</table>'
    rec += '</table>'

    frappe.sendmail(
        # recipients=["siva.m@groupteampro.com","accounts@groupteampro.com"],
        recipients=['sangeetha.a@groupteampro.com'],
        cc=['dineshbabu.k@groupteampro.com', 'sangeetha.s@groupteampro.com', 'accounts@groupteampro.com', 'annie.m@groupteampro.com'],
        subject='Sales Invoice Follow Up-Sales Order Outstanding',
        message="""
        Dear Mam,<br>
        <p>Collection Outstanding Report For Further Action.</p>
        REC : AS/AM
        <br>
        {}
        <br>
        {}<br><br>
        "This email has been automatically generated. PLEASE DONOT REPLY, Initiate further action and intimate your direct manager through email."
            <br><br>
            "With Best Wishes & Regards "
            <br><br>
            <span style="color:#203ed5;">
            "TEN – Auto Mail "
            </span>
            <br><br>
            <span style="color:#203ed5;">
                "Disclaimers:<br>
                This email and any files transmitted with it are confidential and intended solely for the use of the individual or entity to whom they are addressed. If you have received this email in error please notify the system manager. Please note that any views or opinions presented in this email are solely those of the author and do not necessarily represent those of the company. Finally, the recipient should check this email and any attachments for the presence of viruses. The company accepts no liability for any damage caused by any virus transmitted by this email."
            </span>
        """.format(additional_table, rec)
    )

    
    additional_table = '<br><br><table border=1><tr><td style="background-color:#063970;color:white">S.No</td><td style="background-color:#063970;color:white">Service</td><td style="background-color:#063970;color:white">Grand Total</td><td style="background-color:#063970;color:white">Outstanding</td></tr>'
    itsw = '<table border=1><tr style="text-align: center"><td style="background-color:#063970;color:white">ID</td><td style="background-color:#063970;color:white">Account Manager</td><td style="background-color:#063970;color:white">Service</td><td style="background-color:#063970;color:white">Status</td><td style="background-color:#063970;color:white">Customer Name</td><td style="background-color:#063970;color:white">Company</td><td style="background-color:#063970;color:white">Date</td><td style="background-color:#063970;color:white">Grand Total</td><td style="background-color:#063970;color:white">% Amount Billed</td><td style="background-color:#063970;color:white">Advance Paid</td><td style="background-color:#063970;color:white">To Be Billed</td></tr>'
    
    total_amount_itsw = 0
    grand_total_itsw = 0
    serial_number = 1  

    for i in sales_orders:
        if i.get('service') in ['IT-SW', 'IT-IS']:
            to_be_billed = i.get('base_grand_total') - (i.get('advance_paid') + ((i.get('per_billed') / 100) * i.get('base_grand_total')))
            total_amount_itsw += to_be_billed
            grand_total_itsw += i.get('base_grand_total')
            formatted_date = i.get("transaction_date").strftime('%d-%m-%Y')
            itsw += '<tr style="font-size:14px"><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td style="text-align:left;">{}</td><td style="text-align:right;" nowrap>{}</td><td style="text-align:right;">{}</td><td style="text-align:right;">{}</td><td style="text-align:right;">{}</td><td style="text-align:right;">{}</td></tr>'.format(
                i.get('name'), i.get('account_manager'), i.get('service'), i.get('status'), i.get('customer'), i.get('company'), formatted_date, "{:,.0f}".format(i.get('base_grand_total')), "{:,.0f}".format(i.get('per_billed')), "{:,.0f}".format(i.get('advance_paid')), "{:,.0f}".format(to_be_billed))
    
    additional_table += '<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(serial_number, "IT-SW/IT-IS", "{:,.0f}".format(grand_total_itsw), "{:,.0f}".format(total_amount_itsw))
    
    itsw += '<tr><td style="text-align:center;" colspan=7>Total</td><td style="text-align:right">{}</td><td></td><td></td><td style="text-align:right;">{}</td></tr>'.format("{:,.0f}".format(grand_total_itsw), "{:,.0f}".format(total_amount_itsw))
    additional_table += '</table>'
    itsw += '</table>'

    frappe.sendmail(
        # recipients=["siva.m@groupteampro.com","accounts@groupteampro.com"],
        recipients=['dineshbabu.k@groupteampro.com'],
        cc=[ 'sangeetha.s@groupteampro.com', 'accounts@groupteampro.com'],
        subject='Sales Invoice Follow Up-Sales Order Outstanding',
        message="""
        Dear Sir,<br>
        <p>Collection Outstanding Report For Further Action.</p>
        IT-SW/IT-IS : DKB/APP
        <br>
        {}
        <br>
        {}<br><br>
        "This email has been automatically generated. PLEASE DONOT REPLY, Initiate further action and intimate your direct manager through email."
            <br><br>
            "With Best Wishes & Regards "
            <br><br>
            <span style="color:#203ed5;">
            "TEN – Auto Mail "
            </span>
            <br><br>
            <span style="color:#203ed5;">
                "Disclaimers:<br>
                This email and any files transmitted with it are confidential and intended solely for the use of the individual or entity to whom they are addressed. If you have received this email in error please notify the system manager. Please note that any views or opinions presented in this email are solely those of the author and do not necessarily represent those of the company. Finally, the recipient should check this email and any attachments for the presence of viruses. The company accepts no liability for any damage caused by any virus transmitted by this email."
            </span>
        """.format(additional_table, itsw)
    )

    additional_table = '<br><br><table border=1><tr><td style="background-color:#063970;color:white">S.No</td><td style="background-color:#063970;color:white">Service</td><td style="background-color:#063970;color:white">Grand Total</td><td style="background-color:#063970;color:white">Outstanding</td></tr>'
    tgt = ''
    tgt += '<table border=1><tr style="text-align: center"><td style="background-color:#063970;color:white">ID</td><td style="background-color:#063970;color:white">Account Manager</td><td style="background-color:#063970;color:white">Service</td><td style="background-color:#063970;color:white">Status</td><td style="background-color:#063970;color:white">Customer Name</td><td style="background-color:#063970;color:white">Company</td><td style="background-color:#063970;color:white">Date</td><td style="background-color:#063970;color:white">Grand Total</td><td style="background-color:#063970;color:white">% Amount Billed</td><td style="background-color:#063970;color:white">Advance Paid</td><td style="background-color:#063970;color:white">To Be Billed</td></tr>'

    total_amount_tgt = 0
    grand_total_tgt = 0
    serial_number = 1 

    for i in sales_orders:
        if i.service == 'TGT':
            to_be_billed = i.base_grand_total - (i.advance_paid + ((i.per_billed / 100) * i.base_grand_total))
            total_amount_tgt += to_be_billed
            grand_total_tgt += i.base_grand_total
            formatted_date = i.transaction_date.strftime('%d-%m-%Y')
            tgt += '<tr style="font-size:14px"><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td style="text-align:left;">{}</td><td style="text-align:right;" nowrap>{}</td><td style="text-align:right;">{}</td><td style="text-align:right;">{}</td><td style="text-align:right;">{}</td><td style="text-align:right;">{}</td></tr>'.format(
                i.name, i.account_manager, i.service, i.status, i.customer, i.company, formatted_date, "{:,.0f}".format(i.base_grand_total), "{:,.0f}".format(i.per_billed), "{:,.0f}".format(i.advance_paid), "{:,.0f}".format(to_be_billed))

    additional_table += '<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(serial_number, "TGT", "{:,.0f}".format(grand_total_tgt), "{:,.0f}".format(total_amount_tgt))
    tgt += '<tr><td style="text-align:center;" colspan=7>Total</td><td style="text-align:right">{}</td><td></td><td></td><td style="text-align:right;">{}</td></tr>'.format("{:,.0f}".format(grand_total_tgt), "{:,.0f}".format(total_amount_tgt))
    additional_table += '</table>'
    tgt += '</table>'

    frappe.sendmail(
        # recipients=["siva.m@groupteampro.com","accounts@groupteampro.com"],
        recipients=['sangeetha.s@groupteampro.com'],
        cc=['dineshbabu.k@groupteampro.com', 'accounts@groupteampro.com'],
        subject='Sales Invoice Follow Up-Sales Order Outstanding',
        message="""
        Dear Mam,<br>
        <p>Collection Outstanding Report For Further Action.</p>
        TGT : SBMK
        <br>
        {}
        <br>
        {}<br><br>
        "This email has been automatically generated. PLEASE DONOT REPLY, Initiate further action and intimate your direct manager through email."
            <br><br>
            "With Best Wishes & Regards "
            <br><br>
            <span style="color:#203ed5;">
            "TEN – Auto Mail "
            </span>
            <br><br>
            <span style="color:#203ed5;">
                "Disclaimers:<br>
                This email and any files transmitted with it are confidential and intended solely for the use of the individual or entity to whom they are addressed. If you have received this email in error please notify the system manager. Please note that any views or opinions presented in this email are solely those of the author and do not necessarily represent those of the company. Finally, the recipient should check this email and any attachments for the presence of viruses. The company accepts no liability for any damage caused by any virus transmitted by this email."
            </span>
        """.format(additional_table, tgt)
    )

# Overall Service Report With Excel Sheet Attachment

import frappe
import openpyxl
from openpyxl.styles import PatternFill
from frappe.utils import nowdate, add_days
from io import BytesIO

@frappe.whitelist()
def sales_order_follow_up_test():
    def send_sales_report_with_table():
        filename = "Sales_Order_Follow_Up_" + nowdate() + ".xlsx"
        xlsx_file = build_xlsx_response_sales(filename)
        html_table, total_count = sales_next_action()
        send_mail_with_attachment_and_html(filename, xlsx_file, html_table)

    def build_xlsx_response_sales(filename):
        return make_xlsx_sales(filename)

    def make_xlsx_sales(filename, sheet_name=None, wb=None, column_widths=None):
        if wb is None:
            wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name or filename
        default_column_widths = [15, 25, 25, 15, 25, 20]
        column_widths = column_widths or default_column_widths
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        header_fill = PatternFill(start_color="87CEFA", end_color="87CEFA", fill_type="solid")
        headers = ["name", "account_manager", "service", "status", "customer", "company", "transaction_date", "grand_total", "per_billed", "advance_paid", "to_be_billed"]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill

        sales_orders = frappe.get_list(
            "Sales Order",
            filters={"status": ["not in", ["Hold", "To Deliver", "Closed", "Cancelled", "Completed"]]},
            fields=["name", "account_manager", "service", "status", "customer", "company", "transaction_date", "base_grand_total", "per_billed", "advance_paid"]
        )

        service_summary = {}
        total_outstanding = 0

        for order in sales_orders:
            to_be_billed = order.base_grand_total - (order.advance_paid + ((order.per_billed / 100) * order.base_grand_total))
            ws.append([
                order.name, order.account_manager, order.service, order.status, order.customer, order.company,
                order.transaction_date.strftime("%d-%m-%Y"), round(order.base_grand_total, 2), round(order.per_billed, 2),
                round(order.advance_paid, 2), round(to_be_billed, 2)
            ])
            if order.service not in service_summary:
                service_summary[order.service] = {"base_grand_total": 0, "outstanding": 0}
            service_summary[order.service]["base_grand_total"] += order.base_grand_total
            service_summary[order.service]["outstanding"] += to_be_billed
            total_outstanding += to_be_billed

        # Add total row
        ws.append([""] * 9 + ["Total", round(total_outstanding, 2)])

        with BytesIO() as b:
            wb.save(b)
            b.seek(0)
            return b.read()

    def sales_next_action():
        sales_orders = frappe.get_list(
            "Sales Order",
            filters={"status": ["not in", ["Hold", "To Deliver", "Closed", "Cancelled", "Completed"]]},
            fields=["name", "account_manager", "service", "status", "customer", "company", "transaction_date", "base_grand_total", "per_billed", "advance_paid"]
        )

        service_summary = {}
        detailed_rows = []

        for order in sales_orders:
            if order.service not in service_summary:
                service_summary[order.service] = {"base_grand_total": 0, "outstanding": 0}
            
            to_be_billed = order.base_grand_total - (order.advance_paid + ((order.per_billed / 100) * order.base_grand_total))
            service_summary[order.service]["base_grand_total"] += order.base_grand_total
            service_summary[order.service]["outstanding"] += to_be_billed

            transaction_date = order.transaction_date.strftime("%d-%m-%Y")
            
            detailed_rows.append('<tr style="font-size:14px"><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td style="text-align:left;">{}</td><td style="text-align:right;">{}</td><td style="text-align:right;">{}</td><td style="text-align:right;">{}</td><td style="text-align:right;">{}</td><td style="text-align:right;">{}</td></tr>'.format(
                order.name, order.account_manager, order.service, order.status, order.customer, order.company, transaction_date, round(order.base_grand_total, 2), round(order.per_billed, 2), round(order.advance_paid, 2), round(to_be_billed, 2)))

        summary_table = '<table border=1><tr style="text-align: center"><td style="background-color:#063970;color:white">Services</td><td style="background-color:#063970;color:white">Grand Total</td><td style="background-color:#063970;color:white">Outstanding</td></tr>'
        grand_total_amount = 0
        total_outstanding = 0

        for service, amounts in service_summary.items():
            summary_table += '<tr style="font-size:14px"><td>{}</td><td style="text-align:right;">{}</td><td style="text-align:right;">{}</td></tr>'.format(service, round(amounts["base_grand_total"], 2), round(amounts["outstanding"], 2))
            grand_total_amount += amounts["base_grand_total"]
            total_outstanding += amounts["outstanding"]

        summary_table += '<tr><td></td><td style="text-align:center;" colspan=1>Total</td><td style="text-align:right;">{}</td></tr>'.format(round(total_outstanding, 2))
        summary_table += '</table>'
        
        details_table = '<table border=1><tr style="text-align: center"><td style="background-color:#063970;color:white">ID</td><td style="background-color:#063970;color:white">Account Manager</td><td style="background-color:#063970;color:white">Service</td><td style="background-color:#063970;color:white">Status</td><td style="background-color:#063970;color:white">Customer Name</td><td style="background-color:#063970;color:white">Company</td><td style="background-color:#063970;color:white">Date</td><td style="background-color:#063970;color:white">Grand Total</td><td style="background-color:#063970;color:white">% Amount Billed</td><td style="background-color:#063970;color:white">Advance Paid</td><td style="background-color:#063970;color:white">To Be Billed</td></tr>'
        details_table += ''.join(detailed_rows)
        details_table += '<tr><td colspan=9></td><td style="text-align:center;">Total</td><td style="text-align:right;">{}</td></tr>'.format(round(total_outstanding, 2))
        details_table += '</table>'
        
        total_count = len(sales_orders)
        return summary_table + details_table, total_count

    def send_mail_with_attachment_and_html(filename, file_content, html_content):
        attachments = [{"fname": filename, "fcontent": file_content}]
        frappe.sendmail(
            # recipients='siva.m@groupteampro.com',
           recipients='dineshbabu.k@groupteampro.com',
            cc=["accounts@groupteampro.com","sangeetha.s@groupteampro.com","sangeetha.a@groupteampro.com","annie.m@groupteampro.com","amirtham.g@groupteampro.com"],
            subject='Sales Invoice Follow Up-Sales Order Outstanding',
            message="""
            <br>
            <p>Collection Outstanding Report For Further Action.</p>
            REC   : AS/AM<br><br>
            IT-SW : DKB/APP<br><br>
            TFP   : SBMK/AM<br><br>
            BCS   : SBMK<br><br>
            TGT   : SBMK<br><br>
            <br>
            {0}
            <br><br>
            Thanks & Regards,<br>TEAMPRO<br>"This email has been automatically generated. Please do not reply"<br><br>"Initiate further action and intimate a direct manager through email."
            """.format(html_content),
            attachments=attachments,
        )

    send_sales_report_with_table()


########   Sales Invoice Overall Service Excel Sheet Attachment   ###########


import frappe
import openpyxl
from openpyxl.styles import PatternFill
from frappe.utils import nowdate
from io import BytesIO

@frappe.whitelist()
def sales_invoice_follow_up_test():
    def send_sales_report_with_table():
        filename = "Sales_Invoice_Follow_Up_" + nowdate() + ".xlsx"
        xlsx_file = build_xlsx_response_sales(filename)
        html_table, total_count = sales_next_action()
        send_mail_with_attachment_and_html(filename, xlsx_file, html_table)

    def build_xlsx_response_sales(filename):
        return make_xlsx_sales(filename)

    def make_xlsx_sales(filename, sheet_name=None, wb=None, column_widths=None):
        import openpyxl
        from openpyxl.styles import PatternFill
        from io import BytesIO

        if wb is None:
            wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name or filename
        default_column_widths = [15, 25, 25, 15, 25, 20]
        column_widths = column_widths or default_column_widths
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        header_fill = PatternFill(start_color="87CEFA", end_color="87CEFA", fill_type="solid")
        headers = ["ID", "Account Manager", "Service", "Customer Name", "Company", "Date", "Grand Total", "Outstanding Amount", "Age"]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill

        sales_invoice = frappe.get_list("Sales Invoice", filters={"status": ["not in", ["Return", "Credit Note Issued", "Paid", "Cancelled"]]}, fields=["name", "company", "customer", "services", "posting_date", "due_date", "grand_total", "outstanding_amount", "account_manager", "delivery_manager"])

        service_summary = {}
        total_outstanding = 0

        for order in sales_invoice:
            todate = date.today()
            grand_total = round(order.grand_total, 2)
            outstanding_amount = round(order.outstanding_amount, 2)
            total_outstanding += outstanding_amount
            postingdate1 =(order.posting_date)
            age = (todate - postingdate1).days

            ws.append([
                order.name, order.account_manager, order.services, order.customer, order.company,
                order.posting_date.strftime("%d-%m-%Y"), grand_total, outstanding_amount, age
            ])

            if order.services not in service_summary:
                service_summary[order.services] = {"grand_total": 0, "outstanding": 0}
            service_summary[order.services]["grand_total"] += grand_total
            service_summary[order.services]["outstanding"] += outstanding_amount

        
        ws.append([""] * 6 + ["Total", round(total_outstanding, 2)])

        with BytesIO() as b:
            wb.save(b)
            b.seek(0)
            return b.read()

    def sales_next_action():
        sales_invoice = frappe.get_list("Sales Invoice", filters={"status": ["not in", ["Return", "Credit Note Issued", "Paid", "Cancelled"]]}, fields=["name", "company", "customer", "services", "posting_date", "due_date", "grand_total", "outstanding_amount", "account_manager", "delivery_manager"])

        service_summary = {}
        detailed_rows = []

        for order in sales_invoice:
            grand_total = round(order.grand_total, 2)
            outstanding_amount = round(order.outstanding_amount, 2)

            if order.services not in service_summary:
                service_summary[order.services] = {"grand_total": 0, "outstanding": 0}
            service_summary[order.services]["grand_total"] += grand_total
            service_summary[order.services]["outstanding"] += outstanding_amount

            transaction_date = (order.posting_date.strftime("%d-%m-%Y"))
            todate = date.today()
            age = (todate - order.posting_date).days
            
            detailed_rows.append('<tr style="font-size:14px"><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td style="text-align:left;">{}</td><td style="text-align:right;">{}</td><td style="text-align:right;">{}</td><td style="text-align:right;">{}</td></tr>'.format(
                order.name, order.account_manager, order.services, order.customer, order.company,transaction_date, grand_total, outstanding_amount, age))

        summary_table = '<table border=1><tr style="text-align: center"><td style="background-color:#063970;color:white">Services</td><td style="background-color:#063970;color:white">Grand Total</td><td style="background-color:#063970;color:white">Outstanding</td></tr>'
        grand_total_amount = 0
        total_outstanding = 0

        for service, amounts in service_summary.items():
            summary_table += '<tr style="font-size:14px"><td>{}</td><td style="text-align:right;">{}</td><td style="text-align:right;">{}</td></tr>'.format(service, round(amounts["grand_total"], 2), round(amounts["outstanding"], 2))
            grand_total_amount += amounts["grand_total"]
            total_outstanding += amounts["outstanding"]

        summary_table += '<tr><td></td><td style="text-align:center;" colspan=1>Total</td><td style="text-align:right;">{}</td></tr>'.format(round(total_outstanding, 2))
        summary_table += '</table>'
        
        details_table = '<table border=1><tr style="text-align: center"><td style="background-color:#063970;color:white">ID</td><td style="background-color:#063970;color:white">Account Manager</td><td style="background-color:#063970;color:white">Service</td><td style="background-color:#063970;color:white">Customer Name</td><td style="background-color:#063970;color:white">Company</td><td style="background-color:#063970;color:white">Date</td><td style="background-color:#063970;color:white">Grand Total</td><td style="background-color:#063970;color:white">Outstanding Amount</td><td style="background-color:#063970;color:white">Age</td></tr>'
        details_table += ''.join(detailed_rows)
        details_table += '<tr><td colspan=6></td><td style="text-align:center;">Total</td><td style="text-align:right;">{}</td></tr>'.format(round(total_outstanding, 2))
        details_table += '</table>'
        
        total_count = len(sales_invoice)
        return summary_table + details_table, total_count

    def send_mail_with_attachment_and_html(filename, file_content, html_content):
        attachments = [{"fname": filename, "fcontent": file_content}]
        frappe.sendmail(
            recipients='dineshbabu.k@groupteampro.com',
            cc=["accounts@groupteampro.com","sangeetha.s@groupteampro.com","sangeetha.a@groupteampro.com","annie.m@groupteampro.com","amirtham.g@groupteampro.com"],
            subject='Collection Follow Up-Sales Invoice Report',
            message="""
            <br>
            <p>Collection Outstanding Report For Further Action.</p>
            REC   : AS/AM<br><br>
            IT-SW : DKB/APP<br><br>
            TFP   : SBMK/AM<br><br>
            BCS   : SBMK<br><br>
            TGT   : SBMK<br><br>
            <br>
            {0}
            <br><br>
            Thanks & Regards,<br>TEAMPRO<br>"This email has been automatically generated. Please do not reply"<br><br>"Initiate further action and intimate a direct manager through email."
            """.format(html_content),
            attachments=attachments,
        )

    send_sales_report_with_table()



import frappe
from frappe.utils import nowdate, today
import openpyxl
from io import BytesIO
from frappe.utils.pdf import get_pdf

@frappe.whitelist()
def candidate_excel_format():
    next_date=today()
    next_dates=datetime.strptime(next_date, '%Y-%m-%d')
    formatted_next_date=next_dates.strftime('%Y-%m-%d')
    filename = "Candidate_Details_" + today() + ".xlsx"
    pdffilename = "Candidate_Details_" + today() + ".pdf"
    candidates = frappe.get_all(
        "Candidate",
        filters={'submitted_date': formatted_next_date},
        fields=["candidate_created_by"],
        group_by='candidate_created_by'
    )

    for user in candidates:
        user_id = user.candidate_created_by
        xlsx_file = make_xlsx_candidate(filename, user_id)
        pdf_content = make_pdf_candidate(pdffilename, user_id)
        candidate_status_mail_test(filename, xlsx_file.getvalue(), pdffilename, pdf_content, user_id)

def candidate_status_mail_test(filename, file_content, pdffilename, pdf_content, user_id):
    next_date=today()
    next_dates=datetime.strptime(next_date, '%Y-%m-%d')
    formatted_next_date=next_dates.strftime('%Y-%m-%d')
    data=""
    s_no = 0
    candidates = frappe.db.sql(
        """
        SELECT c.name, c.passport_number, c.given_name, c.highest_degree,
               c.total_experience, c.overseas_experience, c.current_employer,
               c.current_ctc, c.expected_ctc, c.location, c.notice_period_months,
               c.remarks_1, c.position, c.currency_ctc
        FROM `tabCandidate` c
        INNER JOIN `tabCandidate status` cs ON c.name = cs.parent
        WHERE c.candidate_created_by = %s
        AND cs.status = %s
        AND DATE(cs.sourced_date) = %s
        """,
        (user_id, "Pending QC", formatted_next_date),
        as_dict=True
    )

    grouped_candidates = {}
    for candidate in candidates:
        position = candidate.get("position", "")
        currency = candidate.get("currency_ctc", "")  # Default to SAR
        current_ctc = candidate.get("current_ctc", 0)  # Default to 0
        formatted_ctc = f"{currency} {current_ctc}" if current_ctc else " "

        if position not in grouped_candidates:
            grouped_candidates[position] = []
        grouped_candidates[position].append([
            candidate.get("name", "-"),
            candidate.get("passport_number", "-"),
            candidate.get("given_name", "-"),
            candidate.get("highest_degree", "-"),
            candidate.get("total_experience", "-"),
            candidate.get("overseas_experience", "-"),
            candidate.get("current_employer", "-"),
            formatted_ctc,
            candidate.get("expected_ctc", "-"),
            candidate.get("location", "-"),
            candidate.get("notice_period_months", "-"),
            candidate.get("remarks_1", "-"),
        ])

    # Define headers for the table
    headers = [
        "Candidate ID", "PP Number", "Candidate Name", "Qualification", 
        "Total Yrs of Exp", "Overseas Exp", "Current Employer", 
        "Current Salary", "Exp. Salary", "Current Location", 
        "Notice Period", "Remarks"
    ]

    for position, candidates in grouped_candidates.items():
        # Add position header and start table
        data += f"""
        <table class='table table-bordered' style='border: 1px solid black; border-collapse: collapse; width: 100%;'>
        <tr style='border: 1px solid black; background-color: #0f1568; color: white;'>
        <th colspan="12" style="text-align: center; font-size: 18px;">Position: {position}</th>
        </tr>
        <tr style='border: 1px solid black; background-color: #98D7F5; color: black;'>
        """
        # Add headers to the table
        for header in headers:
            data += f"<th style='border: 1px solid black;'>{header}</th>"
        data += "</tr>"

        # Add rows for each candidate under the position
        for candidate in candidates:
            data += "<tr style='border: 1px solid black;'>"
            for value in candidate:
                data += f"<td style='border: 1px solid black;'>{value}</td>"
            data += "</tr>"
        data += "</table><br>"

    subject = f"Candidates Submitted - {nowdate()}"
    message = f"""
    Dear Sir/Madam,<br><br>
    Kindly find the below list of candidates you submitted today:<br><br>{data if data else ''}<br><br>
    Thanks & Regards,<br>TEAM ERP<br>
    <i>This email has been automatically generated. Please do not reply</i>
    """

    frappe.sendmail(
        recipients=[user_id],
        # recipients=["divya.p@groupteampro.com"],
        subject=subject,
        message=message,
        attachments=[
            {"fname": filename, "fcontent": file_content},
            {"fname": pdffilename, "fcontent": pdf_content}
        ]
    )

def make_xlsx_candidate(filename, user_id):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Candidates'
    
    # Define column width
    for col in range(ord('A'), ord('M')):  # Columns A to L
        ws.column_dimensions[chr(col)].width = 20

    # Define headers
    headers = ["Candidate ID", "PP Number", "Candidate Name", "Qualification", 
               "Total Yrs of Exp", "Overseas Exp", "Current Employer", 
               "Current Salary", "Exp. Salary", "Current Location", 
               "Notice Period", "Remarks"]

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Fetch and group candidates by position
    position_candidates = get_data_grouped_by_position_candidate(user_id)
    
    # Debug: Check if any positions are retrieved
    print(f"Positions found: {len(position_candidates)}")
    
    if not position_candidates:
        print("No candidates found for the given user.")
    
    for position, candidates in position_candidates.items():
        # Add position row
        position_row = ws.max_row + 1
        ws.merge_cells(start_row=position_row, start_column=1, end_row=position_row, end_column=12)
        cell = ws.cell(row=position_row, column=1)
        cell.value = f"{position}"
        cell.fill = PatternFill(start_color="0F1568", end_color="0F1568", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add headers
        header_row = ws.max_row + 1
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.fill = PatternFill(start_color="98D7F5", end_color="98D7F5", fill_type="solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Add candidate rows
        for candidate in candidates:
            row_num = ws.max_row + 1
            for col_num, value in enumerate(candidate, start=1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = thin_border  # Apply border to each cell

        # Add an empty row for separation
        ws.append([])

    # Save the file into a BytesIO stream
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file

from frappe.utils.pdf import get_pdf
def make_pdf_candidate(pdffilename, user_id):
    html = """
    <html>
    <head>
    <style>
    table { width: 100%; border-collapse: collapse; }
    table, th, td { border: 1px solid black; }
    th, td { padding: 5px; text-align: left; }
    th { background-color: #98D7F5; color: black; }  /* Header background color */
    td.position { background-color: #0F1568; color: white; } /* Position row background color */
    </style>
    </head>
    <body>
    <h2>Candidate Details</h2>
    """
    
    next_date = today()
    next_dates = datetime.strptime(next_date, '%Y-%m-%d')
    formatted_next_date = next_dates.strftime('%Y-%m-%d')
    
    candidates = frappe.db.sql(
        """
        SELECT c.name, c.passport_number, c.given_name, c.highest_degree,
               c.total_experience, c.overseas_experience, c.current_employer,
               c.current_ctc, c.expected_ctc, c.location, c.notice_period_months,
               c.remarks_1, c.position, c.currency_ctc
        FROM `tabCandidate` c
        INNER JOIN `tabCandidate status` cs ON c.name = cs.parent
        WHERE c.candidate_created_by = %s
        AND cs.status = %s
        AND DATE(cs.sourced_date) = %s
        """,
        (user_id, "Pending QC", formatted_next_date),
        as_dict=True
    )

    # Group candidates by position
    grouped_candidates = {}
    for candidate in candidates:
        position = candidate.get("position", "")
        currency = candidate.get("currency_ctc", "")  # Default to SAR
        current_ctc = candidate.get("current_ctc", 0)  # Default to 0
        formatted_ctc = f"{currency} {current_ctc}" if current_ctc else " "

        if position not in grouped_candidates:
            grouped_candidates[position] = []
        grouped_candidates[position].append([ 
            candidate.get("name", "-"),
            candidate.get("passport_number", "-"),
            candidate.get("given_name", "-"),
            candidate.get("highest_degree", "-"),
            candidate.get("total_experience", "-"),
            candidate.get("overseas_experience", "-"),
            candidate.get("current_employer", "-"),
            formatted_ctc,
            candidate.get("expected_ctc", "-"),
            candidate.get("location", "-"),
            candidate.get("notice_period_months", "-"),
            candidate.get("remarks_1", "-"),
        ])

    # Define table headers
    headers = [
        "Candidate ID", "PP Number", "Candidate Name", "Qualification", 
        "Total Yrs of Exp", "Overseas Exp", "Current Employer", 
        "Current Salary", "Exp. Salary", "Current Location", 
        "Notice Period", "Remarks"
    ]

    # Add data position-wise to the HTML
    for position, candidates in grouped_candidates.items():
        # Add position header with custom color
        html += f"""
        <table>
        <tr>
        <td class="position" colspan="12">{position}</td>
        </tr>
        """
        
        # Add table headers
        html += "<tr>"
        for header in headers:
            html += f"<th>{header}</th>"
        html += "</tr>"

        # Add candidate rows
        for candidate in candidates:
            html += "<tr>"
            for value in candidate:
                html += f"<td>{value}</td>"
            html += "</tr>"
        html += "</table>"

    html += """
    </body>
    </html>
    """

    # Generate PDF from the HTML
    pdf_content = get_pdf(html)
    return pdf_content

def get_data_grouped_by_position_candidate(user_id):
    """
    Fetch candidate data grouped by position using SQL query.
    """
    data = {}
    next_date=today()
    next_dates=datetime.strptime(next_date, '%Y-%m-%d')
    formatted_next_date=next_dates.strftime('%Y-%m-%d')

    # Execute the SQL query to fetch candidates
    candidates = frappe.db.sql(
        """
        SELECT c.name, c.passport_number, c.given_name, c.highest_degree,
               c.total_experience, c.overseas_experience, c.current_employer,
               c.current_ctc, c.expected_ctc, c.location, c.notice_period_months,
               c.remarks_1, c.position, c.currency_ctc
        FROM `tabCandidate` c
        INNER JOIN `tabCandidate status` cs ON c.name = cs.parent
        WHERE c.candidate_created_by = %s
        AND cs.status = %s
        AND DATE(cs.sourced_date) = %s
        """,
        (user_id, "Pending QC", formatted_next_date),
        as_dict=True
    )

    # Group candidates by position
    for candidate in candidates:
        position = candidate.get("position", "")
        currency = candidate.get("currency_ctc", "")  # Default to SAR if not specified
        current_ctc = candidate.get("current_ctc", 0)  # Default to 0 if not specified
        formatted_ctc = f"{currency} {current_ctc}" if current_ctc else " "

        if position not in data:
            data[position] = []
        data[position].append([
            candidate.name, candidate.passport_number, candidate.given_name,
            candidate.highest_degree, candidate.total_experience, 
            candidate.overseas_experience, candidate.current_employer,
            formatted_ctc, candidate.expected_ctc, candidate.location, 
            candidate.notice_period_months, candidate.remarks_1
        ])

    return data


# Daily DPR Mail Trigger # 

import frappe
from datetime import datetime
from collections import defaultdict

@frappe.whitelist()
def task_mail():
    current_date = datetime.now().strftime("%d-%m-%Y")
    
    
    table_style = 'style="width: 100%; border-collapse: collapse;"'
    th_style = 'style="background-color:#063970; color:white; text-align:center; padding: 5px;"'
    td_style = 'style="text-align:center; padding: 5px;"'    
    open_issues_data_template = '''

    <table {0} border="1">
        <thead>
            <tr>
                <th {1} colspan="5">Open Issues - {2}</th>
            </tr>
            <tr>
                <th {1}>S.NO</th>
                <th {1}>Subject</th>
                <th {1}>Customer</th>
                <th {1}>Project</th>
                <th {1}>Count</th>
            </tr>
        </thead>
        <tbody>
    '''.format(table_style, th_style, current_date)

    assigned_to_list = frappe.db.sql("""
        SELECT project, subject, customer
        FROM `tabIssue`
        WHERE status = 'Open'
    """, as_dict=True)

    issue_counts = {}
    for issue in assigned_to_list:
        project = issue['project']
        if project not in issue_counts:
            issue_counts[project] = {'count': 0, 'subject': issue['subject'], 'customer': issue['customer']}
        issue_counts[project]['count'] += 1

    total_count = 0
    for idx, (project, data) in enumerate(issue_counts.items(), start=1):
        total_count += data['count']
        open_issues_data_template += '''
        <tr>
             <td {0}>{1}</td>
             <td {0}>{2}</td>
             <td {0}>{3}</td>
             <td {0}>{4}</td>
             <td {0}>{5}</td>
        </tr>'''.format(td_style, idx, data['subject'], data['customer'], project, data['count'])

    open_issues_data_template += '''
        <tr>
            <td {0} colspan="4"><strong>Total</strong></td>
            <td {0}><strong>{1}</strong></td>
        </tr>'''.format(td_style, total_count)

    open_issues_data_template += '''
            </tbody>
        </table>
        <br><br>
    '''
    
    
    open_meetings_data_template = '''
    <table {0} border="1">
        <thead>
            <tr>
                <th {1} colspan="3">Open Meetings - {2}</th>
            </tr>
            <tr>
                <th {1}>S.NO</th>
                <th {1}>Project</th>
                <th {1}>Count</th>
            </tr>
        </thead>
        <tbody>
    '''.format(table_style, th_style, current_date)

    assigned_to_list = frappe.db.get_all('Meeting', 
        filters={'status': ['not in', ['Completed', 'Cancelled']], 'custom_department': 'ITS - THIS'}, 
        fields=['project'])

    meeting_counts = {
        item['project']: frappe.db.count('Meeting', 
            filters={'status': ['not in', ['Completed', 'Cancelled']],'custom_department': 'ITS - THIS', 'project': item['project']})
        for item in assigned_to_list
    }

    total_count = 0
    for idx, (project, count) in enumerate(meeting_counts.items(), start=1):
        total_count += count
        open_meetings_data_template += '''
        <tr>
             <td {0}>{1}</td>
             <td {0}>{2}</td>
             <td {0}>{3}</td>
        </tr>'''.format(td_style, idx, project, count)

    open_meetings_data_template += '''
        <tr>
            <td {0} colspan="2"><strong>Total</strong></td>
            <td {0}><strong>{1}</strong></td>
        </tr>'''.format(td_style, total_count)

    open_meetings_data_template += '''
            </tbody>
        </table>
        <br><br>
    '''
    
    
    task_rt_data_template = '''
    <table {0} border="1">
        <thead>
            <tr>
                <th {1} colspan="3">Task Available RT - {2}</th>
            </tr>
            <tr>
                <th {1}>S.NO</th>
                <th {1}>CB</th>
                <th {1}>Count</th>
            </tr>
        </thead>
        <tbody>
    '''.format(table_style, th_style, current_date)
    
    tasks = frappe.db.get_all("Task", filters={'status': ['in', ['Open', 'Overdue', 'Working']], 'cb': ['not in', ['SM', 'JA']], 'service': 'IT-SW'}, fields=["cb", "rt"])
    
    cb_summary = defaultdict(lambda: {'total_rt': 0})
    
    for task in tasks:
        cb = task.get('cb', '')
        rt = task.get('rt', 0)
        
        cb_summary[cb]['total_rt'] += rt
    
    total_rt_overall = 0
    data_rows = ''
    
    for idx, (cb, summary) in enumerate(sorted(cb_summary.items()), start=1):
        total_rt = summary['total_rt']
        total_rt_overall += total_rt
        data_rows += '''
            <tr>
                <td {0}>{1}</td>
                <td {0}>{2}</td>
                <td {0}>{3}</td>
            </tr>
        '''.format(td_style, idx, cb, total_rt)
    
    task_rt_data_template += data_rows
    
    task_rt_data_template += '''
            <tr>
                <td {0} colspan="2"><strong>Total</strong></td>
                <td {0}><strong>{1}</strong></td>
            </tr>
            </tbody>
        </table>
        <br><br>
    '''.format(td_style, total_rt_overall)
    
    
    combined_data = '''
    <html>
    <body>
    '''
    
    combined_data += open_issues_data_template + open_meetings_data_template + task_rt_data_template
    
    combined_data += '''
    </body>
    </html>
    '''
    
    frappe.sendmail(
            recipients=['siva.m@groupteampro.com','abdulla.pi@groupteampro.com','dineshbabu.k@groupteampro.com'],
            subject='Task-Issue-Meeting - {}'.format(current_date),
            message=combined_data
        )

@frappe.whitelist()
def rename_case(doc, method):
    # Split batch parts
    batch_parts = doc.batch.split('-')
    batch_suffix = "-".join(batch_parts[1:])
    date_ddmmyy = batch_suffix[:6]
    batch_part = doc.batch.split('-')
    batch_suff ="-".join(batch_parts[2:])
    date_dd =  batch_suff[:5]
    # Create prefix for search
    case_prefix = f"{doc.customer_short_code}-{date_ddmmyy}-{date_dd}"
    # Query to get the highest series number for the given prefix
    highest_case = frappe.db.sql("""
        SELECT name FROM `tabCase`
        WHERE name LIKE %s
        ORDER BY name DESC
        LIMIT 1
    """, (case_prefix + '%'), as_dict=True)

    if highest_case:
        highest_series = int(highest_case[0]['name'].split('-')[-1]) 
        new_series = highest_series + 1
    else:
        new_series = 1

    series_str = str(new_series).zfill(5)
    case_id = f"{doc.customer_short_code}-{date_ddmmyy}-{date_dd}-{series_str}"
    frappe.rename_doc("Case", doc.name, case_id, force=1)




@frappe.whitelist()
def total_wh_hrs(in_time, out_time):
    if in_time and out_time:
        wh = time_diff_in_hours(out_time, in_time)
        return wh

def time_diff_in_hours(out_time_str, in_time_str):
    out_time = datetime.strptime(out_time_str, '%Y-%m-%d %H:%M:%S')
    in_time = datetime.strptime(in_time_str, '%Y-%m-%d %H:%M:%S')
    time_difference = out_time - in_time
    
    hours = time_difference.total_seconds() / 3600
    return round(hours, 2)

    

@frappe.whitelist()
def task_mail_notification():
    projects=frappe.get_all("Project",{'status':'Open','service':('in',['REC-D','REC-I'])},['*'])
    table = '<table text-align="center" border="1" width="100%" style="border-collapse: collapse;text-align: left;">'
    row=0
    for project in projects:
        tasks = frappe.get_all("Task", {'status': ('in', ['Open', 'Working','Overdue','Pending Review']),'project':project.name,'service':('in',['REC-D','REC-I'])},['*'])
        candidate_count=frappe.db.count("Candidate", {'project':project.name,'pending_for':('not in',['IDB','Sourced','Proposed PSL'])})
        if candidate_count>0:
            if row>0:
                table += """<tr><td style="border: none; border-left: hidden; border-right: hidden; height: 40px;" colspan=6></td></tr>"""
            # table+="""<tr><td style="border-left: hidden; border-right: hidden; border-top:hidden; border-bottom:hidden;"colspan=6></td></tr>"""
            table+="""<tr><td style="border-left: none; border-right: none;text-align: left;"colspan=2>Customer Name</td><td style="text-align: left;" colspan=4>%s</td></tr>"""%(project.customer)
            table+="""<tr><td style="border-left: none; border-right: none;text-align: left;"colspan=2>Spoc</td><td style="text-align: left;"colspan=4>%s</td></tr>"""%(project.spoc)
            table+="""<tr><td style="border-left: none; border-right: none;text-align: left;"colspan=2>Account Manager</td><td style="text-align: left;"colspan=4>%s</td></tr>"""%(project.account_manager)
            table += '<tr style="background-color: #0f1568"><td style="width: 2%; font-weight: bold;color: white;text-align: left;">ID</td><td style="width: 1%; font-weight: bold;color: white;text-align: left;">Status</td><td style="width: 5%; font-weight: bold; color: white;text-align: left;">Given Name</td><td style="width:3%; font-weight: bold; color: white;text-align: left;">Position</td><td style="width: 3%; font-weight: bold;color: white;text-align: left;">Passport Number</td><td style="width: 2%; font-weight: bold;color: white;text-align: left;">Project</td></tr>'
        for j in tasks:
            candidate=frappe.get_all("Candidate",{'pending_for':('not in',['IDB','Sourced','Proposed PSL']),'task':j.name},['name','pending_for','given_name','position','passport_number','project'])
            for ca in candidate:
                table+="""<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>""" % (ca.name,ca.pending_for,ca.given_name,ca.position,ca.passport_number or '-',ca.project)
                row+=1
    table += '</table>'
    frappe.sendmail(
        recipients=['sangeetha.a@groupteampro.com','lokeshkumar.a@groupteampro.com','ponkamaleshwari.i@groupteampro.com','rama.a@groupteampro.com'],
        cc=['sangeetha.s@groupteampro.com','dineshbabu.k@groupteampro.com'],
        subject='Candidate Feedback Pending Report',
        message=f"""
        <br>
         <p>As per the mail, Profile feedback pending list.</p>
        
        
          {table}<br><br>
        "This email has been automatically generated. PLEASE DONOT REPLY, Initiate further action and intimate your direct manager through email."
            <br><br>
            "With Best Wishes & Regards "
            <br><br>
            <span style="color:#203ed5;">
            "TEN – Auto Mail "
            </span>
            <br><br>
            <span style="color:#203ed5;">
                "Disclaimers:<br>
                This email and any files transmitted with it are confidential and intended solely for the use of the individual or entity to whom they are addressed. If you have received this email in error please notify the system manager. Please note that any views or opinions presented in this email are solely those of the author and do not necessarily represent those of the company. Finally, the recipient should check this email and any attachments for the presence of viruses. The company accepts no liability for any damage caused by any virus transmitted by this email."
            </span>
        """
    )

@frappe.whitelist()
def update_doc_in_sfu(doc, method):
    # appointment = frappe.db.sql("""select name from `tabAppointment` where customer_name = '%s'""" %(doc.organization_name), as_dict=1)
    # frappe.log_error(message=appointment, title="Appointment")
    # doc.appointment_clone = appointment[0]['name']
    # doc.save()
    if doc.appointment_with == 'Lead':
        lead = frappe.db.get_value("Lead", {"company_name": doc.customer_name}, "name")
        if not doc.party:
            if not lead:
                new_lead = frappe.get_doc({
                    "doctype": "Lead",
                    "company_name": doc.customer_name,
                    "status": "Lead",
                    "lead_owner": frappe.session.user,
                    "email_id": doc.customer_email,
                    "mobile_no": doc.customer_phone_number
                })
                new_lead.insert(ignore_permissions=True)
                frappe.msgprint("Lead Created Successfully")
            else:
                frappe.msgprint("Lead Exists")
        sfu = frappe.db.get_value("Sales Follow Up", {"organization_name": doc.customer_name}, "name")
        
        if not sfu:
            new_sfu = frappe.get_doc({
                "doctype": "Sales Follow Up",
                "organization_name": doc.customer_name,
                "follow_up_to": "Appointment",
                "account_manager": frappe.session.user,
                "lead_owner": frappe.session.user,
                "appointment_clone": doc.name,
                "appointments":doc.name,
                "appointment_date":doc.scheduled_time
            })
            new_sfu.insert(ignore_permissions=True)
            frappe.msgprint("Sales Follow Up Created Successfully")
        else:
            frappe.msgprint("Sales Follow Up Exists")


# @frappe.whitelist()
# def case_status_update_exisiting():
#     filename='06cc110038Case Status.csv'
#     from frappe.utils.file_manager import get_file
#     filepath = get_file(filename)
#     pps = read_csv_content(filepath[1])
#     ind=0
#     for pp in pps:
#         frappe.db.set_value("Case",{"name":pp[0]},"case_status",pp[1])
#         ind+=1
#         print(pp[0])
#         print(pp[1])
#     print(ind)



@frappe.whitelist()
def update_sfp_type():
    sales=frappe.db.get_all("Sales Follow Up",{'account_manager':['!=',''],'lead_owner':['!=','']},["*"])
    count=0
    for i in sales:
        # owner=frappe.db.get_value("Customer",{'name':i.customer},['account_manager'])
        # frappe.db.set_value("Sales Follow Up",i.name,'account_manager_lead_owner',owner)
        # frappe.db.set_value("Sales Follow Up",i.name,'party_from',"Customer")
        frappe.db.set_value("Sales Follow Up",i.name,'account_manager_lead_owner',i.account_manager)
        count+=1
    print(count)

# @frappe.whitelist()
# def batch_status_update_exisiting():
#     filename='36d1fb5d397f410Case Update.csv'
#     from frappe.utils.file_manager import get_file
#     filepath = get_file(filename)
#     pps = read_csv_content(filepath[1])
#     ind=0
#     for pp in pps:
#         frappe.db.set_value("Case",{"name":pp[0]},"case_status",pp[1])
#         ind+=1
#         print(pp[0])
#         print(pp[1])
#     print(ind)


# @frappe.whitelist()
# def batch_status_update_exisiting():
#     filename='5493e09f59Custom Field (2).csv'
#     from frappe.utils.file_manager import get_file
#     filepath = get_file(filename)
#     pps = read_csv_content(filepath[1])
#     ind=0
#     for pp in pps:
#         invoice = frappe.new_doc("Custom Field")
#         invoice.dt = pp[0]
#         invoice.fieldname = pp[1]
#         invoice.fieldtype = pp[2]
#         invoice.options = pp[3]
#         invoice.insert_after = pp[4]
#         invoice.label =pp[5]
#         invoice.save()
#         # frappe.db.set_value("",{"name":pp[0]},"batch_status",pp[1])
#         ind+=1
#         # print(pp[0])
#         # print(pp[1])
#     print(ind)


import frappe

@frappe.whitelist()
def update_case_status_in_batch(doc, method):
    batch = frappe.get_doc("Batch", {"name": doc.batch})
    batch.casewise_status = []
    cases = frappe.db.get_all("Case", filters={"batch": doc.batch}, fields=["name", "case_status"])
    completed = 0
    insuff = 0
    pending = 0
    drop=0
    for case in cases:
        batch.append("casewise_status", {
            "case_id": case["name"],
            "case_status": case["case_status"]
        })
        if case["case_status"] in ["Case Completed","To be Billed","SO Created","Case Report Completed","Generate Report","Billed"]:
            completed += 1
        elif case["case_status"] in ["Entry-Insuff","Execution-Insuff"]:
            insuff += 1
        elif case["case_status"] in ["Drop"]:
            drop+=1
        else:
            pending += 1
        batch.comp=completed
        batch.insuff=insuff
        batch.pending=pending
        batch.custom_drop=drop
    batch.save()
    frappe.db.commit()


@frappe.whitelist()
def update_case_status_existing_batch():
    batches = frappe.get_all(
        "Batch", 
       filters={"batch_status": "Completed", "pending": [">", 0]},
        fields=["name"]
    )
    ind=0
    for batch_data in batches:
        ind+=1
        print(batch_data)
        batch = frappe.get_doc("Batch", batch_data["name"])
        batch.casewise_status = []
        cases = frappe.db.get_all(
            "Case", 
            filters={"batch": batch.name}, 
            fields=["name", "case_status"]
        )
        completed = 0
        insuff = 0
        pending = 0
        case=0
        drop=0
        for case in cases:
            if case.case_status in ["Case Completed","To be Billed","SO Created","Case Report Completed","Generate Report"]:
                completed += 1
            elif case.case_status in ["Entry-Insuff","Execution-Insuff"]:
                insuff += 1
            elif case["case_status"] in ["Drop"]:
                drop+=1
            else:
                pending += 1
            batch.append("casewise_status", {
                "case_id": case["name"],
                "case_status": case["case_status"]
            })
            batch.comp=completed
            batch.insuff=insuff
            batch.pending=pending
            batch.custom_drop=drop

            print(completed)
            print(insuff)
            print(pending)
        batch.save()
    frappe.db.commit()

# @frappe.whitelist()
# def task_mail_notification_status ():
#     job = frappe.db.exists('Scheduled Job Type','update_case_status_existing_batch')
#     if not job:
#         task = frappe.new_doc("Scheduled Job Type")
#         task.update({
#             "method": 'checkpro.custom.update_case_status_existing_batch',
#             "frequency": 'Cron',
#             "cron_format": '0 0 * * *'
#         })
#         task.save(ignore_permissions=True)

# @frappe.whitelist()
# def batch_status_update_in_test():
#     batch=frappe.db.get_all("Batch",{"batch_status":"Open"},["*"])
#     for j in batch:
#         cases=frappe.db.get_all("Case",{"batch":j.name},["*"])
#         case_sts=[]
#         tat_status=[]
#         batch_status=''
#         for i in cases:
#             case_sts.append(i.case_status)
#             tat_status.append(i.tat_monitor)
#             if any(status == "Draft" for status in case_sts):
#                 batch_status="Open"
#             elif any(status == "Entry-Insuff" for status in case_sts):
#                 batch_status="Open with Insuff"
#             elif any(status == "Execution-Insuff" for status in case_sts):
#                 batch_status="Open with Insuff"
#             elif any(status == "Entry-QC" for status in case_sts):
#                 batch_status="Open"
#             elif any(status == "Execution" for status in case_sts):
#                 batch_status="Open"
#             elif any(status == "Entry Completed" for status in case_sts):
#                 batch_status="Open"
#             elif  any(status == "Entry-Insuff" or status == "Execution-Insuff"  for status in case_sts):
#                 if any(tat=="Out TAT" for tat in tat_status):
#                     batch_status="Overdue with Insuff"
#             elif any(status == "Entry-Insuff" or status == "Execution-Insuff"  for status in case_sts):
#                 if any(tat=="In TAT" for tat in tat_status):
#                     batch_status="Open with Insuff"
#             elif any(status == "Entry Completed" or status == "Execution" or status == "Draft"  for status in case_sts):
#                 if any(tat=="In TAT" for tat in tat_status):
#                     batch_status="Open"
#             elif any(status == "Entry Completed" or status == "Execution" or status == "Draft"  for status in case_sts):
#                 if any(tat=="Out TAT" for tat in tat_status):
#                     batch_status="Overdue"
#             elif any(status == "Case Report Completed" or status == "Case Completed" or status == "To be Billed" or status == "SO Created" or status == "Drop" for status in case_sts):
#                 batch_status="Completed"
#         frappe.db.set_value("Batch",j.name,"batch_status",batch_status)


# @frappe.whitelist()
# def download_excel():
#     filename = "Candidate Details"
#     build_xlsx_response_new(filename)

# def build_xlsx_response_new(filename):
#     xlsx_file = make_xlsx_new(filename)
#     frappe.response['filename'] = filename + '.xlsx'
#     frappe.response['filecontent'] = xlsx_file.getvalue()
#     frappe.response['type'] = 'binary'

# def make_xlsx_new(data, sheet_name="Candidates", wb=None, column_widths=None):
#     args = frappe.local.form_dict
#     column_widths = column_widths or []
    
#     if wb is None:
#         wb = Workbook()
#     ws = wb.create_sheet(sheet_name, 0)

#     # Set column widths
#     for col in range(ord('A'), ord('M')):  # Columns A to L
#         ws.column_dimensions[chr(col)].width = 20

#     # Define headers
#     headers = ["Candidate ID", "PP Number", "Candidate Name", "Qualification", 
#                "Total Yrs of Exp", "Overseas Exp", "Current Employer", 
#                "Current Salary", "Exp. Salary", "Current Location", 
#                "Notice Period", "Remarks"]

#     # Define styles
#     position_fill = PatternFill(start_color="0F1568", end_color="0F1568", fill_type="solid")
#     position_font = Font(color="FFFFFF", bold=True)
#     header_fill = PatternFill(start_color="98D7F5", end_color="98D7F5", fill_type="solid")
#     header_font = Font(bold=True)
#     black_border = Border(
#         left=Side(border_style="thin", color="000000"),
#         right=Side(border_style="thin", color="000000"),
#         top=Side(border_style="thin", color="000000"),
#         bottom=Side(border_style="thin", color="000000")
#     )
#     # Fetch candidate data grouped by positions
#     position_candidates = get_data_new(args)

#     for position, candidates in position_candidates.items():
#         # Add position row
#         position_row = ws.max_row + 1
#         ws.merge_cells(start_row=position_row, start_column=1, end_row=position_row, end_column=12)
#         cell = ws.cell(row=position_row, column=1)
#         cell.value = f"{position}"
#         cell.fill = position_fill
#         cell.font = position_font
#         cell.alignment = Alignment(horizontal="center", vertical="center")
#         cell.border = black_border

#         # Add headers
#         header_row = ws.max_row + 1
#         for col_num, header in enumerate(headers, start=1):
#             cell = ws.cell(row=header_row, column=col_num)
#             cell.value = header
#             cell.fill = header_fill
#             cell.font = header_font
#             cell.alignment = Alignment(horizontal="center", vertical="center")
#             cell.border = black_border

#         # Add details for the position
#         for candidate in candidates:
#             # ws.append(candidate)
#             row_num = ws.max_row + 1
#             for col_num, value in enumerate(candidate, start=1):
#                 cell = ws.cell(row=row_num, column=col_num)
#                 cell.value = value
#                 cell.border = black_border  # Apply border to each cell

#         # Add an empty row for separation
#         ws.append([])

#     xlsx_file = BytesIO()
#     wb.save(xlsx_file)
#     xlsx_file.seek(0)
#     return xlsx_file

# import json
# import frappe

# def get_data_new(args):
#     if args is None:
#         args = {}

#     if isinstance(args.get('args'), str):
#         try:
#             args['filters'] = json.loads(args['args'])
#         except json.JSONDecodeError as e:
#             frappe.log_error(title='JSON Decode Error', message=str(e))
#             return []
    
#     args['filters'] = args.get('filters', {})

#     data = []
#     date_filter = args['filters'].get('custom_date_filter')
#     candidate_created_by_filter = args['filters'].get('custom_candidate_status_filter')
#     candidate_status = args['filters'].get('custom_status_filter')
#     # print(candidate_created_by_filter)
#     # print(date_filter)
#     if not date_filter or not candidate_created_by_filter:
#         frappe.log_error(title='Missing filters', message='Date filter or candidate_created_by filter is missing.')
#         return data

#     filters = {}

#     candidate_condition = candidate_created_by_filter.get('condition')
#     candidate_value = candidate_created_by_filter.get('value')

#     if candidate_condition in ('!=', '=',):
#         filters['candidate_created_by'] = [candidate_condition, candidate_value]
#     elif candidate_condition in ('like', 'not like') and isinstance(candidate_value, str):
#         filters['candidate_created_by'] = [candidate_condition, candidate_value]
#     elif candidate_condition == 'is' and candidate_value == 'set':
#         filters['candidate_created_by'] = ["is", "set"]
#     elif candidate_condition == 'is' and candidate_value == 'not set':
#         filters['candidate_created_by'] = ["is", "not set"]
#     elif candidate_condition in ('in', 'not in') and isinstance(candidate_value, list):
#         filters['candidate_created_by'] = [candidate_condition, candidate_value]
#     else:
#         return data

#     date_condition = date_filter.get('condition')
#     date_value = date_filter.get('value')

#     if date_condition == 'Between' and isinstance(date_value, list) and len(date_value) == 2:
#         filters['submitted_date'] = ['between', [date_value[0], date_value[1]]]
#     elif date_condition == 'in' and isinstance(date_value, list):
#         filters['submitted_date'] = ['in', date_value]
#     elif date_condition == 'not in' and isinstance(date_value, list):
#         filters['submitted_date'] = ['not in', date_value]
#     elif date_condition == 'is' and date_value == 'set':
#         filters['submitted_date'] = ['is', 'set'] 
#     elif date_condition == 'is' and date_value == 'not set':
#         filters['submitted_date'] = ['is', 'not set'] 
#     elif date_condition in ('<', '<=', '>', '>=') and isinstance(date_value, str):
#         filters['submitted_date'] = [date_condition, date_value]
#     elif date_condition in ('=', '!=') and isinstance(date_value, str):
#         filters['submitted_date'] = [date_condition, date_value]  
#     elif date_condition == 'Timespan' and isinstance(date_value, str):
#         # print(date_value)
#         from_date, to_date = get_timespan_custom(date_value)
#         filters['submitted_date'] = ['between', [from_date, to_date]]
#     elif date_condition == 'fiscal year' and isinstance(date_value, str):
#         fiscal_year_start, fiscal_year_end = get_fiscal_year_custom(date_value)
#         filters['submitted_date'] = ['between', [fiscal_year_start, fiscal_year_end]]
#     else:
#         # frappe.log_error(title='Invalid Date Filter', message='Date filter is not set properly.')
#         return data
    
#     if candidate_status:
#         status_condition = candidate_status.get('condition')
#         status_value = candidate_status.get('value')
#         if status_condition and status_value:
#             if status_value =="Submit(SPOC)" or status_value == "Submitted(Client)":
#                 filters['pending_for'] = [status_condition, status_value]
    
#     data = {}
#     candidates = frappe.get_all(
#         "Candidate",
#         filters=filters,
#         fields=["name", "passport_number", "given_name", "highest_degree",
#                 "total_experience", "overseas_experience", "current_employer",
#                 "current_ctc", "expected_ctc", "location", "notice_period_months",
#                 "remarks_1", "position","currency_ctc"]
#     )
#     for candidate in candidates:
#         position = candidate.get("position", "")
#         currency=candidate.currency_ctc
#         formatted_ctc = f"{currency} {candidate.current_ctc}" if candidate.current_ctc else "0"
#         if position not in data:
#             data[position] = []
#         data[position].append([
#             candidate.name, candidate.passport_number, candidate.given_name,
#             candidate.highest_degree, candidate.total_experience, 
#             candidate.overseas_experience, candidate.current_employer,
#             formatted_ctc, candidate.expected_ctc, candidate.location, 
#             candidate.notice_period_months, candidate.remarks_1
#         ])

#     return data



# def get_timespan_custom(timespan):
#     print(nowdate())
#     today = nowdate()
#     if timespan == "last week":
#         start_date = add_days(today, -7)
#         end_date = today
#     elif timespan == "last month":
#         start_date = add_months(today, -1)
#         end_date = today
#     elif timespan == "last quarter":
#         start_date = add_months(today, -3)
#         end_date = today
#     elif timespan == "last year":
#         start_date = add_months(today, -12)
#         end_date = today
#     elif timespan == "last 6 months":
#         start_date = add_months(today, -6)
#         end_date = today
#     elif timespan == "today":
#         start_date = end_date = today
#     elif timespan == "yesterday":
#         start_date = end_date = add_days(today, -1)
#     elif timespan == "tomorrow":
#         start_date = end_date = add_days(today, 1)
#     elif timespan == "next month":
#         start_date = add_months(today, 1)
#         end_date = add_days(add_months(today, 1), -1)
#     elif timespan == "next week":
#         start_date = today
#         end_date = add_days(today, 7)
#     elif timespan == "next quarter":
#         start_date = today
#         end_date = add_months(today, 3)
#     elif timespan == "next year":
#         start_date = today
#         end_date = add_months(today, 12)
#     elif timespan == "next 6 months":
#         start_date = today
#         end_date = add_months(today, 6)
#     elif timespan == "this week":
#         start_date = get_first_day(today, "week")  
#         end_date = add_days(start_date, 6) 
#     elif timespan == "this month":
#         start_date = get_first_day(today, "month")  
#         end_date = get_last_day(today, "month")  
#     elif timespan == "last month":
#         start_date = add_months(today, -1)
#         end_date = today
#     elif timespan == "this quarter":
#         start_date = get_first_day(today, "quarter")  
#         end_date = get_last_day(today, "quarter")  
#     elif timespan == "this year":
#         start_date = get_first_day(today, "year")  
#         end_date = get_last_day(today, "year")
#     else:
#         raise ValueError(f"Unsupported timespan: {timespan}")
    
#     return start_date, end_date

# def get_fiscal_year_custom(fiscal_year):
#     fiscal_year_split = fiscal_year.split('-')
#     start_year = fiscal_year_split[0]
#     end_year = fiscal_year_split[1]

#     start_date = date(int(start_year), 1, 1) 
#     end_date = date(int(end_year), 12, 31)    

#     return start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')


# @frappe.whitelist()
# def update_tat_existing():
#     batches=frappe.db.get_all("Batch",{"expected_start_date": ["between", ["2025-01-15", "2025-01-20"]]},["name"])
#     ind=0
#     for i in batches:
#         cases=frappe.db.get_all("Case",{"batch":i.name},["name"])
#         for j in cases:
#             doc=frappe.get_doc("Case",j.name)
#             if doc.insufficiency_closed:
#                 from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
#                 holiday_list_name = 'TEAMPRO 2023 - Checkpro'
#                 start_date = doc.insufficiency_closed
#                 working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
#                 current_date = start_date
#                 holiday = []
#                 while working_days > 0:
#                     if not is_holiday(holiday_list_name, current_date):
#                         holiday.append(current_date)
#                         working_days -= 1
#                     current_date = add_days(current_date, 1)
#                 frappe.db.set_value("Case",doc.name,"end_date",holiday[-1])



@frappe.whitelist()
def check_holidays(date1, date2,name):
    doc=frappe.get_doc("Case",name)
    if doc.date_of_initiating:
        from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
        holiday_list_name = 'TEAMPRO 2023 - Checkpro'
        start_date = doc.date_of_initiating
        working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
        current_date = start_date
        holiday = []
        while working_days > 0:
            if not is_holiday(holiday_list_name, current_date):
                holiday.append(current_date)
                working_days -= 1
            current_date = add_days(current_date, 1)
        sql_query = f"""
            SELECT COUNT(*) 
            FROM `tabHoliday` 
            WHERE parent = 'TEAMPRO 2023 - Checkpro' 
            AND holiday_date BETWEEN '{date1}' AND '{date2}'
        """

        count = frappe.db.sql(sql_query, as_list=True)[0][0]

        # return count
        return holiday[-1],count
    

@frappe.whitelist()
def case_status_report_excel():
    next_date=today()
    next_dates=datetime.strptime(next_date, '%Y-%m-%d')
    filename = "Case_Status_Report" + today() + ".xlsx"
    xlsx_file = make_xlsx_case_status(filename)
    case_status_report(filename, xlsx_file.getvalue())

def make_xlsx_case_status(filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Case Status Report"
    text_wrap_left = Alignment(vertical="center", horizontal="center")
    # Setting column widths
    for col in range(ord('A'), ord('M') + 1):  # Adjust for header range
        ws.column_dimensions[chr(col)].width = 20

    # Adding headers
    headers = [
        "Sr.no", "ID", "Employee Name", "Customer", "Check Package", "Batch", 
        "Case Status", "Case Report", "Client Employee Code", "Initiation Date",
        "Entry Allocated To", "Case Completion Date", "TAT Completion Date",
        "Insufficiency Closed", "Insufficiency Reported", "Actual Age",
        "0 to 5", "6 to 10", "11 to 15", ">15"
    ]
    ws.append(headers)  # Adding headers to the sheet

    # Formatting headers
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")  # White font color for better visibility
        cell.fill = PatternFill(start_color="FF002060", end_color="FF002060", fill_type="solid")  # aRGB format
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.alignment=text_wrap_left

    # Fetching case details
    case_details = get_case_report_detaiils()

    sr_no = 1
    for case in case_details:
        tat_counts = {
            "0 to 5": "",
            "6 to 10": "",
            "11 to 15": "",
            ">15": ""
        }
        # Determine the age range for `actual_tat`
        tat_range = {
            "0 to 5": 0 <= case["actual_tat"] <= 5,
            "6 to 10": 6 <= case["actual_tat"] <= 10,
            "11 to 15": 11 <= case["actual_tat"] <= 15,
            ">15": case["actual_tat"] > 15
        }

        tat_counts = {key: 1 if condition else 0 for key, condition in tat_range.items()}

        # Append data row
        ws.append([
            sr_no,
            case.get("name"),
            case.get("case_name"),
            case.get("customer"),
            case.get("check_package"),
            case.get("batch"),
            case.get("case_status"),
            case.get("case_report"),
            case.get("client_employee_code"),
            case.get("date_of_initiating"),
            case.get("allocated_to"),
            case.get("case_completion_date"),
            case.get("end_date"),
            case.get("insufficiency_closed"),
            case.get("insufficiency_reported"),
            case.get("actual_tat"),
            tat_counts["0 to 5"] or "",
            tat_counts["6 to 10"] or "",
            tat_counts["11 to 15"] or "",
            tat_counts[">15"] or ""
        ])
        sr_no += 1
        for cell in ws[ws.max_row]:
            cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # Save to BytesIO object
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file


def get_case_report_detaiils():
    cases=frappe.db.get_all("Case",{"case_status":("not in",["Final-QC","Generate Report","Case Report Completed","Case Completed","To be Billed","SO Created","Drop"])},["*"])
    return cases

@frappe.whitelist()
def case_status_report(filename, file_content):
    data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
    data += '<tr style="background-color: #002060; color: white;">' \
        '<td style="text-align:center; font-weight:bold; color:white;">Customer</td>' \
        '<td style="text-align:center; font-weight:bold; color:white;">0-5</td>' \
        '<td style="text-align:center; font-weight:bold; color:white;">6-10</td>' \
        '<td style="text-align:center; font-weight:bold; color:white;">11-15</td>' \
        '<td style="text-align:center; font-weight:bold; color:white;">>15</td>' \
        '<td style="text-align:center; font-weight:bold; color:white;">Entry Grand Total</td>' \
        '<td style="text-align:center; font-weight:bold; color:white;">Entry-Insuff</td>' \
        '<td style="text-align:center; font-weight:bold; color:white;">Execution-Insuff</td>' \
        '<td style="text-align:center; font-weight:bold; color:white;">Grand Total</td>' \
        '</tr>'

    # Fetch all batches with batch_status not "Completed"
    batches = frappe.db.get_all("Batch", {"batch_status": ("!=", "Completed")}, ["name", "customer"])

    customer_data = {}
    grand_totals = {
        "0-5": 0,
        "6-10": 0,
        "11-15": 0,
        ">15": 0,
        "Entry-Insuff": 0,
        "Execution-Insuff": 0
    }

    for batch in batches:
        customer = batch.customer
        if customer not in customer_data:
            customer_data[customer] = {
                "0-5": 0,
                "6-10": 0,
                "11-15": 0,
                ">15": 0,
                "Entry-Insuff": 0,
                "Execution-Insuff": 0
            }

        # Get cases for the current batch
        cases = frappe.db.get_all("Case", {"batch": batch.name}, ["name", "case_status", "actual_tat"])

        for case in cases:
            if case["case_status"] in ["Draft", "Entry Completed", "Entry-QC", "Execution"]:
                if 0 <= case["actual_tat"] <= 5:
                    customer_data[customer]["0-5"] += 1
                elif 6 <= case["actual_tat"] <= 10:
                    customer_data[customer]["6-10"] += 1
                elif 11 <= case["actual_tat"] <= 15:
                    customer_data[customer]["11-15"] += 1
                elif case["actual_tat"] > 15:
                    customer_data[customer][">15"] += 1

            if case["case_status"] == "Entry-Insuff":
                customer_data[customer]["Entry-Insuff"] += 1
            if case["case_status"] == "Execution-Insuff":
                customer_data[customer]["Execution-Insuff"] += 1

    # Populate the table with customer data
    for customer, counts in customer_data.items():
        # Calculate row-level totals
        entry_grand_total = counts["0-5"] + counts["6-10"] + counts["11-15"] + counts[">15"]
        grand_total = counts["Entry-Insuff"] + counts["Execution-Insuff"]

        # Update grand totals
        for key in grand_totals:
            grand_totals[key] += counts[key]

        # Append row data
        data += f'<tr>' \
            f'<td style="text-align:center;">{customer}</td>' \
            f'<td style="text-align:center;">{counts["0-5"] or ""}</td>' \
            f'<td style="text-align:center;">{counts["6-10"] or ""}</td>' \
            f'<td style="text-align:center;">{counts["11-15"] or ""}</td>' \
            f'<td style="text-align:center;">{counts[">15"] or ""}</td>' \
            f'<td style="text-align:center;">{entry_grand_total or ""}</td>' \
            f'<td style="text-align:center;">{counts["Entry-Insuff"] or ""}</td>' \
            f'<td style="text-align:center;">{counts["Execution-Insuff"] or ""}</td>' \
            f'<td style="text-align:center;">{grand_total or ""}</td>' \
            f'</tr>'

    # Append grand total row
    overall_grand_total = grand_totals["Entry-Insuff"] + grand_totals["Execution-Insuff"]
    entry_grand_total_sum = grand_totals["0-5"] + grand_totals["6-10"] + grand_totals["11-15"] + grand_totals[">15"]

    data += f'<tr style="font-weight: bold; background-color: #f2f2f2;">' \
        f'<td style="text-align:center;">Grand Total</td>' \
        f'<td style="text-align:center;">{grand_totals["0-5"] or ""}</td>' \
        f'<td style="text-align:center;">{grand_totals["6-10"] or ""}</td>' \
        f'<td style="text-align:center;">{grand_totals["11-15"] or ""}</td>' \
        f'<td style="text-align:center;">{grand_totals[">15"] or ""}</td>' \
        f'<td style="text-align:center;">{entry_grand_total_sum or ""}</td>' \
        f'<td style="text-align:center;">{grand_totals["Entry-Insuff"] or ""}</td>' \
        f'<td style="text-align:center;">{grand_totals["Execution-Insuff"] or ""}</td>' \
        f'<td style="text-align:center;">{overall_grand_total or ""}</td>' \
        f'</tr>'

    data += '</table>'

    frappe.sendmail(
        recipients=["sangeetha.a@groupteampro.com","sangeetha.s@groupteampro.com","dineshbabu.k@groupteampro.com","keerthana.b@groupteampro.com"],
        subject=_("Case Status Report"),
        message=f"""
            Dear Sir/Madam,<br><br>
            Kindly find the below list of Case Status Report:<br>{data}<br>
            Thanks & Regards,<br>
            TEAM ERP<br>
            <i>This email has been automatically generated. Please do not reply.</i>
        """,
        attachments=[
            {"fname": filename, "fcontent": file_content},
        ]
    )

# Add entry allocated to and allocated date in case in list view
@frappe.whitelist()
def update_entry_allocated_to(case_id,allocated_to,allocated_date):
    doc_name = json.loads(case_id)
    for i in doc_name:
        frappe.set_value("Case",i,"allocated_to",allocated_to)
        frappe.set_value("Case",i,"custom_allocation_date",allocated_date)
            
@frappe.whitelist()
def update_tat_case():
    frappe.enqueue(
        update_tat_completion_date_daily, 
        queue="long",
        timeout=36000,
        is_async=True, 
        now=False, 
        job_name='Tat Updation',
        enqueue_after_commit=False,
    )
# Daily cron update
@frappe.whitelist()
def update_tat_completion_date_daily():
    cases=frappe.db.get_all("Case",{"case_status":("not in",["Case Completed","To be Billed","SO Created","Billed"])},["name"])
    # cases=frappe.db.get_all("Case",{"name":("in",["KCP-221025-17038-00001","KBL-101025-16829-00014","KBL-101025-16828-00004","EID-110425-15820-00003"])},["name"])
    for i in cases:
        doc=frappe.get_doc("Case",i.name)
        if doc.insufficiency_closed:
            from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
            holiday_list_name = 'TEAMPRO 2023 - Checkpro'
            start_date = doc.insufficiency_closed
            working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
            current_date = start_date
            holiday = []
            while working_days > 0:
                if not is_holiday(holiday_list_name, current_date):
                    holiday.append(current_date)
                    working_days -= 1
                current_date = add_days(current_date, 1)
            sql_query = f"""
                SELECT COUNT(*) 
                FROM `tabHoliday` 
                WHERE parent = 'TEAMPRO 2023 - Checkpro' 
                AND holiday_date BETWEEN '{doc.insufficiency_closed}' AND '{holiday[-1]}'
            """
            count = frappe.db.sql(sql_query, as_list=True)[0][0]
            frappe.db.set_value("Case",doc.name,"end_date",holiday[-1])
            frappe.db.set_value("Case",doc.name,"holidays",count)
            # doc.end_date=holiday[-1]
            # doc.holidays=count
            # doc.save(ignore_permissions=True)
            # frappe.db.commit()
# @frappe.whitelist()
# def task_mail_notification_status ():
#     job = frappe.db.exists('Scheduled Job Type','update_holiday_tat_case')
#     if not job:
#         task = frappe.new_doc("Scheduled Job Type")
#         task.update({
#             "method": 'checkpro.custom.update_holiday_tat_case',
#             "frequency": 'Cron',
#             "cron_format": '0 0 * * *'
#         })
#         task.save(ignore_permissions=True)

@frappe.whitelist()
def update_holiday_tat_case():
    frappe.enqueue(
        check_holidays_not_insuff, 
        queue="long",
        timeout=36000,
        is_async=True, 
        now=False, 
        job_name='Tat Holiday Updation',
        enqueue_after_commit=False,
    )

@frappe.whitelist()
def check_holidays_not_insuff():
    cases=frappe.db.get_all("Case",{"case_status":("not in",["Case Completed","To be Billed","SO Created","Billed"])},["*"])
    for i in cases:
        doc=frappe.get_doc("Case",i.name)
        if doc.date_of_initiating and not doc.insufficiency_closed:
            from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
            holiday_list_name = 'TEAMPRO 2023 - Checkpro'
            start_date = doc.date_of_initiating
            working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
            current_date = start_date
            holiday = []
            while working_days > 0:
                if not is_holiday(holiday_list_name, current_date):
                    holiday.append(current_date)
                    working_days -= 1
                current_date = add_days(current_date, 1)
            sql_query = f"""
                SELECT COUNT(*) 
                FROM `tabHoliday` 
                WHERE parent = 'TEAMPRO 2023 - Checkpro' 
                AND holiday_date BETWEEN '{doc.date_of_initiating}' AND '{doc.end_date}'
            """

            count = frappe.db.sql(sql_query, as_list=True)[0][0]
            frappe.db.set_value("Case",doc.name,"end_date",holiday[-1])
            frappe.db.set_value("Case",doc.name,"holidays",count)
            # doc.end_date=holiday[-1]
            # doc.holidays=count
            # doc.save(ignore_permissions=True)
            # frappe.db.commit()


@frappe.whitelist()
def insuff_tat_daily():
    list = ["Education Checks","Family","Reference Check","Court","Social Media","Criminal","Employment","Identity Aadhar","Address Check"]
    cases=frappe.db.get_all("Case",{"case_status":("not in",["Case Completed","To be Billed","SO Created"])},["*"])
    for j in cases:
        mini=[]
        maxi=[]
        for i in list:
            doc=frappe.get_all(i,{"case_id":j.name},["name","insufficiency_date","workflow_state","insuff_closed"])
            for j in doc:
                if j.insufficiency_date:
                    mini.append(j.insufficiency_date)
                if j.insuff_closed and j.workflow_state!="Insufficient Data":
                    maxi.append(j.insuff_closed)
        first_date= min(mini) if mini else None
        last_time = max(maxi) if maxi else None
        # frappe.db.set_value("Case",j.name,"insufficiency_reported",first_date)
        # frappe.db.set_value("Case",j.name,"insufficiency_closed",last_time)

@frappe.whitelist()
def update_insuff_days_daily():
    cases=frappe.db.get_all("Case",{"case_status":("not in",["Case Completed","To be Billed","SO Created"])},["*"])
    for i in cases:
        doc=frappe.get_doc("Case",i.name)
        if doc.insufficiency_reported and doc.insufficiency_closed:
            date=(date_diff(doc.insufficiency_reported,doc.insufficiency_closed))
            sql_query = f"""
                SELECT COUNT(*) 
                FROM `tabHoliday` 
                WHERE parent = 'TEAMPRO 2023 - Checkpro' 
                AND holiday_date BETWEEN '{doc.insufficiency_reported}' AND '{doc.insufficiency_closed}'
            """
            count = frappe.db.sql(sql_query, as_list=True)[0][0]
            date1 = (date-count)+1
            # doc.insufficiency_days=date1
            # doc.save(ignore_permissions=True)
            # frappe.db.commit()

@frappe.whitelist()
def update_tat_completion_date_ed_daily():
    checks=frappe.db.get_all("Education Checks",{"check_status":("!=",["Report Completed"])},["*"])
    for i in checks:
        doc=frappe.get_doc("Education Checks",i.name)
        if doc.clear_insufficiency:
            from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
            holiday_list_name = 'TEAMPRO 2023 - Checkpro'
            start_date = doc.clear_insufficiency
            working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
            current_date = start_date
            holiday = []
            while working_days > 0:
                if not is_holiday(holiday_list_name, current_date):
                    holiday.append(current_date)
                    working_days -= 1
                current_date = add_days(current_date, 1)
            # doc.tat_completion_date=holiday[-1]
            frappe.db.set_value("Education Checks",doc.name,"tat_completion_date",holiday[-1])
        elif not doc.clear_insufficiency and doc.check_creation_date:
            from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
            holiday_list_name = 'TEAMPRO 2023 - Checkpro'
            start_date = doc.check_creation_date
            working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
            current_date = start_date
            holiday = []
            while working_days > 0:
                if not is_holiday(holiday_list_name, current_date):
                    holiday.append(current_date)
                    working_days -= 1
                current_date = add_days(current_date, 1)
            # doc.tat_completion_date=holiday[-1]
            frappe.db.set_value("Education Checks",doc.name,"tat_completion_date",holiday[-1])

@frappe.whitelist()
def update_holidays_daily_edu():
    checks=frappe.db.get_all("Education Checks",{"check_status":("!=",["Report Completed"])},["*"])
    for i in checks:
        doc=frappe.get_doc("Education Checks",i.name)  
        if doc.check_creation_date and doc.check_completion_date:
            sql_query = f"""
            SELECT COUNT(*) 
            FROM `tabHoliday` 
            WHERE parent = 'TEAMPRO 2023 - Checkpro' 
            AND holiday_date BETWEEN '{doc.check_creation_date}' AND '{doc.check_completion_date}'
        """

            count = frappe.db.sql(sql_query, as_list=True)[0][0]
            frappe.db.set_value("Education Checks",doc.name,"holidays",count)

@frappe.whitelist()
def update_tat_completion_date_employment_daily():
    checks=frappe.db.get_all("Employment",{"check_status":("!=",["Report Completed"])},["*"])
    for i in checks:
        doc=frappe.get_doc("Employment",i.name)
        if doc.insuff_cleared_on:
            from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
            holiday_list_name = 'TEAMPRO 2023 - Checkpro'
            start_date = doc.insuff_cleared_on
            working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
            current_date = start_date
            holiday = []
            while working_days > 0:
                if not is_holiday(holiday_list_name, current_date):
                    holiday.append(current_date)
                    working_days -= 1
                current_date = add_days(current_date, 1)
            doc.tat_completion_date = holiday[-1]
        elif doc.check_creation_date and not doc.insuff_cleared_on:
            from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
            holiday_list_name = 'TEAMPRO 2023 - Checkpro'
            start_date = doc.check_creation_date
            working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
            current_date = start_date
            holiday = []
            while working_days > 0:
                if not is_holiday(holiday_list_name, current_date):
                    holiday.append(current_date)
                    working_days -= 1
                current_date = add_days(current_date, 1)
            doc.tat_completion_date = holiday[-1]

@frappe.whitelist()
def update_holidays_daily_emp():
    checks=frappe.db.get_all("Employment",{"check_status":("!=",["Report Completed"])},["*"])
    for i in checks:
        doc=frappe.get_doc("Employment",i.name)  
        if doc.check_creation_date and doc.check_completion_date:
            sql_query = f"""
            SELECT COUNT(*) 
            FROM `tabHoliday` 
            WHERE parent = 'TEAMPRO 2023 - Checkpro' 
            AND holiday_date BETWEEN '{doc.check_creation_date}' AND '{doc.check_completion_date}'
        """

            count = frappe.db.sql(sql_query, as_list=True)[0][0]
            frappe.db.set_value("Employment",doc.name,"holidays",count)


@frappe.whitelist()
def update_tat_completion_date_address_daily():
    checks=frappe.db.get_all("Address Check",{"check_status":("!=",["Report Completed"])},["*"])
    for i in checks:
        doc=frappe.get_doc("Address Check",i.name)
        if doc.clear_insufficiency:
            from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
            holiday_list_name = 'TEAMPRO 2023 - Checkpro'
            start_date = doc.clear_insufficiency
            if doc.check_package:
                working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
            current_date = start_date
            holiday = []
            while working_days > 0:
                if not is_holiday(holiday_list_name, current_date):
                    holiday.append(current_date)
                    working_days -= 1
                current_date = add_days(current_date, 1)
            doc.custom_tat_completion_date=holiday[-1]
        elif doc.check_creation_date and not doc.clear_insufficiency:
            from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
            holiday_list_name = 'TEAMPRO 2023 - Checkpro'
            start_date = doc.check_creation_date
            if doc.check_package:
                working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
            current_date = start_date
            holiday = []
            while working_days > 0:
                if not is_holiday(holiday_list_name, current_date):
                    holiday.append(current_date)
                    working_days -= 1
                current_date = add_days(current_date, 1)
            doc.custom_tat_completion_date=holiday[-1]

@frappe.whitelist()
def update_holidays_daily_add():
    checks=frappe.db.get_all("Address Check",{"check_status":("!=",["Report Completed"])},["*"])
    for i in checks:
        doc=frappe.get_doc("Address Check",i.name)  
        if doc.check_creation_date and doc.check_completion_date:
            sql_query = f"""
            SELECT COUNT(*) 
            FROM `tabHoliday` 
            WHERE parent = 'TEAMPRO 2023 - Checkpro' 
            AND holiday_date BETWEEN '{doc.check_creation_date}' AND '{doc.check_completion_date}'
        """

            count = frappe.db.sql(sql_query, as_list=True)[0][0]
            frappe.db.set_value("Address Check",doc.name,"holidays",count)


@frappe.whitelist()
def update_tat_completion_date_inchecks_daily():
    checks=frappe.db.get_all("Criminal",{"check_status":("!=",["Report Completed"])},["*"])
    for i in checks:
        doc=frappe.get_doc("Criminal",i.name)
        if doc.insuff_cleared_on:
            from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
            holiday_list_name = 'TEAMPRO 2023 - Checkpro'
            start_date = doc.insuff_cleared_on
            working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
            current_date = start_date
            holiday = []
            while working_days > 0:
                if not is_holiday(holiday_list_name, current_date):
                    holiday.append(current_date)
                    working_days -= 1
                current_date = add_days(current_date, 1)
            doc.tat_completion_date=holiday[-1]
        elif doc.check_creation_date and not doc.insuff_cleared_on:
            from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
            holiday_list_name = 'TEAMPRO 2023 - Checkpro'
            start_date = doc.check_creation_date
            working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
            current_date = start_date
            holiday = []
            while working_days > 0:
                if not is_holiday(holiday_list_name, current_date):
                    holiday.append(current_date)
                    working_days -= 1
                current_date = add_days(current_date, 1)
            doc.tat_completion_date=holiday[-1]

@frappe.whitelist()
def update_holidays_daily_criminal():
    checks=frappe.db.get_all("Criminal",{"check_status":("!=",["Report Completed"])},["*"])
    for i in checks:
        doc=frappe.get_doc("Criminal",i.name)  
        if doc.check_creation_date and doc.check_completion_date:
            sql_query = f"""
            SELECT COUNT(*) 
            FROM `tabHoliday` 
            WHERE parent = 'TEAMPRO 2023 - Checkpro' 
            AND holiday_date BETWEEN '{doc.check_creation_date}' AND '{doc.check_completion_date}'
        """

            count = frappe.db.sql(sql_query, as_list=True)[0][0]
            frappe.db.set_value("Criminal",doc.name,"holidays",count)

@frappe.whitelist()
def update_tat_completion_date_court_daily():
    checks=frappe.db.get_all("Court",{"check_status":("!=",["Report Completed"])},["*"])
    for i in checks:
        doc=frappe.get_doc("Court",i.name)
        if doc.insuff_cleared_on:
            from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
            holiday_list_name = 'TEAMPRO 2023 - Checkpro'
            start_date = doc.insuff_cleared_on
            if doc.check_package:
                working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
            current_date = start_date
            holiday = []
            while working_days > 0:
                if not is_holiday(holiday_list_name, current_date):
                    holiday.append(current_date)
                    working_days -= 1
                current_date = add_days(current_date, 1)
            doc.tat_completion_date=holiday[-1]
        elif doc.check_creation_date and not doc.insuff_cleared_on:
            from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
            holiday_list_name = 'TEAMPRO 2023 - Checkpro'
            start_date = doc.check_creation_date
            if doc.check_package:
                working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
            current_date = start_date
            holiday = []
            while working_days > 0:
                if not is_holiday(holiday_list_name, current_date):
                    holiday.append(current_date)
                    working_days -= 1
                current_date = add_days(current_date, 1)
            doc.tat_completion_date=holiday[-1]

@frappe.whitelist()
def update_holidays_daily_court():
    checks=frappe.db.get_all("Court",{"check_status":("!=",["Report Completed"])},["*"])
    for i in checks:
        doc=frappe.get_doc("Court",i.name)  
        if doc.check_creation_date and doc.check_completion_date:
            sql_query = f"""
            SELECT COUNT(*) 
            FROM `tabHoliday` 
            WHERE parent = 'TEAMPRO 2023 - Checkpro' 
            AND holiday_date BETWEEN '{doc.check_creation_date}' AND '{doc.check_completion_date}'
        """

            count = frappe.db.sql(sql_query, as_list=True)[0][0]
            frappe.db.set_value("Court",doc.name,"holidays",count)


@frappe.whitelist()
def update_tat_completion_date_reference_daily():
    checks=frappe.db.get_all("Reference Check",{"check_status":("!=",["Report Completed"])},["*"])
    for i in checks:
        doc=frappe.get_doc("Reference Check",i.name)
        if doc.insuff_cleared_on:
            from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
            holiday_list_name = 'TEAMPRO 2023 - Checkpro'
            start_date = doc.insuff_cleared_on
            working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
            current_date = start_date
            holiday = []
            while working_days > 0:
                if not is_holiday(holiday_list_name, current_date):
                    holiday.append(current_date)
                    working_days -= 1
                current_date = add_days(current_date, 1)
            frappe.db.set_value("Reference Check",doc.name,"tat_completion_date",holiday[-1])
        elif doc.check_creation_date and not doc.insuff_cleared_on:
            from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
            holiday_list_name = 'TEAMPRO 2023 - Checkpro'
            start_date = doc.check_creation_date
            if doc.check_package:
                working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
            current_date = start_date
            holiday = []
            while working_days > 0:
                if not is_holiday(holiday_list_name, current_date):
                    holiday.append(current_date)
                    working_days -= 1
                current_date = add_days(current_date, 1)
            frappe.db.set_value("Reference Check",doc.name,"tat_completion_date",holiday[-1])

@frappe.whitelist()
def update_holidays_daily_ref():
    checks=frappe.db.get_all("Reference Check",{"check_status":("!=",["Report Completed"])},["*"])
    for i in checks:
        doc=frappe.get_doc("Reference Check",i.name)  
        if doc.check_creation_date and doc.check_completion_date:
            sql_query = f"""
            SELECT COUNT(*) 
            FROM `tabHoliday` 
            WHERE parent = 'TEAMPRO 2023 - Checkpro' 
            AND holiday_date BETWEEN '{doc.check_creation_date}' AND '{doc.check_completion_date}'
        """

            count = frappe.db.sql(sql_query, as_list=True)[0][0]
            frappe.db.set_value("Reference Check",doc.name,"holidays",count)

@frappe.whitelist()
def update_tat_completion_date_id_daily():
    checks=frappe.db.get_all("Identity Aadhar",{"check_status":("!=",["Report Completed"])},["*"])
    for i in checks:
        doc=frappe.get_doc("Identity Aadhar",i.name)
        if doc.insuff_cleared_on:
            from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
            holiday_list_name = 'TEAMPRO 2023 - Checkpro'
            start_date = doc.insuff_cleared_on
            working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
            current_date = start_date
            holiday = []
            while working_days > 0:
                if not is_holiday(holiday_list_name, current_date):
                    holiday.append(current_date)
                    working_days -= 1
                current_date = add_days(current_date, 1)
            frappe.db.set_value("Identity Aadhar",doc.name,"tat_completion_date",holiday[-1])
        elif doc.check_creation_date and not doc.insuff_cleared_on:
            from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
            holiday_list_name = 'TEAMPRO 2023 - Checkpro'
            start_date = doc.check_creation_date
            working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
            current_date = start_date
            holiday = []
            while working_days > 0:
                if not is_holiday(holiday_list_name, current_date):
                    holiday.append(current_date)
                    working_days -= 1
                current_date = add_days(current_date, 1)
            frappe.db.set_value("Identity Aadhar",doc.name,"tat_completion_date",holiday[-1])

@frappe.whitelist()
def update_holidays_daily_iden():
    checks=frappe.db.get_all("Identity Aadhar",{"check_status":("!=",["Report Completed"])},["*"])
    for i in checks:
        doc=frappe.get_doc("Identity Aadhar",i.name)  
        if doc.check_creation_date and doc.check_completion_date:
            sql_query = f"""
            SELECT COUNT(*) 
            FROM `tabHoliday` 
            WHERE parent = 'TEAMPRO 2023 - Checkpro' 
            AND holiday_date BETWEEN '{doc.check_creation_date}' AND '{doc.check_completion_date}'
        """

            count = frappe.db.sql(sql_query, as_list=True)[0][0]
            frappe.db.set_value("Identity Aadhar",doc.name,"holidays",count)


@frappe.whitelist()
def update_tat_completion_date_sm_daily():
    checks=frappe.db.get_all("Social Media",{"check_status":("!=",["Report Completed"])},["*"])
    for i in checks:
        doc=frappe.get_doc("Social Media",i.name)
        if doc.insuff_cleared_on:
            from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
            holiday_list_name = 'TEAMPRO 2023 - Checkpro'
            start_date = doc.insuff_cleared_on
            working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
            current_date = start_date
            holiday = []
            while working_days > 0:
                if not is_holiday(holiday_list_name, current_date):
                    holiday.append(current_date)
                    working_days -= 1
                current_date = add_days(current_date, 1)
            frappe.db.set_value("Social Media",doc.name,"tat_completion_date",holiday[-1])
        elif doc.check_creation_date and not doc.insuff_cleared_on:
            from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
            holiday_list_name = 'TEAMPRO 2023 - Checkpro'
            start_date = doc.check_creation_date
            working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
            current_date = start_date
            holiday = []
            while working_days > 0:
                if not is_holiday(holiday_list_name, current_date):
                    holiday.append(current_date)
                    working_days -= 1
                current_date = add_days(current_date, 1)
            frappe.db.set_value("Social Media",doc.name,"tat_completion_date",holiday[-1])

@frappe.whitelist()
def update_holidays_daily_soc():
    checks=frappe.db.get_all("Social Media",{"check_status":("!=",["Report Completed"])},["*"])
    for i in checks:
        doc=frappe.get_doc("Social Media",i.name)  
        if doc.check_creation_date and doc.check_completion_date:
            sql_query = f"""
            SELECT COUNT(*) 
            FROM `tabHoliday` 
            WHERE parent = 'TEAMPRO 2023 - Checkpro' 
            AND holiday_date BETWEEN '{doc.check_creation_date}' AND '{doc.check_completion_date}'
        """

            count = frappe.db.sql(sql_query, as_list=True)[0][0]
            frappe.db.set_value("Social Media",doc.name,"holidays",count)


@frappe.whitelist()
def update_tat_completion_date_family_daily():
    checks=frappe.db.get_all("Family",{"check_status":("!=",["Report Completed"])},["*"])
    for i in checks:
        doc=frappe.get_doc("Family",i.name)
        if doc.insuff_cleared_on:
            from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
            holiday_list_name = 'TEAMPRO 2023 - Checkpro'
            start_date = doc.insuff_cleared_on
            working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
            current_date = start_date
            holiday = []
            while working_days > 0:
                if not is_holiday(holiday_list_name, current_date):
                    holiday.append(current_date)
                    working_days -= 1
                current_date = add_days(current_date, 1)
            frappe.db.set_value("Family",doc.name,"tat_completion_date",holiday[-1])
        elif doc.check_creation_date and not doc.insuff_cleared_on:
            from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
            holiday_list_name = 'TEAMPRO 2023 - Checkpro'
            start_date = doc.check_creation_date
            working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
            current_date = start_date
            holiday = []
            while working_days > 0:
                if not is_holiday(holiday_list_name, current_date):
                    holiday.append(current_date)
                    working_days -= 1
                current_date = add_days(current_date, 1)
            frappe.db.set_value("Family",doc.name,"tat_completion_date",holiday[-1])

@frappe.whitelist()
def update_holidays_daily_fam():
    checks=frappe.db.get_all("Family",{"check_status":("!=",["Report Completed"])},["*"])
    for i in checks:
        doc=frappe.get_doc("Family",i.name)  
        if doc.check_creation_date and doc.check_completion_date:
            sql_query = f"""
            SELECT COUNT(*) 
            FROM `tabHoliday` 
            WHERE parent = 'TEAMPRO 2023 - Checkpro' 
            AND holiday_date BETWEEN '{doc.check_creation_date}' AND '{doc.check_completion_date}'
        """

            count = frappe.db.sql(sql_query, as_list=True)[0][0]
            frappe.db.set_value("Family",doc.name,"holidays",count)

@frappe.whitelist()
def update_actual_tat_daily():
    list = ["Education Checks","Family","Reference Check","Court","Social Media","Criminal","Employment","Identity Aadhar","Address Check"]
    actual_tat=0
    tat = 0
    tat_monitor = ''
    date = 0
    dat = 0
    variation = 0
    for i in list:
        doc=frappe.db.get_list(i,["name","workflow_state","check_completion_date","check_creation_date","insufficiency_days","holidays","package_tat"])
        for j in doc:
            if(j.check_completion_date and j.workflow_state=="Report Completed"):
                date=(date_diff(doc.check_completion_date,doc.check_creation_date))+1
                dat=(sum([int(doc.insufficiency_days),int(doc.holidays)]))
                actual_tat=date - dat
                variation = int(actual_tat)-int(doc.package_tat)

                if variation < 0:
                    tat=0
                    tat_monitor = "In TAT"
                else:
                    tat=variation
                    tat_monitor = "Out TAT"
            # frappe.db.set_value(i,j.name,"actual_tat",actual_tat)
            # frappe.db.set_value(i,j.name,"tat_variation",tat)
            # frappe.db.set_value(i,j.name,"tat_monitor",tat_monitor)

@frappe.whitelist()
def nc_for_check_reject(name=None,id=None,allocated=None,class_proposed=None,reason=None):
    if allocated:
        emp_id=frappe.db.get_value("Employee",{'user_id':allocated},['name'])
        reopen_cause='(%s) Check :(%s) Rejected .Reason(%s)' % (name,id,reason)
        nc = frappe.new_doc('Energy Point And Non Conformity')
        nc.emp = emp_id
        nc.action='Non Conformity(NC)'
        nc.class_proposed = class_proposed
        nc.reason_of_ep = reopen_cause
        nc.save(ignore_permissions=True)
        frappe.db.commit()
        frappe.db.set_value("Energy Point And Non Conformity", nc.name, "workflow_state", "Explanation")
        frappe.db.commit()
        return {"status": "success", "message": "NC created"}

@frappe.whitelist()
def send_mail_nc_for_check_reject(name=None,id=None,allocated=None,class_proposed=None,reason=None):
    if allocated:
        emp_id=frappe.db.get_value("Employee",{'user_id':allocated},['name'])
        subject = _("{} - {} Rejected").format(name, id)
        message = """
            <p>Dear {},</p>
            <p><b>{} - {}</b> has been rejected.</p>
            <p><b>Reason:</b> {}</p>
            <p><b>NC Class:</b>{}</p>
            <p>Kindly review and take the necessary action.</p>
            <p>Best Regards,<br>TEAMPRO</p>
            """.format(emp_id,name, id, reason,class_proposed)

        frappe.sendmail(
            recipients=allocated,
            subject=subject,
            message=message
        )

import frappe
import os
import requests
from frappe.utils import get_files_path
from pdf2image import convert_from_path

@frappe.whitelist()
def convert_pdf_to_images(file_url):
    if not file_url:
        frappe.throw("No file URL provided")

    pdf_url = f"https://erp.teamproit.com{file_url}"
    
    frappe.logger().error(f"Fetching PDF from: {pdf_url}")

    output_dir = get_files_path("pdf_images", is_private=False)
    os.makedirs(output_dir, exist_ok=True)
    
    pdf_filename = os.path.basename(file_url).replace(' ', '_')
    pdf_path = os.path.join(output_dir, pdf_filename)

    try:
        response = requests.get(pdf_url, stream=True)
        response.raise_for_status()

        with open(pdf_path, "wb") as f:
            for chunk in response.iter_content(1024):
                f.write(chunk)

    except requests.exceptions.RequestException as e:
        frappe.throw(f"Failed to download PDF: {str(e)}")

    try:
        images = convert_from_path(pdf_path, dpi=150)
    except Exception as e:
        frappe.throw(f"PDF conversion error: {str(e)}")

    image_urls = []
    for i, img in enumerate(images):
        img_filename = f"{pdf_filename}_page_{i + 1}.png"
        img_path = os.path.join(output_dir, img_filename)
        img.save(img_path, "PNG")

        # Generate public URL for the image
        image_urls.append(f"/files/pdf_images/{img_filename}")

    return image_urls

import frappe

@frappe.whitelist()
def send_task_creation_email(doc,method):
    task_doc = frappe.get_doc("Task", doc.name)
    
    if task_doc.service == "IT-SW" and task_doc.type == "OPS":
        subject = f"New Task Created: {task_doc.name}"
        sub=task_doc.subject
        customer = task_doc.customer
        project = task_doc.project
        status = task_doc.status
        priority = task_doc.priority
        production_date = task_doc.custom_production_date.strftime("%d-%m-%Y") if task_doc.custom_production_date else ""
        allocation_date = task_doc.custom_allocated_on.strftime("%d-%m-%Y") if task_doc.custom_allocated_on else ""
        user = task_doc.custom_allocated_to

        recipients = []
        if task_doc.custom_development_spoc:
            recipients.append(task_doc.custom_development_spoc)
        # if task_doc.spoc:
        #     recipients.append(task_doc.spoc)
        if task_doc.project_manager:
            recipients.append(task_doc.project_manager)
        
        # Prepare HTML content with table
        email_body = f"""
        <p><strong>New Task Details:</strong></p>
        <table width='100%' style='border-collapse: collapse; border: 1px solid black; text-align: center;'>
            <thead>
                <tr style="background-color: #0f1568; color: white;">
                    <th colspan='2' style='text-align: center; background-color: #0f1568;color: white; font-size: 17px; border: 1px solid black;'><b>Details</b></th>
                </tr>
            </thead>
            <tbody>
                <tr style='text-align: left;'><td width='25%'style='border: 1px solid black;'><b>Task ID</b></td><td style='border: 1px solid black;'>{task_doc.name}</td></tr>
                <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Subject</b></td><td style='border: 1px solid black;'>{sub}</td></tr>
                <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Customer</b></td><td style='border: 1px solid black;'>{customer}</td></tr>
                <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Project</b></td><td style='border: 1px solid black;'>{project}</td></tr>
                <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Status</b></td><td style='border: 1px solid black;'>{status}</td></tr>
                <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Priority</b></td><td style='border: 1px solid black;'>{priority}</td></tr>
                <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Production Date</b></td><td style='border: 1px solid black;'>{production_date}</td></tr>
                <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Allocation Date</b></td><td style='border: 1px solid black;'>{allocation_date}</td></tr>
                <tr style='text-align: left;'><td style='border: 1px solid black;'><b>User</b></td><td style='border: 1px solid black;'>{user}</td></tr>
            </tbody>
        </table>
        """
        
        # Send email to recipients
        if recipients:
            frappe.sendmail(
                recipients=recipients,
                subject=subject,
                message=email_body
            )


import frappe
from frappe.utils import nowdate

from frappe.utils import nowdate
@frappe.whitelist()
def kt_email():
    
    tasks = frappe.get_all("Task", filters={"kt_confirmed": False, "service": "IT-SW", "type": "OPS", "status": "Working"}, fields=["name", "subject", "customer", "project", "status", "priority", "custom_production_date", "custom_allocated_on", "custom_allocated_to", "project_manager"])

    
    task_details = []
    serial_no = 1  # Initialize serial number

    for task_doc in tasks:
        task = frappe.get_doc("Task", task_doc.name)

      
        if task.service == "IT-SW" and task.type == "OPS" and task.status == "Working":
            production_date = task.custom_production_date.strftime("%d-%m-%Y") if task.custom_production_date else ""
            allocation_date = task.custom_allocated_on.strftime("%d-%m-%Y") if task.custom_allocated_on else ""
            user = task.custom_allocated_to

            # Append task details to the list with a serial number in columns
            task_details.append(f"""
            <tr style='text-align: left;'>
                <td style='border: 1px solid black; text-align: center; padding: 8px;'>{serial_no}</td>
                <td style='border: 1px solid black; text-align: left; padding: 8px;'>{task.name}</td>
                <td style='border: 1px solid black; padding: 8px;'>{task.subject}</td>
                <td style='border: 1px solid black; padding: 8px;'>{task.customer}</td>
                <td style='border: 1px solid black; padding: 8px;'>{task.project}</td>
                <td style='border: 1px solid black; padding: 8px;'>{task.status}</td>
                <td style='border: 1px solid black; padding: 8px;'>{task.priority}</td>
                <td style='border: 1px solid black; padding: 8px;'>{production_date}</td>
                <td style='border: 1px solid black; padding: 8px;'>{allocation_date}</td>
                <td style='border: 1px solid black; padding: 8px;'>{user}</td>
            </tr>
            """)

            serial_no += 1  # Increment serial number for the next task

    # If we have any tasks, send an email
    if task_details:
        # Email Subject
        subject = "KT Not Confirmed Tasks"

        # Prepare HTML content with all task details
        email_body = f"""
        <p><strong>KT Not Confirmed Tasks</strong></p>
        <table width='100%' style='border-collapse: collapse; border: 1px solid black; text-align: center;'>
            <thead>
                <tr style="background-color: #0f1568; color: white;">
                    <th style='text-align: center; background-color: #0f1568;color: white; font-size: 17px; border: 1px solid black; padding: 8px;'><b>S.No</b></th>
                    <th style='text-align: center; background-color: #0f1568;color: white; font-size: 17px; border: 1px solid black; padding: 8px;'><b>Task ID</b></th>
                    <th style='text-align: center; background-color: #0f1568;color: white; font-size: 17px; border: 1px solid black; padding: 8px;'><b>Subject</b></th>
                    <th style='text-align: center; background-color: #0f1568;color: white; font-size: 17px; border: 1px solid black; padding: 8px;'><b>Customer</b></th>
                    <th style='text-align: center; background-color: #0f1568;color: white; font-size: 17px; border: 1px solid black; padding: 8px;'><b>Project</b></th>
                    <th style='text-align: center; background-color: #0f1568;color: white; font-size: 17px; border: 1px solid black; padding: 8px;'><b>Status</b></th>
                    <th style='text-align: center; background-color: #0f1568;color: white; font-size: 17px; border: 1px solid black; padding: 8px;'><b>Priority</b></th>
                    <th style='text-align: center; background-color: #0f1568;color: white; font-size: 17px; border: 1px solid black; padding: 8px;'><b>Production Date</b></th>
                    <th style='text-align: center; background-color: #0f1568;color: white; font-size: 17px; border: 1px solid black; padding: 8px;'><b>Allocation Date</b></th>
                    <th style='text-align: center; background-color: #0f1568;color: white; font-size: 17px; border: 1px solid black; padding: 8px;'><b>Allocated To</b></th>
                </tr>
            </thead>
            <tbody>
                {''.join(task_details)}
            </tbody>
        </table>
        """
        

        
        frappe.sendmail(
            recipients=["abdulla.pi@groupteampro.com","siva.m@groupteampro.com","giftyp@groupteampro.com","jenisha.p@groupteampro.com"],
            subject=subject,
            message=email_body
        )


from frappe.utils import nowdate

@frappe.whitelist()
def sprint_task():
    # Get all tasks where the checkbox is unchecked (disabled)
    tasks = frappe.get_all("Task", filters={"service": "IT-SW", "type": "OPS", "status": "Working"}, fields=["name", "subject", "customer", "project", "status", "priority", "custom_production_date", "custom_allocated_on", "custom_allocated_to", "project_manager","custom_sprint","custom_dev_team"])
    
    task_details = []
    serial_no = 1  # Initialize serial number

    for task_doc in tasks:
        task = frappe.get_doc("Task", task_doc.name)

        # Get the active sprint for the task
        sprint = frappe.get_all('Task Sprint', filters={'Active': True, 'name': task.custom_sprint})
        
        if sprint:
            # Sprint exists for this task
            production_date = task.custom_production_date.strftime("%d-%m-%Y") if task.custom_production_date else ""
            allocation_date = task.custom_allocated_on.strftime("%d-%m-%Y") if task.custom_allocated_on else ""
            user = task.custom_allocated_to

            # Append task details to the list with a serial number in columns
            task_details.append(f"""
            <tr style='text-align: left;'>
                <td style='border: 1px solid black; text-align: center; padding: 8px;'>{serial_no}</td>
                <td style='border: 1px solid black; text-align: left; padding: 8px;'>{task.name}</td>
                <td style='border: 1px solid black; padding: 8px;'>{task.subject}</td>
                <td style='border: 1px solid black; padding: 8px;'>{task.customer}</td>
                <td style='border: 1px solid black; padding: 8px;'>{task.project}</td>
                <td style='border: 1px solid black; padding: 8px;'>{task.status}</td>
                <td style='border: 1px solid black; padding: 8px;'>{task.priority}</td>
                <td style='border: 1px solid black; padding: 8px;'>{production_date}</td>
                <td style='border: 1px solid black; padding: 8px;'>{allocation_date}</td>
                <td style='border: 1px solid black; padding: 8px;'>{user}</td>
            </tr>
            """)

            serial_no += 1  # Increment serial number for the next task

    # If we have any tasks, send an email
    if task_details:
        # Email Subject
        subject = "Sprint Active Tasks"

        # Prepare HTML content with all task details
        email_body = f"""
        <p><strong>Sprint Active Tasks</strong></p>
        <table width='100%' style='border-collapse: collapse; border: 1px solid black; text-align: center;'>
            <thead>
                <tr style="background-color: #0f1568; color: white;">
                    <th style='text-align: center; background-color: #0f1568;color: white; font-size: 17px; border: 1px solid black; padding: 8px;'><b>S.No</b></th>
                    <th style='text-align: center; background-color: #0f1568;color: white; font-size: 17px; border: 1px solid black; padding: 8px;'><b>Task ID</b></th>
                    <th style='text-align: center; background-color: #0f1568;color: white; font-size: 17px; border: 1px solid black; padding: 8px;'><b>Subject</b></th>
                    <th style='text-align: center; background-color: #0f1568;color: white; font-size: 17px; border: 1px solid black; padding: 8px;'><b>Customer</b></th>
                    <th style='text-align: center; background-color: #0f1568;color: white; font-size: 17px; border: 1px solid black; padding: 8px;'><b>Project</b></th>
                    <th style='text-align: center; background-color: #0f1568;color: white; font-size: 17px; border: 1px solid black; padding: 8px;'><b>Status</b></th>
                    <th style='text-align: center; background-color: #0f1568;color: white; font-size: 17px; border: 1px solid black; padding: 8px;'><b>Priority</b></th>
                    <th style='text-align: center; background-color: #0f1568;color: white; font-size: 17px; border: 1px solid black; padding: 8px;'><b>Production Date</b></th>
                    <th style='text-align: center; background-color: #0f1568;color: white; font-size: 17px; border: 1px solid black; padding: 8px;'><b>Allocation Date</b></th>
                    <th style='text-align: center; background-color: #0f1568;color: white; font-size: 17px; border: 1px solid black; padding: 8px;'><b>Allocated To</b></th>
                </tr>
            </thead>
            <tbody>
                {''.join(task_details)}
            </tbody>
        </table>
        """

        # Send email to the recipients (example recipient is hardcoded here)
        frappe.sendmail(
            recipients=["abdulla.pi@groupteampro.com"],
            subject=subject,
            message=email_body
        )


# @frappe.whitelist()
# def task_sprint_mail():
#     job = frappe.db.exists('Scheduled Job Type','sprint_task')
#     if not job:
#         task = frappe.new_doc("Scheduled Job Type")
#         task.update({
#             "method": 'checkpro.custom.sprint_task',
#             "frequency": 'Cron',
#             "cron_format": '00 7 * * *'
#         })
#         task.save(ignore_permissions=True)

@frappe.whitelist()
def task_sprint_mail_1():
    job = frappe.db.exists('Scheduled Job Type','dsr_task_mail_for_cmn_service')
    if not job:
        task = frappe.new_doc("Scheduled Job Type")
        task.update({
            "method": 'teampro.teampro.doctype.daily_monitor.dm_it_dev.dsr_task_mail_for_cmn_service',
            "frequency": 'Cron',
            "cron_format": '00 20 * * *'
        })
        task.save(ignore_permissions=True)


# @frappe.whitelist()
# def sprint_task():
    
#     tasks = frappe.get_all("Task", filters={"service": "IT-SW", "type": "OPS"}, fields=["name", "status","custom_sprint","custom_dev_team"])
    
#     for task_doc in tasks:
#         task = frappe.get_doc("Task", task_doc.name)

#         sprint = frappe.get_all('Task Sprint', filters={'Active': True, 'name': task.custom_sprint})

#         return sprint if sprint else None


# import frappe
# from frappe.utils.file_manager import get_file_path
# from frappe.utils.csvutils import read_csv_content

# @frappe.whitelist()
# def update_sales_follow_up_territory():
#     filename = "e7000bb087territory.csv"  # Ensure correct file name

#     try:
#         # Get full file path (supports private & public files)
#         filepath = get_file_path(filename)
        
#         # Read CSV content
#         csv_data = read_csv_content(filepath)
        
#         updated_count = 0  # Counter for updated records

#         for row in csv_data:
#             # Ensure the CSV has valid data (columns: Doc Name, Territory)
#             if len(row) < 2:
#                 frappe.log_error(f"Invalid data in row: {row}", "CSV Processing Error")
#                 continue

#             doc_name, territory = row[0], row[1]  # Extract values

#             try:
#                 # Update 'Sales Follow Up' document
#                 frappe.db.set_value("Sales Follow Up", {"name": doc_name}, "territory", territory)
#                 updated_count += 1
#             except Exception as e:
#                 frappe.log_error(f"Error updating Sales Follow Up {doc_name}: {str(e)}", "Sales Follow Up Update")

#         # Commit all updates at once for better performance
#         frappe.db.commit()
#         frappe.msgprint(f"{updated_count} records updated successfully.")

#     except Exception as e:
#         frappe.log_error(f"File error: {str(e)}", "Sales Follow Up File Error")
#         frappe.throw("Error processing the file. Please check the file name and format.")


# @frappe.whitelist()
# def emp_checkin_update():
#     filename = 'e7000bb087territory.csv'
#     from frappe.utils.file_manager import get_file
#     from frappe.utils.csvutils import read_csv_content
#     import frappe

#     # Get the file path
#     filepath = get_file(filename)[1]

#     # Read CSV content
#     pps = read_csv_content(filepath)
    
#     ind = 0  # Counter for updated records

#     for pp in pps:
#         # Ensure the CSV has valid data in expected columns
#         if len(pp) < 2:  # Adjusted to match the number of used fields
#             frappe.log_error(f"Invalid data in row: {pp}", "CSV Processing Error")
#             continue
        
#         try:
#             # Update the 'Sales Follow Up' document in one call
#             frappe.db.set_value("Sales Follow Up", {"name": pp[0]}, {
#                 "name": pp[0],
#                 "territory": pp[1]
#             })
#             ind += 1
#         except Exception as e:
#             # Log any errors during the update
#             frappe.log_error(f"Error updating Sales Follow Up {pp[0]}: {str(e)}", "Sales Follow Up Update Error")

#     frappe.db.commit()  # Commit the changes to the database
#     frappe.msgprint(f"{ind} records updated successfully.")




import frappe

def after_insert_employee_onboarding(doc, method):
    frappe.msgprint(f"Employee Onboarding Created for {doc.employee}")

    if doc.employee_onboarding_template:
        activities = frappe.get_all(
            "Employee Boarding Activity",
            fields=[
                "activity_name", "role", "user", "required_for_employee_creation",
                "description", "task_weight", "begin_on", "duration"
            ],
            filters={"parent": doc.employee_onboarding_template, "parenttype": "Employee Onboarding Template"},
            order_by="idx",
        )
        chc=frappe.get_all(
            "Employee Boarding Activity",
            fields=[
                "activity_name", "role", "user", "required_for_employee_creation",
                "description", "task_weight", "begin_on", "duration"
            ],
            filters={"parent": doc.custom_employee_chc_template, "parenttype": "Employee Onboarding Template"},
            order_by="idx",
        )
        for activity in activities:
            new_activity = doc.append("activities", {})
            new_activity.update(activity)
        for i in chc:
            new_chc=doc.append("custom_employee_chc", {})
            new_chc.update(i)
        job_offer=frappe.db.get_value("Job Offer",{"job_applicant":doc.job_applicant},["name"])
        doc.job_offer=job_offer
        doc.save()
        # frappe.msgprint(f"Activities added from template {doc.employee_onboarding_template}")


import frappe

def on_submit_employee_onboarding(doc, method):
    pending_activities = [i.activity_name for i in doc.activities if i.status == "Pending"]
    pending_custom_activities = [j.activity_name for j in doc.custom_employee_chc if j.status == "Pending"]
    if pending_activities or pending_custom_activities:
        pending_list = pending_activities + pending_custom_activities
        frappe.throw(f"The following activities are pending: {', '.join(pending_list)}. Kindly complete them before submission.")
    if frappe.db.exists("Employee Onboarding",{"employee": doc.employee,"docstatus": 1}):
        employee = frappe.get_doc("Employee", doc.employee)
        employee.workflow_state = "Joined"
        employee.save(ignore_permissions=True)
        frappe.msgprint(f"Employee {doc.employee} has been successfully set as active.")


def send_closure_report_with_table_dpr():
    next_date = datetime.today().date() + timedelta(days=1)
    formatted_date = next_date.strftime('%d-%m-%Y')
    filename1 = "Closure_Direct_" + formatted_date
    filename2 = "Closure_Indirect_" + formatted_date
    filename3 = "Closure_bdm_" + formatted_date
    xlsx_files = create_multiple_xlsx_closure_dpr()
    
    html_table, total_count , html_table_2, total_count_2, html_table_3, total_count_3 = closure_next_action_dpr()
    if total_count > 0 and total_count_2 > 0 and total_count_3 >0  :
        send_mail_with_attachment_and_html_dpr(html_table, html_table_2,html_table_3 ,filename1,filename2,filename3, xlsx_files)
    elif total_count > 0 and total_count_2 <= 0 and total_count_3 > 0 :
        send_mail_with_attachment_and_html_dpr(html_table,"",html_table_3, filename1,"",filename3, xlsx_files)
    elif total_count_2 > 0 and total_count <= 0 and total_count_3 >0 :
        send_mail_with_attachment_and_html_dpr("",html_table_2,html_table_3,"", filename2,filename3, xlsx_files)
    elif total_count_2 > 0 and total_count > 0 and total_count_3 <=0 :
        send_mail_with_attachment_and_html_dpr(html_table,html_table_2,"", filename1,filename2,"", xlsx_files)
    elif total_count_2 <= 0 and total_count > 0 and total_count_3 <=0 :
        send_mail_with_attachment_and_html_dpr(html_table,"","", filename1,"","", xlsx_files)
    elif total_count_2 > 0 and total_count <= 0 and total_count_3 <=0 :
        send_mail_with_attachment_and_html_dpr("",html_table_2,"","", filename2,"", xlsx_files)
    elif total_count_2 <= 0 and total_count <= 0 and total_count_3 > 0 :
        send_mail_with_attachment_and_html_dpr("","",html_table_3,"","", filename3, xlsx_files)
            
            

def send_mail_with_attachment_and_html_dpr(html_table = None , html_table_2 = None, html_table_3 = None, filename1 =None, filename2 =None, filename3 =None,file_content = None ):
    next_date_str = add_days(nowdate(), 1)

    
    next_date_obj = datetime.strptime(next_date_str, '%Y-%m-%d')

    
    formatted_date = next_date_obj.strftime('%d-%m-%Y')

    
    subject = "DND DPR - %s" % formatted_date
    message = (
        "Dear Sir/Madam,<br>"
        "Please find attached the attached Report based on Next Action.<br><br>"
        + html_table + "<br>"
        +html_table_2+"<br>"
        +html_table_3+
        "<br>Thanks & Regards,<br>TEAM ERP<br>"
        "This email has been automatically generated. Please do not reply"
    )
    if file_content:
        # if filename1 and filename2 and filename3:
        #     attachments = [
        #         {"fname": filename1 + '.xlsx', "fcontent": file_content[0].getvalue()},
        #         {"fname": filename2 + '.xlsx', "fcontent": file_content[1].getvalue()},
        #         {"fname": filename3 + '.xlsx', "fcontent": file_content[2].getvalue()},
        #     ]
        # elif filename1 and filename2 :
        #     attachments = [
        #         {"fname": filename1 + '.xlsx', "fcontent": file_content[0].getvalue()},
        #         {"fname": filename2 + '.xlsx', "fcontent": file_content[1].getvalue()},
        #     ]
        # elif filename1 and filename3 :
        #     attachments = [
        #         {"fname": filename1 + '.xlsx', "fcontent": file_content[0].getvalue()},
        #         {"fname": filename3 + '.xlsx', "fcontent": file_content[2].getvalue()},
        #     ]
        # elif filename2 and filename3 :
        #     attachments = [
        #         {"fname": filename2 + '.xlsx', "fcontent": file_content[1].getvalue()},
        #         {"fname": filename3 + '.xlsx', "fcontent": file_content[2].getvalue()},
        #     ]
        # else:
        #     attachments=[]
        
        attachments = []
        if filename1:
            attachments.append({"fname": filename1 + '.xlsx', "fcontent": file_content[0].getvalue()})
        if filename2:
            attachments.append({"fname": filename2 + '.xlsx', "fcontent": file_content[1].getvalue()})
        if filename3:
            attachments.append({"fname": filename3 + '.xlsx', "fcontent": file_content[2].getvalue()})
            
    
                
                
             
    frappe.sendmail(
        recipients=['divya.p@groupteampro.com','riyaz.a@groupteampro.com','dc@groupteampro.com','sangeetha.s@groupteampro.com','dineshbabu.k@groupteampro.com'],
        cc=['sangeetha.a@groupteampro.com'],
        # recipients=['dc@groupteampro.com','sangeetha.s@groupteampro.com','dineshbabu.k@groupteampro.com'],
        sender=None,
        subject=subject,
        message=message,
        attachments=attachments,
    )



# def make_xlsx_closure_dpr(filename, sheet_name=None, wb=None, column_widths=None, custom_conditions=None):
#     action = add_days(nowdate(), 1)
#     if wb is None:
#         wb = openpyxl.Workbook()
#     ws = wb.create_sheet(sheet_name or filename, 0)  
#     default_column_widths = [15, 35, 45, 25, 25, 40,25]
#     align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
#     bold_font = Font(bold=True)
#     column_widths = column_widths or default_column_widths    
#     for i, width in enumerate(column_widths, start=1):
#         ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width  
#     header_fill = PatternFill(start_color="87CEFA", end_color="87CEFA", fill_type="solid")
    
#     ws.merge_cells("A1:G1")
    
#     next_date_str = action  
#     formatted_date = datetime.strptime(next_date_str, '%Y-%m-%d').strftime('%d-%m-%Y')

    
    
#     filename1 = "Closure_Direct_" + formatted_date
#     filename2 = "Closure_Indirect_" + formatted_date
#     filename3 = "Closure_bdm_" + formatted_date
    
#     if filename == filename1:
#         ws["A1"]="Direct Follow Up"
#     elif filename == filename2:
#         ws["A1"]="In Direct Follow Up"
#     else:
#         ws["A1"]="BDM Follow Up"    
            
#     ws["A1"].fill = header_fill
#     ws["A1"].font = bold_font
#     ws["A1"].alignment = align_center
#     ws.append(["ID", "Candidate Name", "Customer", "Status", "Next Action", "Remark", "Next Action Date"])
#     for cell in ws[2]: 
#         cell.fill = header_fill
#         cell.font = bold_font
#         cell.alignment = align_center
#     # closures = frappe.get_all("Closure", {"custom_next_follow_up_on":action,'status':["Not In", ['Medical','Biometric','Signed offer','Ticket','Pre Medical','PCC','Emigration']]}, ['*'])
#     closures = frappe.get_all("Closure", custom_conditions, ['*'])
#     if closures:
#         for closure in closures:
            
#             next_action_date = ""
#             if closure.custom_next_follow_up_on:
#                 try:
#                     next_action_date = closure.custom_next_follow_up_on.strftime("%d-%m-%Y")
#                 except AttributeError:
                    
#                     next_action_date = datetime.strptime(str(closure.custom_next_follow_up_on), "%Y-%m-%d").strftime("%d-%m-%Y")
            
            
#             ws.append([closure.name, closure.given_name, closure.customer, closure.status, closure.std_remarks, closure.remark, next_action_date])
#     xlsx_file = BytesIO()
#     wb.save(xlsx_file)
#     xlsx_file.seek(0)    
#     return xlsx_file


def make_xlsx_closure_dpr(filename, sheet_name=None, wb=None, column_widths=None, custom_conditions=None):
    action = add_days(nowdate(), 1)

    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name or filename, 0)  

    default_column_widths = [15, 35, 45, 25, 25, 40, 25]
    column_widths = column_widths or default_column_widths    

    # === Styles ===
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="87CEFA", end_color="87CEFA", fill_type="solid")

    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width  

    ws.merge_cells("A1:G1")
    
    # === Title Row ===
    formatted_date = datetime.strptime(action, '%Y-%m-%d').strftime('%d-%m-%Y')
    filename1 = "Closure_Direct_" + formatted_date
    filename2 = "Closure_Indirect_" + formatted_date
    filename3 = "Closure_bdm_" + formatted_date
    
    if filename == filename1:
        ws["A1"] = "Direct Follow Up"
    elif filename == filename2:
        ws["A1"] = "In Direct Follow Up"
    else:
        ws["A1"] = "BDM Follow Up"    

    ws["A1"].fill = header_fill
    ws["A1"].font = bold_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # === Table Header ===
    ws.append(["ID", "Candidate Name", "Customer", "Status", "Next Action", "Remark", "Next Action Date"])
    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = align_center

    # === Data Rows ===
    closures = frappe.get_all("Closure", custom_conditions, ['*'])
    if closures:
        for closure in closures:
            next_action_date = ""
            if closure.custom_next_follow_up_on:
                try:
                    next_action_date = closure.custom_next_follow_up_on.strftime("%d-%m-%Y")
                except AttributeError:
                    next_action_date = datetime.strptime(str(closure.custom_next_follow_up_on), "%Y-%m-%d").strftime("%d-%m-%Y")

            ws.append([
                closure.name,
                closure.given_name,
                closure.customer,
                closure.status,
                closure.std_remarks,
                closure.remark,
                next_action_date
            ])

    # === Apply Border, Height & Wrap ===
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=7):
          
        for cell in row:
            cell.border = thin_border
            
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=7):
         
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[cell.row].height = 30

    

    # === Save Workbook ===
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)    
    return xlsx_file



def create_multiple_xlsx_closure_dpr():
    action_date = add_days(nowdate(), 1)
    conditions_file1 = {"custom_next_follow_up_on": action_date,'stamping_vendor':("is","not set"),"sa_id": ("is", "not set"),'status':("In", ['Final Medical','Biometric','Signed Offer Letter','Ticket','Premedical','PCC','Emigration'])}
    
    conditions_file2 = {"custom_next_follow_up_on": action_date,'stamping_vendor':("is","set"),"sa_id": ("is", "set"),'status':("In", ['Final Medical','Biometric','Signed Offer Letter','Ticket','Premedical','PCC','Emigration'])}
    
    conditions_file3 = {"custom_next_follow_up_on": action_date,'status':("In", ['Visa','Client Offer Letter','Ticket'])}
    
    next_date_str = action_date  
    formatted_date = datetime.strptime(next_date_str, '%Y-%m-%d').strftime('%d-%m-%Y')

    filename1 = "Closure_Direct_" + formatted_date
    filename2 = "Closure_Indirect_" + formatted_date
    filename3 = "Closure_bdm_" + formatted_date 
    file1 = make_xlsx_closure_dpr(filename1, custom_conditions=conditions_file1)
    file2 = make_xlsx_closure_dpr(filename2, custom_conditions=conditions_file2)
    file3 = make_xlsx_closure_dpr(filename3, custom_conditions=conditions_file3)
    
    
    return [file1, file2, file3]

def closure_next_action_dpr():
    
    records_to_delete = frappe.get_all("DND DPR Records", ["name"])
    for record in records_to_delete:
        frappe.delete_doc("DND DPR Records", record.name, ignore_permissions=True)
    
    action_date = add_days(nowdate(), 1)
    #Direct Follow up
    closures = frappe.get_all("Closure", {"custom_next_follow_up_on": action_date,'stamping_vendor':("is","not set"),"sa_id": ("is", "not set"),'status':("In", ['Final Medical','Biometric','Signed Offer Letter','Ticket','Premedical','PCC','Emigration'])}, ["customer", "status","name"])
    customer_status_count = {}
    for closure in closures:
        customer = closure.customer
        status = closure.status
        name = closure.name
        if customer not in customer_status_count:
            customer_status_count[customer] = {}
        if status not in customer_status_count[customer]:
            customer_status_count[customer][status] = []
        customer_status_count[customer][status].append(name)
    table = '<table text-align="center" border="1" width="100%" style="border-collapse: collapse;text-align: center;">'
    table += '<tr style="background-color: #87CEFA"><td colspan="3" style=" font-weight: bold; text-align: center;">Direct Follow Up</td></tr>'
    table += '<tr style="background-color: #87CEFA"><td style="width: 45%; font-weight: bold; text-align: center;">Customer</td><td style="width: 30%; font-weight: bold; text-align: center;">Status</td><td style="width: 25%; font-weight: bold; text-align: center;">Count</td></tr>'
    for customer, statuses,  in customer_status_count.items():
        # total_counts = sum(statuses.values())
        total_counts = sum(len(ids) for ids in statuses.values()) 
        table += '<tr><td><b>%s</b></td><td></td><td><b>%s</b></td></tr>' % (customer, total_counts)        
        for status, closure_ids  in statuses.items():
            count = len(closure_ids)
            table += '<tr><td></td><td>%s</td><td>%s</td></tr>' % (status, count)
            for i in closure_ids:
                # print(f"[DUPLICATE SKIPPED] Closure ID: {i}")
                doc_1 = frappe.new_doc("DND DPR Records")
                doc_1.closure_id = i
                doc_1.dpr_date = action_date
                doc_1.customer = customer
                doc_1.status = status
                doc_1.count = 1
                doc_1.follow_up = "Direct Follow Up"
                doc_1.insert(ignore_permissions=True)
                frappe.db.commit()
            
    table += '</table>'
    # total_count = sum(sum(status.values()) for status in customer_status_count.values())
    total_count = sum(len(ids) for statuses in customer_status_count.values() for ids in statuses.values())

    
    #InDirect Follow up
    closures_indirect = frappe.get_all("Closure", {"custom_next_follow_up_on": action_date,'stamping_vendor':("is","set"),"sa_id": ("is", "set"),'status':("In", ['Final Medical','Biometric','Signed Offer Letter','Ticket','Premedical','PCC','Emigration'])}, ["customer", "status"])
    customer_status_count_indirect = {}
    for closure in closures_indirect:
        customer = closure.customer
        status = closure.status
        if customer not in customer_status_count_indirect:
            customer_status_count_indirect[customer] = {}
        if status not in customer_status_count_indirect[customer]:
            customer_status_count_indirect[customer][status] = []
        customer_status_count_indirect[customer][status].append(name) 
    table_2 = '<table text-align="center" border="1" width="100%" style="border-collapse: collapse;text-align: center;">'
    table_2 += '<tr style="background-color: #87CEFA"><td colspan="3" style=" font-weight: bold; text-align: center;">InDirect Follow Up</td></tr>'
    table_2 += '<tr style="background-color: #87CEFA"><td style="width: 45%; font-weight: bold; text-align: center;">Customer</td><td style="width: 30%; font-weight: bold; text-align: center;">Status</td><td style="width: 25%; font-weight: bold; text-align: center;">Count</td></tr>'
    for customer, statuses in customer_status_count_indirect.items():
        # total_counts_indirect = sum(statuses.values())
        total_counts_indirect = sum(len(ids) for ids in statuses.values()) 
        table_2 += '<tr><td><b>%s</b></td><td></td><td><b>%s</b></td></tr>' % (customer, total_counts_indirect)        
        for status, closure_ids  in statuses.items():
            count = len(closure_ids)
            table_2 += '<tr><td></td><td>%s</td><td>%s</td></tr>' % (status, count)
            for i in closure_ids:
                doc_2 = frappe.new_doc("DND DPR Records")
                doc_2.closure_id = i
                doc_2.dpr_date = action_date
                doc_2.customer = customer
                doc_2.status = status
                doc_2.count = count
                doc_2.follow_up = "InDirect Follow Up"
                doc_2.insert(ignore_permissions=True)
            frappe.db.commit()
    table_2 += '</table>'
    # total_count_indirect = sum(sum(status.values()) for status in customer_status_count_indirect.values())
    total_count_indirect = sum(len(ids) for statuses in customer_status_count_indirect.values() for ids in statuses.values())
    
    #BDM
    closures_bdm = frappe.get_all("Closure", {"custom_next_follow_up_on": action_date,'status':("In", ['Visa','Client Offer Letter','Ticket'])}, ["customer", "status"])
    customer_status_count_bdm = {}
    for closure in closures_bdm:
        customer = closure.customer
        status = closure.status
        if customer not in customer_status_count_bdm:
            customer_status_count_bdm[customer] = {}
        if status not in customer_status_count_bdm[customer]:
            customer_status_count_bdm[customer][status] = []
        customer_status_count_bdm[customer][status].append(name)
    table_3 = '<table text-align="center" border="1" width="100%" style="border-collapse: collapse;text-align: center;">'
    table_3 += '<tr style="background-color: #87CEFA"><td colspan="3" style=" font-weight: bold; text-align: center;">BDM Follow Up</td></tr>'
    table_3 += '<tr style="background-color: #87CEFA"><td style="width: 45%; font-weight: bold; text-align: center;">Customer</td><td style="width: 30%; font-weight: bold; text-align: center;">Status</td><td style="width: 25%; font-weight: bold; text-align: center;">Count</td></tr>'
    for customer, statuses in customer_status_count_bdm.items():
        # total_counts_bdm = sum(statuses.values())
        total_counts_bdm = sum(len(ids) for ids in statuses.values())
        table_3 += '<tr><td><b>%s</b></td><td></td><td><b>%s</b></td></tr>' % (customer, total_counts_bdm)        
        for status, closure_ids in statuses.items():
            count = len(closure_ids)
            table_3 += '<tr><td></td><td>%s</td><td>%s</td></tr>' % (status, count)
            for i in closure_ids:
                doc_3 = frappe.new_doc("DND DPR Records")
                doc_3.closure_id = i
                doc_3.dpr_date = action_date
                doc_3.customer = customer
                doc_3.status = status
                doc_3.count = count
                doc_3.follow_up = "BDM Follow Up"
                doc_3.insert(ignore_permissions=True)
                frappe.db.commit()
    table_3 += '</table>'
    # total_count_bdm = sum(sum(status.values()) for status in customer_status_count_indirect.values())
    total_count_bdm = sum(len(ids) for statuses in customer_status_count_bdm.values() for ids in statuses.values())
    
    
    return table, total_count ,table_2, total_count_indirect , table_3, total_count_bdm


@frappe.whitelist()
def create_schedule_job_dpr():
    job = frappe.db.exists('Scheduled Job Type', 'send_closure_report_with_table_dpr')
    if not job:
        exp = frappe.new_doc("Scheduled Job Type")
        exp.update({
            "method": 'checkpro.custom.send_closure_report_with_table_dpr',
            "frequency": 'Cron',
            "cron_format": "5 21 * * *"
        })
        exp.save(ignore_permissions=True)


@frappe.whitelist()
def get_vpi_details(name, ins):
    unique_details = set()
    result = []

    edu_checks = frappe.get_all('Education Checks', 
        filters={'custom_institute': ins, 'name': ['!=', name]},
        fields=['name']
    )
    frappe.errprint('HIII')
    frappe.errprint(edu_checks)
    for edu in edu_checks:
        vpi_details = frappe.get_all('Institute Details',
            filters={'parent': edu.name, 'parenttype': 'Education Checks'},
            fields=['verified_by', 'name1', 'designation', 'contact']
        )
        for vpi in vpi_details:
            frappe.errprint(vpi_details)
            detail_tuple = (
                vpi.verified_by or '',
                vpi.name1 or '',
                vpi.designation or '',
                vpi.contact or ''
            )

            if any(detail_tuple): 
                if detail_tuple not in unique_details:
                    unique_details.add(detail_tuple)
                    result.append({
                        'verified_by': vpi.verified_by,
                        'name1': vpi.name1,
                        'designation': vpi.designation,
                        'contact': vpi.contact
                    })

    return result

# @frappe.whitelist()
# def update_emp_check_sts():
#     frappe.db.set_value('Employment','Employment-4477','drop',0)
#     frappe.db.set_value("Employment",'Employment-4477',"workflow_state","Draft")
#     frappe.db.set_value("Employment",'Employment-4477',"check_status","Draft")
#     frappe.db.set_value("Employment",'Employment-4477',"report_status","YTS")

# onsubmit Leave
@frappe.whitelist()
def update_session_leave(doc,method):
    if doc.half_day_date and doc.custom_session:
        attendance=frappe.db.get_value("Attendance",{"employee":doc.employee,"attendance_date":doc.half_day_date},["name"])
        if attendance:
            frappe.db.set_value("Attendance",attendance,"custom_session",doc.custom_session)

# oncancel Leave
@frappe.whitelist()
def update_session_leave_cancel(doc,method):
    if doc.half_day_date and doc.custom_session:
        attendance=frappe.db.get_value("Attendance",{"employee":doc.employee,"attendance_date":doc.half_day_date},["name"])
        if attendance:
            frappe.db.set_value("Attendance",attendance,"custom_session",'')

# onsubmit Attendance Request
@frappe.whitelist()
def update_session_ar(doc,method):
    if doc.half_day_date and doc.custom_session:
        attendance=frappe.db.get_value("Attendance",{"employee":doc.employee,"attendance_date":doc.half_day_date},["name"])
        if attendance:
            frappe.db.set_value("Attendance",attendance,"custom_session",doc.custom_session)

# oncancel Attendance Request
@frappe.whitelist()
def update_session_ar_cancel(doc,method):
    if doc.half_day_date and doc.custom_session:
        attendance=frappe.db.get_value("Attendance",{"employee":doc.employee,"attendance_date":doc.half_day_date},["name"])
        if attendance:
            frappe.db.set_value("Attendance",attendance,"custom_session",'')

# @frappe.whitelist()
# def update_cs_status():
#     frappe.db.set_value("Employment","Employment-2250","report_status","YTS")
#     frappe.db.set_value("Employment","Employment-2249","report_status","YTS")

# @frappe.whitelist()
# def update_candidate_status():
#     candidates=frappe.db.get_all("Candidate",{"pending_for":"QC Cleared"},["name"])
#     ind=0
#     for i in candidates:
#         frappe.db.set_value("Candidate",i.name,"pending_for","Submit(SPOC)")
#         ind+=1
#     print(ind)


#######___DSR___#######

def send_closure_report_with_table_dsr():
    today_date = datetime.today().date()
    formatted_date = today_date.strftime('%d-%m-%Y')
    filename1 = "Closure_Direct_" + formatted_date
    filename2 = "Closure_Indirect_" + formatted_date
    filename3 = "Closure_bdm_" + formatted_date
    xlsx_files = create_multiple_xlsx_closure_dsr()
    
    html_table, total_count , html_table_2, total_count_2, html_table_3, total_count_3 = closure_next_action_dsr()
    if total_count > 0 and total_count_2 > 0 and total_count_3 >0  :
        send_mail_with_attachment_and_html_dsr(html_table, html_table_2,html_table_3 ,filename1,filename2,filename3, xlsx_files)
    elif total_count > 0 and total_count_2 <= 0 and total_count_3 > 0 :
        send_mail_with_attachment_and_html_dsr(html_table,"",html_table_3, filename1,"",filename3, xlsx_files)
    elif total_count_2 > 0 and total_count <= 0 and total_count_3 >0 :
        send_mail_with_attachment_and_html_dsr("",html_table_2,html_table_3,"", filename2,filename3, xlsx_files)
    elif total_count_2 > 0 and total_count > 0 and total_count_3 <=0 :
        send_mail_with_attachment_and_html_dsr(html_table,html_table_2,"", filename1,filename2,"", xlsx_files)
    elif total_count_2 <= 0 and total_count > 0 and total_count_3 <=0 :
        send_mail_with_attachment_and_html_dsr(html_table,"","", filename1,"","", xlsx_files)
    elif total_count_2 > 0 and total_count <= 0 and total_count_3 <=0 :
        send_mail_with_attachment_and_html_dsr("",html_table_2,"","", filename2,"", xlsx_files)
    elif total_count_2 <= 0 and total_count <= 0 and total_count_3 > 0 :
        send_mail_with_attachment_and_html_dsr("","",html_table_3,"","", filename3, xlsx_files)
            
            

def send_mail_with_attachment_and_html_dsr(html_table = None , html_table_2 = None, html_table_3 = None, filename1 =None, filename2 =None, filename3 =None,file_content = None ):
    date_str = nowdate()  

    
    date_obj = datetime.strptime(date_str, '%Y-%m-%d')

    
    formatted_date = date_obj.strftime('%d-%m-%Y')

    
    subject = "DND DSR - %s" % formatted_date
    
    
    message = (
        "Dear Sir/Madam,<br>"
        "Please find attached the attached Report based on Next Action.<br><br>"
        + html_table + "<br>"
        +html_table_2+"<br>"
        +html_table_3+
        "<br>Thanks & Regards,<br>TEAM ERP<br>"
        "This email has been automatically generated. Please do not reply"
    )
    if file_content:
        
        
        attachments = []
        if filename1:
            attachments.append({"fname": filename1 + '.xlsx', "fcontent": file_content[0].getvalue()})
        if filename2:
            attachments.append({"fname": filename2 + '.xlsx', "fcontent": file_content[1].getvalue()})
        if filename3:
            attachments.append({"fname": filename3 + '.xlsx', "fcontent": file_content[2].getvalue()})

            
    
                
                
             
    frappe.sendmail(
        
        recipients=['divya.p@groupteampro.com','riyaz.a@groupteampro.com','dc@groupteampro.com','sangeetha.s@groupteampro.com','dineshbabu.k@groupteampro.com'],
        cc=['sangeetha.a@groupteampro.com'],
        # recipients=['riyaz.a@groupteampro.com'],
        sender=None,
        subject=subject,
        message=message,
        attachments=attachments,
    )




# def make_xlsx_closure_dsr(filename, sheet_name=None, wb=None, column_widths=None, custom_conditions=None):
#     action = nowdate()

#     if wb is None:
#         wb = openpyxl.Workbook()
#     ws = wb.create_sheet(sheet_name or filename, 0)

#     default_column_widths = [45, 25, 20, 30, 25,20,20]
#     column_widths = column_widths or default_column_widths

#     align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
#     bold_font = Font(bold=True)
#     header_fill = PatternFill(start_color="87CEFA", end_color="87CEFA", fill_type="solid")

#     for i, width in enumerate(column_widths, start=1):
#         ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

#     ws.merge_cells("A1:G1")
#     ws["A1"] = (
#         "Direct Follow Up" if "Direct" in filename
#         else "In Direct Follow Up" if "Indirect" in filename
#         else "BDM Follow Up"
#     )
#     ws["A1"].fill = header_fill
#     ws["A1"].font = bold_font
#     ws["A1"].alignment = align_center

#     # Table Header
#     ws.append([
#         "Customer",
#         "Closure Status", "Closure Count", "Closure IDs",
#         "DPR Status", "DPR Count", "DPR IDs"
#     ])
#     for cell in ws[2]:
#         cell.fill = header_fill
#         cell.font = bold_font
#         cell.alignment = align_center

#     # Get Closure Records
#     closures = frappe.get_all("Closure", custom_conditions, ["name", "customer", "status"])

#     # Organize closure data by customer
#     closure_by_customer = {}
#     for c in closures:
#         closure_by_customer.setdefault(c.customer, {}).setdefault(c.status, []).append(c.name)

#     for customer, closure_statuses in closure_by_customer.items():
#         # Determine follow_up_type from filename
#         follow_up_type = (
#             "Direct Follow Up" if "Direct" in filename
#             else "InDirect Follow Up" if "Indirect" in filename
#             else "BDM Follow Up"
#         )

#         dpr_records = frappe.get_all("DND DPR Records", {
#             "dpr_date": action,
#             "customer": customer,
#             "follow_up": follow_up_type
#         }, ["name", "status"])

#         # Organize DPR records
#         dpr_statuses = {}
#         for d in dpr_records:
#             dpr_statuses.setdefault(d.status, []).append(d.name)

#         closure_items = list(closure_statuses.items())
#         dpr_items = list(dpr_statuses.items())
#         max_rows = max(len(closure_items), len(dpr_items))

#         for i in range(max_rows):
#             # Closure info
#             closure_status, closure_ids = closure_items[i] if i < len(closure_items) else ("", [])
#             closure_count = len(closure_ids)
#             closure_ids_str = ", ".join(closure_ids)

#             # DPR info
#             dpr_status, dpr_ids = dpr_items[i] if i < len(dpr_items) else ("", [])
#             dpr_count = len(dpr_ids)
#             dpr_ids_str = ", ".join(dpr_ids)

#             row = [
#                 customer if i == 0 else "",
#                 closure_status,
#                 closure_count,
#                 closure_ids_str,
#                 dpr_status,
#                 dpr_count,
#                 dpr_ids_str
#             ]
#             ws.append(row)
        
        
        
#     thin_border = Border(
#         left=Side(style='thin', color='000000'),
#         right=Side(style='thin', color='000000'),
#         top=Side(style='thin', color='000000'),
#         bottom=Side(style='thin', color='000000')
#     )
#     for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=7):
#         for cell in row:
#             cell.border = thin_border    

#     # Save the workbook to a BytesIO object
#     xlsx_file = BytesIO()
#     wb.save(xlsx_file)
#     xlsx_file.seek(0)
#     return xlsx_file


def make_xlsx_closure_dsr(filename, sheet_name=None, wb=None, column_widths=None, custom_conditions=None):
    action = nowdate()

    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name or filename, 0)

    default_column_widths = [45, 25, 20, 30, 25, 20, 20]
    column_widths = column_widths or default_column_widths

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="87CEFA", end_color="87CEFA", fill_type="solid")

    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    ws.merge_cells("A1:G1")
    ws["A1"] = (
        "Direct Follow Up" if "Direct" in filename
        else "In Direct Follow Up" if "Indirect" in filename
        else "BDM Follow Up"
    )
    ws["A1"].fill = header_fill
    ws["A1"].font = bold_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Table Header
    ws.append([
        "Customer",
        "Closure Status", "Closure Count", "Closure IDs",
        "DPR Status", "DPR Count", "DPR IDs"
    ])
    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = align_center

    # Get Closure Records
    closures = frappe.get_all("Closure", custom_conditions, ["name", "customer", "status"])

    # Organize closure data by customer
    closure_by_customer = {}
    for c in closures:
        closure_by_customer.setdefault(c.customer, {}).setdefault(c.status, []).append(c.name)

    # ✅ Collect all customers from both Closure and DPR
    all_customers = set(closure_by_customer.keys())
    follow_up_type = (
        "Direct Follow Up" if "Direct" in filename
        else "InDirect Follow Up" if "Indirect" in filename
        else "BDM Follow Up"
    )

    for d in frappe.get_all("DND DPR Records", {
        "dpr_date": action,
        "follow_up": follow_up_type
    }, ["customer"]):
        all_customers.add(d.customer)

    # ✅ Now loop through all customers
    for customer in all_customers:
        closure_statuses = closure_by_customer.get(customer, {})

        dpr_records = frappe.get_all("DND DPR Records", {
            "dpr_date": action,
            "customer": customer,
            "follow_up": follow_up_type
        }, ["name", "status"])

        # Organize DPR records
        dpr_statuses = {}
        for d in dpr_records:
            dpr_statuses.setdefault(d.status, []).append(d.name)

        closure_items = list(closure_statuses.items())
        dpr_items = list(dpr_statuses.items())
        max_rows = max(len(closure_items), len(dpr_items))

        for i in range(max_rows):
            # Closure info
            closure_status, closure_ids = closure_items[i] if i < len(closure_items) else ("", [])
            closure_count = len(closure_ids)
            closure_ids_str = ", ".join(closure_ids)

            # DPR info
            dpr_status, dpr_ids = dpr_items[i] if i < len(dpr_items) else ("", [])
            dpr_count = len(dpr_ids)
            dpr_ids_str = ", ".join(dpr_ids)

            ws.append([
                customer if i == 0 else "",
                closure_status,
                closure_count,
                closure_ids_str,
                dpr_status,
                dpr_count,
                dpr_ids_str
            ])

    # ✅ Add thin border to all cells
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=7):
        for cell in row:
            cell.border = thin_border
            
            
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[cell.row].height = 30

    # Save workbook
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file



def create_multiple_xlsx_closure_dsr():
    action_date = nowdate()
    
    # action_date = "31-10-2025"
    
    conditions_file1 = {"last_updated_on": action_date,'stamping_vendor':("is", "not set"),"sa_id": ("is", "not set"),'status':("In", ['Final Medical','Biometric','Signed Offer Letter','Ticket','Premedical','PCC','Emigration'])}
    
    conditions_file2 = {"last_updated_on": action_date,'stamping_vendor':("is", "set"),"sa_id": ("is", "set"),'status':("In", ['Final Medical','Biometric','Signed Offer Letter','Ticket','Premedical','PCC','Emigration'])}
    
    conditions_file3 = {"last_updated_on": action_date,'status':("In", ['Visa','Client Offer Letter','Ticket'])}

    filename1 = "Closure_Direct_" + today()
    filename2 = "Closure_Indirect_" + today()
    filename3 = "Closure_bdm_" + today() 
    file1 = make_xlsx_closure_dsr(filename1, custom_conditions=conditions_file1)
    file2 = make_xlsx_closure_dsr(filename2, custom_conditions=conditions_file2)
    file3 = make_xlsx_closure_dsr(filename3, custom_conditions=conditions_file3)
    
    
    return [file1, file2, file3]


def closure_next_action_dsr():
    
    action_date = nowdate()
    # action_date = "31-10-2025"
    
    #Direct2
    # Fetch Closure records for Direct Follow Up
    closures = frappe.get_all(
        "Closure",
        {
            "last_updated_on": action_date,
            'stamping_vendor': ("is", "not set"),
            "sa_id": ("is", "not set"),
            'status': ("in", ['Final Medical', 'Biometric', 'Signed Offer Letter', 'Ticket', 'Premedical', 'PCC', 'Emigration'])
        },
        ["customer", "status", "name"]
    )

    # Fetch DND DPR Records for Direct Follow Up
    dnd_dpr_records = frappe.get_all(
        "DND DPR Records",
        {
            "dpr_date": action_date,
            "follow_up": "Direct Follow Up"
        },
        ["customer", "status", "name"]
    )

    # Group closure records by customer
    closure_data_by_customer = {}
    for c in closures:
        closure_data_by_customer.setdefault(c.customer, []).append(c)

    # Group DND DPR records by customer
    dpr_data_by_customer = {}
    for d in dnd_dpr_records:
        dpr_data_by_customer.setdefault(d.customer, []).append(d)

    # Get union of all customers
    all_customers = set(closure_data_by_customer.keys()).union(dpr_data_by_customer.keys())

    # Start HTML table
    table = '<table text-align="center" border="1" width="100%" style="border-collapse: collapse;text-align: center;">'
    table += '<tr style="background-color: #87CEFA"><td colspan="5" style=" font-weight: bold; text-align: center;">Direct Follow Up</td></tr>'
    table += '<tr style="background-color: #87CEFA">'
    table += '<td style="width: 20%; font-weight: bold; text-align: center;">Customer</td>'
    table += '<td style="width: 20%; font-weight: bold; text-align: center;">Closure Status</td>'
    table += '<td style="width: 10%; font-weight: bold; text-align: center;">Closure Count</td>'
    table += '<td style="width: 20%; font-weight: bold; text-align: center;">DPR Status</td>'
    table += '<td style="width: 10%; font-weight: bold; text-align: center;">DPR Count</td>'
    table += '</tr>'

    total_count = 0

    for customer in all_customers:
        closure_statuses = {}
        for closure in closure_data_by_customer.get(customer, []):
            closure_statuses.setdefault(closure.status, []).append(closure.name)

        dpr_statuses = {}
        for dpr in dpr_data_by_customer.get(customer, []):
            dpr_statuses.setdefault(dpr.status, []).append(dpr.name)

        closure_items = list(closure_statuses.items())
        dpr_items = list(dpr_statuses.items())
        max_rows = max(len(closure_items), len(dpr_items))

        for i in range(max_rows):
            closure_row = closure_items[i] if i < len(closure_items) else ("", [])
            dpr_row = dpr_items[i] if i < len(dpr_items) else ("", [])

            closure_status, closure_ids = closure_row
            dpr_status, dpr_ids = dpr_row

            # Add total closure count
            if closure_status:
                total_count += len(closure_ids)

            table += (
                f"<tr>"
                f"<td>{customer if i == 0 else ''}</td>"
                f"<td>{closure_status}</td><td>{len(closure_ids)}</td>"
                f"<td>{dpr_status}</td><td>{len(dpr_ids)}</td>"
                f"</tr>"
            )

    table += '</table>'

    # Log total closure count
    frappe.log_error(message=str(total_count), title="Total Direct Closure Count")

    

    
        
    #InDirect2
    
    # Indirect Follow up
    closures_indirect = frappe.get_all("Closure", {
        "last_updated_on": action_date,
        'stamping_vendor': ("is", "set"),
        "sa_id": ("is", "set"),
        'status': ("In", ['Final Medical', 'Biometric', 'Signed Offer Letter', 'Ticket', 'Premedical', 'PCC', 'Emigration'])
    }, ["customer", "status", "name"])

    dnd_dpr_records_indirect = frappe.db.get_all("DND DPR Records", {
        "dpr_date": action_date,
        "follow_up": "InDirect Follow Up"
    }, ["customer", "status", "name"])

    # Organize data
    closure_data_by_customer_indirect = {}
    dpr_data_by_customer_indirect = {}

    for c in closures_indirect:
        closure_data_by_customer_indirect.setdefault(c.customer, []).append(c)

    for d in dnd_dpr_records_indirect:
        dpr_data_by_customer_indirect.setdefault(d.customer, []).append(d)

    customers_indirect = set(closure_data_by_customer_indirect.keys()).union(dpr_data_by_customer_indirect.keys())

    table_2 = '<table text-align="center" border="1" width="100%" style="border-collapse: collapse;text-align: center;">'
    table_2 += '<tr style="background-color: #87CEFA"><td colspan="5" style=" font-weight: bold; text-align: center;">InDirect Follow Up</td></tr>'
    table_2 += '<tr style="background-color: #87CEFA"><td style="width: 20%; font-weight: bold;">Customer</td><td style="width: 20%; font-weight: bold;">Closure Status</td><td style="width: 10%; font-weight: bold;">Closure Count</td><td style="width: 20%; font-weight: bold;">DPR Status</td><td style="width: 10%; font-weight: bold;">DPR Count</td></tr>'

    for customer in customers_indirect:
        closure_statuses = {}
        dpr_statuses = {}

        for closure in closure_data_by_customer_indirect.get(customer, []):
            closure_statuses.setdefault(closure.status, []).append(closure.name)

        for dpr in dpr_data_by_customer_indirect.get(customer, []):
            dpr_statuses.setdefault(dpr.status, []).append(dpr.name)

        closure_items = list(closure_statuses.items())
        dpr_items = list(dpr_statuses.items())
        max_rows = max(len(closure_items), len(dpr_items), 1)

        for i in range(max_rows):
            closure_row = closure_items[i] if i < len(closure_items) else ("", [])
            dpr_row = dpr_items[i] if i < len(dpr_items) else ("", [])

            closure_status, closure_ids = closure_row
            dpr_status, dpr_ids = dpr_row

            table_2 += (
                f"<tr>"
                f"<td>{customer if i == 0 else ''}</td>"
                f"<td>{closure_status}</td><td>{len(closure_ids)}</td>"
                f"<td>{dpr_status}</td><td>{len(dpr_ids)}</td>"
                f"</tr>"
            )

    table_2 += '</table>'

    total_count_indirect = sum(len(ids) for statuses in closure_data_by_customer_indirect.values() for ids in statuses)


    
    
    
    #BDM2
    
    # BDM Follow Up
    closures_bdm = frappe.get_all("Closure", {
        "last_updated_on": action_date,
        "status": ("In", ['Visa', 'Client Offer Letter', 'Ticket'])
    }, ["customer", "status", "name"])

    dnd_dpr_records_bdm = frappe.db.get_all("DND DPR Records", {
        "dpr_date": action_date,
        "follow_up": "BDM Follow Up"
    }, ["customer", "status", "name"])

    # Organize data
    closure_data_by_customer_bdm = {}
    dpr_data_by_customer_bdm = {}

    for c in closures_bdm:
        closure_data_by_customer_bdm.setdefault(c.customer, []).append(c)

    for d in dnd_dpr_records_bdm:
        dpr_data_by_customer_bdm.setdefault(d.customer, []).append(d)

    customers_bdm = set(closure_data_by_customer_bdm.keys()).union(dpr_data_by_customer_bdm.keys())

    table_3 = '<table text-align="center" border="1" width="100%" style="border-collapse: collapse;text-align: center;">'
    table_3 += '<tr style="background-color: #87CEFA"><td colspan="5" style=" font-weight: bold; text-align: center;">BDM Follow Up</td></tr>'
    table_3 += '<tr style="background-color: #87CEFA"><td style="width: 20%; font-weight: bold;">Customer</td><td style="width: 20%; font-weight: bold;">Closure Status</td><td style="width: 10%; font-weight: bold;">Closure Count</td><td style="width: 20%; font-weight: bold;">DPR Status</td><td style="width: 10%; font-weight: bold;">DPR Count</td></tr>'

    for customer in customers_bdm:
        closure_statuses = {}
        dpr_statuses = {}

        for closure in closure_data_by_customer_bdm.get(customer, []):
            closure_statuses.setdefault(closure.status, []).append(closure.name)

        for dpr in dpr_data_by_customer_bdm.get(customer, []):
            dpr_statuses.setdefault(dpr.status, []).append(dpr.name)

        closure_items = list(closure_statuses.items())
        dpr_items = list(dpr_statuses.items())
        max_rows = max(len(closure_items), len(dpr_items), 1)

        for i in range(max_rows):
            closure_row = closure_items[i] if i < len(closure_items) else ("", [])
            dpr_row = dpr_items[i] if i < len(dpr_items) else ("", [])

            closure_status, closure_ids = closure_row
            dpr_status, dpr_ids = dpr_row

            table_3 += (
                f"<tr>"
                f"<td>{customer if i == 0 else ''}</td>"
                f"<td>{closure_status}</td><td>{len(closure_ids)}</td>"
                f"<td>{dpr_status}</td><td>{len(dpr_ids)}</td>"
                f"</tr>"
            )

    table_3 += '</table>'

    total_count_bdm = sum(len(ids) for statuses in closure_data_by_customer_bdm.values() for ids in statuses)
    return table, total_count ,table_2, total_count_indirect , table_3, total_count_bdm

@frappe.whitelist()
def create_schedule_job_dsr():
    job = frappe.db.exists('Scheduled Job Type', 'send_closure_report_with_table_dsr')
    if not job:
        exp = frappe.new_doc("Scheduled Job Type")
        exp.update({
            "method": 'checkpro.custom.send_closure_report_with_table_dsr',
            "frequency": 'Cron',
            "cron_format": "00 21 * * *"
        })
        exp.save(ignore_permissions=True)

@frappe.whitelist()
def update_check_age():
    list = ["Education Checks"]

    age=0
    tat_var=0
    tat_mon=''
    tat_sts=''
    for i in list:
        doc=frappe.db.get_list(i,{"name":"Education Checks-22528"},["name","check_creation_date","workflow_state",'package_tat','insufficiency_days'])
        for j in doc:
            if j.workflow_state not in ('Report Completed', '', 'Drop', 'Dropped', 'Not Applicable'):
                if j.check_creation_date:
                    date=(date_diff(nowdate(),j.check_creation_date))+1
                    sql_query = f"""
                        SELECT COUNT(*)
                        FROM `tabHoliday`
                        WHERE parent = 'TEAMPRO 2023 - Checkpro'
                        AND holiday_date BETWEEN '{j.check_creation_date}' AND '{nowdate()}'
                    """
                    count = frappe.db.sql(sql_query, as_list=True)[0][0]
                    print(count)
                    print(date)
                    if count==0:
                        age=date-j.insufficiency_days
                    else:
                        age = date-(count+j.insufficiency_days)
                    tat_var=int(j.package_tat)-age
                    if tat_var>0:
                        tat_mon='In TAT'
                    else:
                        tat_mon='Out TAT'
                    if age<(0.4*int(j.package_tat)):
                        tat_sts='Regular'
                    elif age<(0.65*int(j.package_tat)):
                        tat_sts='Critical'
                    else:
                        tat_sts='Most Critical'
                    print(j.name)
                    print(age)
                    # frappe.db.set_value(i,j.name,"actual_tat",age)
                    # frappe.db.set_value(i,j.name,"holidays",count)
                    # frappe.db.set_value(i,j.name,"tat_variation",tat_var)
                    # frappe.db.set_value(i,j.name,"tat_monitor",tat_mon)
                    # frappe.db.set_value(i,j.name,"custom_tat_status",tat_sts)

# Daily cron update
@frappe.whitelist()
def update_tat_completion_date_during_update(name):
    
    doc=frappe.get_doc("Case",name)
    if doc.insufficiency_closed:
        from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
        holiday_list_name = 'TEAMPRO 2023 - Checkpro'
        start_date = doc.insufficiency_closed
        working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
        current_date = start_date
        holiday = []
        while working_days > 0:
            if not is_holiday(holiday_list_name, current_date):
                holiday.append(current_date)
                working_days -= 1
            current_date = add_days(current_date, 1)
        sql_query = f"""
            SELECT COUNT(*) 
            FROM `tabHoliday` 
            WHERE parent = 'TEAMPRO 2023 - Checkpro' 
            AND holiday_date BETWEEN '{doc.insufficiency_closed}' AND '{holiday[-1]}'
        """
        count = frappe.db.sql(sql_query, as_list=True)[0][0]
        frappe.db.set_value("Case",doc.name,"end_date",holiday[-1])
        frappe.db.set_value("Case",doc.name,"holidays",count)

